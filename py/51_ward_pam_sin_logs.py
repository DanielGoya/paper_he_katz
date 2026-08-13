"""Bloque K.8 — Ward sin ajustes contra PAM ponderado con balanza normalizada
y SIN logaritmos, k = 5, en los cuatro paises.

POR QUE EXISTE (DG, 2026-08-13). La hoja `Chile PAM+ajustes` de `T_K14` es
economicamente ilegible: el cobre queda con educacion publica, suministro de agua
y administracion publica; hay un segundo grupo de 35 actividades con el comercio
y toda la manufactura; y la fruticultura queda sola. La descomposicion en
escalera —una correccion por vez, mismo universo y mismo k— identifico al
culpable, y no es el algoritmo ni el universo:

  1. LOS LOGARITMOS meten al cobre con educacion. Sus puntajes z en 2003 pasan
     de 6,60 / 6,15 / 7,10 (productividad, remuneracion, balanza) a
     5,35 / 3,91 / 1,75: el log y la normalizacion borran justo las tres
     dimensiones en que el cobre se distingue. Y como con 0,29 % del empleo PAM
     ponderado nunca le va a dar medoide propio, queda asignado POR RESIDUO al
     medoide grande mas cercano. La compania de educacion y agua no es afinidad:
     es lo que sobra.
  2. LA PONDERACION deja solas a fruticultura y construccion, y en PAM pega mas
     fuerte que en Ward: los cinco medoides de 2003 son literalmente las cinco
     actividades de mas empleo (Comercio 14,6 %, Construccion 15,5 %,
     Fruticultura 5,7 %, Servicios empresariales 10,5 %, Agricultura 3,3 %),
     mientras que sin ponderar son actividades diminutas (conservas 0,3 %). PAM
     planta el centro sobre una OBSERVACION real, asi que ponderar decide donde
     se planta —sobre la masa—; en Ward el peso solo altera el coste de fusion.
  3. NO es el universo ni el k: el cobre ya caia con educacion en T_K13 (COU
     completa), y la patologia es identica a k = 4 y k = 6.

LO QUE ESTE SCRIPT CAMBIA, entonces, es UNA cosa respecto de `T_K14`: **se sacan
los logaritmos**. Queda PAM Manhattan ponderado por empleo con balanza
normalizada. Sobre el universo UAM + cobre y k = 5 eso devuelve el enclave en los
dos anios chilenos —6 actividades con 7,7 % del empleo, 50,9 % de las
exportaciones y masa salarial de 22,2 % en 2003; 16 con 5,6 %, 82,2 % y 18,0 % en
2023—, con la fruticultura junto al cobre, que es donde la economia la pone.

MATIZ SOBRE UNA REGLA YA REGISTRADA. Estaba anotado que logaritmar SIN ponderar
empeora la invariancia a la desagregacion (ARI -0,086 en 2003), de donde salio
«las dos correcciones van juntas». El reverso vale solo para PAM: logaritmar Y
ponderar SOBRE-CORRIGE, porque las dos actuan sobre lo mismo —la influencia de
los extremos— y entre las dos dejan al cobre sin dimension propia y sin poder
anclar un grupo. Con Ward no se ve, porque el centro es una media y no una
observacion. La invariancia de esta especificacion la sostiene la balanza
normalizada, que es un ratio acotado, mas la ponderacion; no los logaritmos.

LAS DOS ESPECIFICACIONES:

    Ward         sobre distancia euclidiana, las siete variables de la UAM en
                 niveles, sin ponderar, balanza comercial en niveles. Es la
                 especificacion publicada, sin ningun ajuste, con el corte comun
                 k = 5. Va como contraste, igual que en T_K14.
    PAM+ajustes  PAM con distancia de Manhattan, PONDERADO por empleo y con
                 balanza normalizada (X-M)/(X+M). SIN logaritmos.

EL UNIVERSO, igual que en T_K14: Chile, las actividades que la UAM clasifica MAS
el cobre reintegrado (63 en 2003, 102 en 2023); Brasil, Mexico y Argentina, la
COU completa, porque ahi no hay lista de la UAM que restringir.

SALIDA. `T_K15`, un libro con nueve hojas: lectura, y dos por pais, llamadas
«<pais> Ward» y «<pais> PAM+ajustes».

DEPENDE DE. `50_ward_pam_universo_uam.py`, de donde salen los cargadores del
universo y el armado de bloques; `planilla_clusters.py` y `26_pam_medoides.py`.
"""

from __future__ import annotations

import importlib.util
import sys
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402
import planilla_clusters as pc  # noqa: E402


def _cargar_modulo(archivo: str, nombre: str):
    """Los modulos de `py/` empiezan con digito y no se pueden importar por nombre."""
    spec = importlib.util.spec_from_file_location(
        nombre, Path(__file__).resolve().parent / archivo)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


pam_mod = _cargar_modulo("26_pam_medoides.py", "pam_medoides")
# De aca salen `cargar()` (universo UAM+cobre para Chile, COU completa para el
# resto), `bloque()` y `extremas()`: son identicos y no se duplican.
k14 = _cargar_modulo("50_ward_pam_universo_uam.py", "ward_pam_universo_uam")
uam = pc.uam

HOY = date.today().strftime("%Y.%m.%d")
SALIDA = rutas.OUT / f"{HOY} T_K15 Ward contra PAM ponderado sin logs k=5.xlsx"

K = 5
PAISES = k14.PAISES

SPEC_WARD = dict(log=False, peso=False, balanza="nivel", sinfbcf=False, comercio=False)
# La UNICA diferencia con T_K14: log=False.
SPEC_PAM = dict(log=False, peso=True, balanza="norm", sinfbcf=False, comercio=False)

NOTAS = {
    "Ward": [
        ("Ward sobre distancia euclidiana, las siete variables de la UAM en niveles, sin "
         "ponderar y con la balanza comercial en niveles: la especificacion publicada, SIN "
         f"ningun ajuste. Lo unico que se le impone es el corte comun k = {K}, para que las dos "
         "hojas del pais sean comparables. Va como contraste.", "000000"),
        ("Conglomerados numerados de mayor a menor empleo. Al pie, para cada anio, las medias "
         "por conglomerado en niveles y en puntajes z.", "000000"),
    ],
    "PAM+ajustes": [
        (f"PAM (k-medoides) con distancia de Manhattan, PONDERADO por empleo y con balanza "
         f"normalizada (X-M)/(X+M). k = {K}. SIN LOGARITMOS: es la unica diferencia con T_K14.",
         "000000"),
        ("Los logaritmos se sacaron por diagnostico, no por gusto: con ellos el cobre pierde "
         "las tres dimensiones en que se distingue (z de productividad 6,60 -> 5,35, "
         "remuneracion 6,15 -> 3,91, balanza 7,10 -> 1,75) y, como con 0,29 % del empleo PAM "
         "ponderado nunca le da medoide propio, quedaba asignado por residuo al grupo de "
         "educacion y administracion publica. Sin logaritmos vuelve al enclave exportador.",
         "000000"),
        ("La invariancia a la desagregacion la sostienen la balanza normalizada —un ratio "
         "acotado— y la ponderacion, no los logaritmos.", "000000"),
        ("PAM trabaja con MEDOIDES, no con medias: las tablas del pie van como descripcion. "
         "Conglomerados numerados de mayor a menor empleo.", "000000"),
    ],
}


def clasificar(d: pd.DataFrame, metodo: str, k: int) -> np.ndarray:
    """Devuelve la particion, renumerada de mayor a menor empleo."""
    if metodo == "Ward":
        cl = pc.clasificar(d, k, **SPEC_WARD)
    elif metodo == "PAM+ajustes":
        z, pesos = pc.matriz_z(d, **SPEC_PAM)
        D = pam_mod.matriz_distancias(z, "manhattan")
        _, et = pam_mod.pam(D, k, pesos)
        cl = et + 1
    else:
        raise ValueError(metodo)
    return pc.renumerar(cl, d["empleo"].to_numpy())


def main() -> None:
    rutas.preparar_directorios()
    pam_mod.verificar()

    datos, cobres = k14.cargar()
    print(f"Chile: {(datos[(datos.pais == 'Chile') & (datos.anio == 2003)]).shape[0]} actividades "
          f"en 2003 y {(datos[(datos.pais == 'Chile') & (datos.anio == 2023)]).shape[0]} en 2023 "
          "— clasificadas por la UAM + cobre reintegrado\n")

    wb = Workbook()
    wb.remove(wb.active)
    resumen = []

    for pais, anios in PAISES:
        for metodo in ("Ward", "PAM+ajustes"):
            bloques, tablas = [], []
            spec = SPEC_WARD if metodo == "Ward" else SPEC_PAM

            for anio in anios:
                d = datos[(datos.pais == pais) & (datos.anio == anio)].reset_index(drop=True)
                if d.empty:
                    print(f"  AVISO: no hay datos de {pais} {anio}")
                    continue

                if pais == "Chile":
                    marcadas = {cobres[anio]}
                else:
                    ext = k14.extremas(d)
                    i_ext, i_exp = int(np.argmax(ext)), int(np.argmax(d["expo"].to_numpy()))
                    marcadas = {d.loc[i_ext, "codigo"], d.loc[i_exp, "codigo"]}

                cl = clasificar(d, metodo, K)
                titulo = f"{pais} {anio} ({len(d)} actividades, k = {K})"
                bloques.append((k14.bloque(d, pais, anio, cl, marcadas), titulo))

                t_niv = pc.medias_en_niveles(d, cl)
                tablas.append([
                    (t_niv, f"{anio}: medias por conglomerado, en niveles (ratios = cociente "
                            "de las sumas; indices con la economia = 100)"),
                    (pc.medias_en_z(d, cl, **spec),
                     f"{anio}: medias en puntajes z, con las variables tal como entran a ESTA "
                     "especificacion"),
                ])

                g = int(t_niv["% expo"].idxmax())
                fila = {
                    "pais": pais, "anio": anio, "metodo": metodo,
                    "n actividades": len(d),
                    "tamanos": " / ".join(str(x) for x in t_niv["n"].astype(int)),
                    "% empleo": " / ".join(f"{x:.1f}" for x in t_niv["% empleo"]),
                    "productividad": " / ".join(f"{x:.0f}" for x in t_niv["productividad (=100)"]),
                    "empleo del mayor (%)": float(t_niv["% empleo"].max()),
                    "grupos <1% empleo": int((t_niv["% empleo"] < 1.0).sum()),
                    "grupos de 1 actividad": int((t_niv["n"] == 1).sum()),
                    "grupo con mas expo: actividades": int(t_niv.loc[g, "n"]),
                    "grupo con mas expo: % empleo": float(t_niv.loc[g, "% empleo"]),
                    "grupo con mas expo: % expo": float(t_niv.loc[g, "% expo"]),
                    "grupo con mas expo: masa salarial/VA (%)":
                        100 * float(t_niv.loc[g, "masa salarial / VA"]),
                    "eta2 log productividad": pam_mod.eta2_productividad(
                        d, cl, d["empleo"].to_numpy(dtype=float)),
                }
                # En Chile se sabe cual es el cobre: se reporta con quien queda.
                if pais == "Chile":
                    ic = int(np.where(d["codigo"].to_numpy() == cobres[anio])[0][0])
                    sub = d[cl == cl[ic]]
                    fila["cobre: actividades en su grupo"] = int(len(sub))
                    fila["cobre: % expo de su grupo"] = 100 * sub["expo"].sum() / d["expo"].sum()
                    fila["cobre: masa salarial/VA de su grupo (%)"] = (
                        100 * sub["remunera"].sum() / sub["valor_agrega"].sum())
                resumen.append(fila)

            notas = NOTAS[metodo] + [
                k14.NOTA_UNIVERSO["Chile" if pais == "Chile" else "otros"]]
            pc.escribir_hoja(wb, f"{pais} {metodo}", notas, bloques,
                             con_marca=True, tablas=tablas)
            print(f"  {pais:<10} {metodo:<12} escrita ({len(bloques)} periodo/s)")

    T = pd.DataFrame(resumen)
    ari = []
    for pais, anios in PAISES:
        for anio in anios:
            d = datos[(datos.pais == pais) & (datos.anio == anio)].reset_index(drop=True)
            if d.empty:
                continue
            ari.append({
                "pais": pais, "anio": anio,
                "ARI Ward vs PAM+ajustes": uam.rand_ajustado(
                    clasificar(d, "Ward", K), clasificar(d, "PAM+ajustes", K)),
            })
    T = T.merge(pd.DataFrame(ari), on=["pais", "anio"], how="left")

    ws = wb.create_sheet("0 Lectura", 0)
    lineas = [
        (f"T_K15 — Ward sin ajustes contra PAM ponderado SIN logaritmos, k = {K}", True, 13),
        ("Es T_K14 con UNA diferencia: se sacan los logaritmos de la especificacion PAM. Queda "
         "PAM Manhattan ponderado por empleo con balanza normalizada. Todo lo demas —universo, "
         "corte, hojas, medias al pie— es identico.", False, 9),
        ("El motivo es un diagnostico, no una preferencia. Con logaritmos, el cobre pierde las "
         "tres dimensiones en que se distingue (z de productividad 6,60 -> 5,35, remuneracion "
         "6,15 -> 3,91, balanza 7,10 -> 1,75) y, como con 0,29 % del empleo PAM ponderado nunca "
         "le da medoide propio, quedaba asignado por residuo al grupo de educacion publica, "
         "administracion publica y suministro de agua. Ademas la ponderacion plantaba los cinco "
         "medoides sobre las cinco actividades de mas empleo, dejando solas a Construccion y "
         "Fruticultura. Sin logaritmos, el enclave exportador vuelve.", False, 9),
        ("Dos hojas por pais: «<pais> Ward» es la especificacion publicada sin ningun ajuste, "
         "como contraste; «<pais> PAM+ajustes» es la contrapropuesta. Los anios van lado a "
         "lado y las medias por conglomerado al pie de cada uno.", False, 9),
        ("Universo: Chile, las actividades que la UAM clasifica MAS el cobre reintegrado (63 en "
         "2003, 102 en 2023). Brasil, Mexico y Argentina, la COU completa, porque ahi no hay "
         "lista de la UAM que restringir.", False, 9),
        (f"Generada el {HOY} por py/51_ward_pam_sin_logs.py.", False, 9),
        ("", False, 9),
        ("Resumen (una fila por pais, anio y metodo):", True, 10),
    ]
    for i, (texto, negrita, tam) in enumerate(lineas):
        ws.cell(1 + i, 1, texto).font = Font(bold=negrita, italic=not negrita, size=tam)
    f0 = len(lineas) + 2
    for j, h in enumerate(T.columns):
        ws.cell(f0, 1 + j, h).font = Font(bold=True, size=9)
    for i, fila in enumerate(T.itertuples(index=False)):
        for j, v in enumerate(fila):
            c = ws.cell(f0 + 1 + i, 1 + j,
                        None if (isinstance(v, float) and np.isnan(v)) else v)
            if isinstance(v, float):
                c.number_format = "0.000"
    ws.column_dimensions["A"].width = 120

    wb.save(SALIDA)
    print(f"\nescrito: {SALIDA}\n")
    cols = [c for c in T.columns if not c.startswith("cobre:")]
    print(T[cols].round(3).to_string(index=False))
    if "cobre: actividades en su grupo" in T.columns:
        print("\n== Donde queda el cobre (Chile) ==")
        print(T[T.pais == "Chile"][
            ["anio", "metodo", "cobre: actividades en su grupo", "cobre: % expo de su grupo",
             "cobre: masa salarial/VA de su grupo (%)"]].round(1).to_string(index=False))


if __name__ == "__main__":
    main()

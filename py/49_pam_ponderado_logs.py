"""Bloque K.6 — la especificacion de la UAM contra PAM Manhattan ponderado y con
logaritmos, k = 5, en los cuatro paises y en un solo libro.

LA PREGUNTA (DG, 2026-08-13). Las corridas multipais anteriores separaban las
piezas a proposito: `42_cluster_multipais.py` lleva afuera el paquete completo de
mejoras con Ward, y `47_pam_multipais.py` (T_K11) prueba que hace PAM POR SI
SOLO, sin ponderar ni logaritmar. Falta la combinacion que el capitulo defiende
—las dos correcciones de forma JUNTO CON el algoritmo robusto— y falta verla al
lado de lo publicado, hoja contra hoja, en un solo archivo.

QUE SE COMPARA, dos hojas por pais:

    UAM    Ward sobre distancia euclidiana, las siete variables en niveles, sin
           ponderar, balanza comercial en niveles. Es la especificacion
           publicada; lo unico que se le impone es el corte comun k = 5.
    PAM    PAM con distancia de Manhattan, PONDERADO por empleo, con logaritmos
           en productividad y remuneracion media y balanza normalizada
           (X-M)/(X+M). Es la especificacion 7 de T_I13, o sea la que combina la
           estrategia de robustez del proyecto con la ortodoxa.

El mismo k en las dos: la comparacion es de METODO, y cortar en distinto numero
de grupos la volveria ilegible. k = 5 es lo pedido; no se recorre el corte, y
por eso las conclusiones sobre el numero de estratos siguen viniendo del salto de
costes (bloque I.5) y del eta cuadrado (bloque I.6), no de aca.

POR QUE PONDERAR Y LOGARITMAR VAN JUNTOS. Es un resultado propio, registrado en
MEMORY.md: logaritmar SIN ponderar empeora la invariancia a la desagregacion
(ARI 0,780 en 2023 y -0,086 en 2003, contra 0,949 y 0,858 en niveles), porque el
logaritmo comprime a los grandes y expande a los chicos. Con pesos, la invariancia
vuelve a ser exacta. De ahi que esta hoja no ofrezca la variante «logs sin peso».

EL UNIVERSO. La COU COMPLETA de cada pais, sin excluir nada, igual que en T_K11:
afuera de Chile no hay subconjunto de la UAM que respetar, porque su criterio de
exclusion no esta documentado (bloque K.4). Chile va con sus 73 y 111
actividades, del insumo de la UAM, que es la unica fuente chilena con la COU
entera Y empleo.

LAS MEDIAS AL PIE. Cada hoja lleva, debajo de las membresias y para CADA ano, las
medias por conglomerado en dos versiones: en niveles —ratios calculados como
cociente de las sumas, con productividad y remuneracion media indexadas a la
economia = 100— y en puntajes z, que son los centroides con que trabajo el propio
algoritmo. Sin ellas la hoja dice quien esta con quien pero no por que.

SALIDA. `T_K13`, un libro con nueve hojas: lectura, y dos por pais (Chile,
Brasil, Mexico, Argentina).

DEPENDE DE. `40_uam_metodo.py`, `41_cou_multipais.py`, `planilla_clusters.py` y
`26_pam_medoides.py`.
"""

from __future__ import annotations

import importlib.util
import sys
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402
import planilla_clusters as pc  # noqa: E402


def _cargar_modulo(archivo: str, nombre: str):
    """Los modulos de `py/` empiezan con digito y no se pueden importar por nombre."""
    spec = importlib.util.spec_from_file_location(
        nombre, Path(__file__).resolve().parent / archivo)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


pam_mod = _cargar_modulo("26_pam_medoides.py", "pam_medoides")
uam = pc.uam

HOY = date.today().strftime("%Y.%m.%d")
SALIDA = rutas.OUT / f"{HOY} T_K13 UAM contra PAM ponderado con logs k=5.xlsx"

K = 5

# Un pais por par de hojas; los anios son los unicos con COU desagregada Y empleo.
PAISES = [
    ("Chile", [2003, 2023]),
    ("Brasil", [2003, 2021]),
    ("Mexico", [2008, 2018]),
    ("Argentina", [1997]),
]

# Las dos especificaciones que se enfrentan, en el vocabulario de `planilla_clusters`.
SPEC_UAM = dict(log=False, peso=False, balanza="nivel", sinfbcf=False, comercio=False)
SPEC_PAM = dict(log=True, peso=True, balanza="norm", sinfbcf=False, comercio=False)

NOTAS = {
    "UAM": [
        ("Ward sobre distancia euclidiana, las siete variables de la UAM en niveles, sin "
         "ponderar y con la balanza comercial en niveles. Es la especificacion publicada; lo "
         f"unico que se le impone es el corte comun k = {K}, para que las dos hojas del pais "
         "sean comparables.", "000000"),
        ("Conglomerados numerados de mayor a menor empleo. Al pie, para cada anio, las medias "
         "por conglomerado en niveles y en puntajes z.", "000000"),
    ],
    "PAM": [
        (f"PAM (k-medoides) con distancia de Manhattan, PONDERADO por empleo, con logaritmos en "
         f"productividad y remuneracion media y balanza normalizada (X-M)/(X+M). k = {K}. Es la "
         "especificacion 7 de T_I13: combina la estrategia de robustez del proyecto con la "
         "ortodoxa.", "000000"),
        ("Ponderar y logaritmar van juntos a proposito: logaritmar SIN ponderar empeora la "
         "invariancia a la desagregacion (ARI 0,780 en 2023 y -0,086 en 2003, contra 0,949 y "
         "0,858 en niveles); con pesos la invariancia es exacta.", "000000"),
        ("PAM trabaja con MEDOIDES, no con medias: las tablas del pie van como descripcion. "
         "Conglomerados numerados de mayor a menor empleo.", "000000"),
    ],
}


# ==========================================================================
# Carga: la COU completa de cada pais
# ==========================================================================

def cargar_chile() -> pd.DataFrame:
    """Chile con las 73 y 111 actividades, del insumo de la UAM.

    `cou_multipais.csv` trae a Chile ya recortado al subconjunto que la UAM
    clasifica (62 y 101); para tener la COU completa —y con ella el cobre— hay
    que ir al insumo, la unica fuente chilena con todas las actividades Y empleo.
    """
    ins = pd.read_excel(rutas.exigir(rutas.D_UAM_INSUMO, "el insumo de la UAM"))
    return pd.DataFrame({
        "pais": "Chile", "anio": ins["anio"], "periodo": ins["periodo"],
        "codigo": ins["codigo_3dig"].astype(str), "desc": ins["desc"],
        "vbp": ins["vbp"], "valor_agrega": ins["valor_agrega"],
        "remunera": ins["remunera"], "empleo": ins["empleo"],
        "fbcf": ins["fbcf"], "expo": ins["expo"], "impo": ins["impo"],
    })


def cargar() -> pd.DataFrame:
    d = pd.read_csv(rutas.exigir(rutas.INTER / "cou_multipais.csv", "la extraccion multipais"))
    # Una actividad sin remuneraciones declaradas se conserva con cero, que es lo
    # que dice la fuente; en NaN caeria despues en la matriz de distancias.
    d["remunera"] = d["remunera"].fillna(0.0)
    d["codigo"] = d["codigo"].astype(str)
    return pd.concat([cargar_chile(), d[d.pais != "Chile"]], ignore_index=True)


# ==========================================================================
# Particion
# ==========================================================================

def clasificar(d: pd.DataFrame, metodo: str, k: int) -> np.ndarray:
    """Devuelve la particion, renumerada de mayor a menor empleo."""
    if metodo == "UAM":
        cl = pc.clasificar(d, k, **SPEC_UAM)
    elif metodo == "PAM":
        z, pesos = pc.matriz_z(d, **SPEC_PAM)
        D = pam_mod.matriz_distancias(z, "manhattan")
        _, et = pam_mod.pam(D, k, pesos)
        cl = et + 1
    else:
        raise ValueError(metodo)
    return pc.renumerar(cl, d["empleo"].to_numpy())


def bloque(d: pd.DataFrame, pais: str, anio: int, cl: np.ndarray,
           marcadas: set[str]) -> pd.DataFrame:
    """Las cinco columnas del formato UAM, ordenadas por conglomerado."""
    b = pd.DataFrame({
        "pais": pais.upper(),
        "periodo": anio,
        "cluster": cl,
        "codigo_3dig": d["codigo"].to_numpy(),
        "desc": d["desc"].to_numpy(),
    })
    b["marcada"] = b["codigo_3dig"].isin(marcadas)
    b["_ord"] = pd.to_numeric(b["codigo_3dig"], errors="coerce")
    return b.sort_values(["cluster", "_ord"]).drop(columns="_ord").reset_index(drop=True)


def extremas(d: pd.DataFrame) -> np.ndarray:
    """max |z| de cada actividad en las siete variables, sobre la especificacion
    de la UAM. Sirve para MARCAR el analogo del cobre de cada pais, no para sacar
    nada: es el sustituto del criterio de exclusion del bloque K.4."""
    r = uam.construir_variables(d, balanza="nivel")
    return np.abs(uam.estandarizar(r, uam.VARS_UAM)).max(axis=1)


def main() -> None:
    rutas.preparar_directorios()
    pam_mod.verificar()

    datos = cargar()
    wb = Workbook()
    wb.remove(wb.active)
    resumen = []

    for pais, anios in PAISES:
        for metodo in ("UAM", "PAM"):
            bloques, tablas = [], []
            spec = SPEC_UAM if metodo == "UAM" else SPEC_PAM

            for anio in anios:
                d = datos[(datos.pais == pais) & (datos.anio == anio)].reset_index(drop=True)
                if d.empty:
                    print(f"  AVISO: no hay datos de {pais} {anio}")
                    continue

                # Las dos actividades que hacen las veces del cobre en cada pais.
                ext = extremas(d)
                i_ext, i_exp = int(np.argmax(ext)), int(np.argmax(d["expo"].to_numpy()))
                marcadas = {d.loc[i_ext, "codigo"], d.loc[i_exp, "codigo"]}

                cl = clasificar(d, metodo, K)
                titulo = f"{pais} {anio} ({len(d)} actividades, k = {K})"
                bloques.append((bloque(d, pais, anio, cl, marcadas), titulo))

                t_niv = pc.medias_en_niveles(d, cl)
                tablas.append([
                    (t_niv, f"{anio}: medias por conglomerado, en niveles (ratios = cociente "
                            "de las sumas; indices con la economia = 100)"),
                    (pc.medias_en_z(d, cl, **spec),
                     f"{anio}: medias en puntajes z, con las variables tal como entran a ESTA "
                     "especificacion"),
                ])

                # El conglomerado que concentra mas exportaciones. NO se lo llama
                # «enclave» en la salida: con la especificacion ponderada suele
                # ser el grupo grande, y el enclave del capitulo es otra cosa
                # —pocas actividades, poco empleo y masa salarial de ~20 % del
                # VA—. La etiqueta la pone quien lee la fila, no el codigo.
                g = int(t_niv["% expo"].idxmax())
                resumen.append({
                    "pais": pais, "anio": anio, "metodo": metodo,
                    "n actividades": len(d),
                    "tamanos": " / ".join(str(x) for x in t_niv["n"].astype(int)),
                    "% empleo": " / ".join(f"{x:.1f}" for x in t_niv["% empleo"]),
                    "productividad": " / ".join(f"{x:.0f}" for x in t_niv["productividad (=100)"]),
                    "empleo del mayor (%)": float(t_niv["% empleo"].max()),
                    "grupos <1% empleo": int((t_niv["% empleo"] < 1.0).sum()),
                    "grupo con mas expo: actividades": int(t_niv.loc[g, "n"]),
                    "grupo con mas expo: % empleo": float(t_niv.loc[g, "% empleo"]),
                    "grupo con mas expo: % expo": float(t_niv.loc[g, "% expo"]),
                    "grupo con mas expo: masa salarial/VA (%)":
                        100 * float(t_niv.loc[g, "masa salarial / VA"]),
                    "eta2 log productividad": pam_mod.eta2_productividad(
                        d, cl, d["empleo"].to_numpy(dtype=float)),
                })

            notas = NOTAS[metodo] + [
                ("Universo: la COU COMPLETA, sin excluir ninguna actividad. En rojo y cursiva, "
                 "la actividad mas extrema (mayor max |z| en las siete variables) y el mayor "
                 "exportador: son el analogo del cobre chileno.", pc.ROJO),
            ]
            pc.escribir_hoja(wb, f"{pais} {metodo}", notas, bloques,
                             con_marca=True, tablas=tablas)
            print(f"  {pais:<10} {metodo}  escrita ({len(bloques)} periodo/s)")

    # ---- ARI entre los dos metodos, pais por pais
    T = pd.DataFrame(resumen)
    ari = []
    for pais, anios in PAISES:
        for anio in anios:
            d = datos[(datos.pais == pais) & (datos.anio == anio)].reset_index(drop=True)
            if d.empty:
                continue
            ari.append({
                "pais": pais, "anio": anio,
                "ARI UAM vs PAM ponderado con logs": uam.rand_ajustado(
                    clasificar(d, "UAM", K), clasificar(d, "PAM", K)),
            })
    T = T.merge(pd.DataFrame(ari), on=["pais", "anio"], how="left")

    ws = wb.create_sheet("0 Lectura", 0)
    lineas = [
        (f"T_K13 — la especificacion de la UAM contra PAM Manhattan ponderado y con "
         f"logaritmos, k = {K}", True, 13),
        ("Dos hojas por pais, con los dos anios lado a lado y las medias por conglomerado al "
         "pie de cada uno. La hoja «UAM» es Ward euclidiano con las siete variables en "
         "niveles, sin ponderar y con la balanza en niveles: la especificacion publicada. La "
         "hoja «PAM» es PAM Manhattan ponderado por empleo, con logaritmos en productividad y "
         "remuneracion media y balanza normalizada.", False, 9),
        (f"El corte es el mismo en las dos, k = {K}: la comparacion es de METODO. No se recorre "
         "el numero de grupos, asi que el criterio para elegir k sigue viniendo del salto de "
         "costes (bloque I.5) y del eta cuadrado (bloque I.6), no de este libro.", False, 9),
        ("Ponderar y logaritmar van juntos a proposito: logaritmar sin ponderar empeora la "
         "invariancia a la desagregacion; con pesos es exacta. Es un resultado propio del "
         "proyecto, no una cita.", False, 9),
        ("Universo: la COU completa de cada pais, sin excluir nada. Chile va con sus 73 y 111 "
         "actividades, del insumo de la UAM; los otros tres, de la extraccion multipais. "
         "Argentina tiene un solo anio: 1997 es el unico con remuneraciones y empleo.", False, 9),
        ("Empleo: Brasil, ocupaciones de la TRU del IBGE; Mexico, puestos de trabajo de la MIP "
         "del INEGI; Argentina, MIP del INDEC. Los anios no son elegibles: son los ultimos con "
         "COU desagregada y empleo.", False, 9),
        (f"Generada el {HOY} por py/49_pam_ponderado_logs.py.", False, 9),
        ("", False, 9),
        ("Resumen (una fila por pais, anio y metodo):", True, 10),
    ]
    for i, (texto, negrita, tam) in enumerate(lineas):
        ws.cell(1 + i, 1, texto).font = Font(bold=negrita, italic=not negrita, size=tam)
    f0 = len(lineas) + 2
    for j, h in enumerate(T.columns):
        ws.cell(f0, 1 + j, h).font = Font(bold=True, size=9)
    for i, fila in enumerate(T.itertuples(index=False)):
        for j, v in enumerate(fila):
            c = ws.cell(f0 + 1 + i, 1 + j,
                        None if (isinstance(v, float) and np.isnan(v)) else v)
            if isinstance(v, float):
                c.number_format = "0.000"
    ws.column_dimensions["A"].width = 120

    wb.save(SALIDA)
    print(f"\nescrito: {SALIDA}\n")
    print(T.round(3).to_string(index=False))


if __name__ == "__main__":
    main()

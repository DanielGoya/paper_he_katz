"""Piezas comunes de las planillas de clasificacion T_I22, T_I23 y T_I24.

Las tres tienen el mismo formato —el de la hoja `sectores_cluster` de los libros
3a y 3b de la UAM, con los dos periodos lado a lado— y se diferencian solo en el
universo de actividades y en como se elige el numero de conglomerados. Todo lo
que comparten vive aca.

Este archivo NO se ejecuta solo: lo importan 28, 29 y 31.
"""

from __future__ import annotations

import importlib.util
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_spec = importlib.util.spec_from_file_location(
    "uam_metodo", Path(__file__).resolve().parent / "40_uam_metodo.py"
)
uam = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(uam)


# --------------------------------------------------------------------------
# Las cuatro piezas de especificacion, cruzadas de forma independiente
# --------------------------------------------------------------------------

# Se cruzan de forma INDEPENDIENTE: logaritmar no arrastra la balanza
# normalizada. Es lo que permite ver, comparando «log+peso» con
# «log+peso+balanza», cuanto aporta cada correccion por separado. El costo es
# que en las hojas con logaritmos y balanza en nivel esa variable queda cruda al
# lado de las logaritmadas: en niveles es negativa en buena parte de las
# actividades y no admite logaritmo.
ESPECIFICACIONES = [
    ("2 log",                       dict(log=True,  peso=False, balanza="nivel", sinfbcf=False)),
    ("3 peso empleo",               dict(log=False, peso=True,  balanza="nivel", sinfbcf=False)),
    ("4 balanza ajustada",          dict(log=False, peso=False, balanza="norm",  sinfbcf=False)),
    ("5 sin FBCF",                  dict(log=False, peso=False, balanza="nivel", sinfbcf=True)),
    ("6 log+peso",                  dict(log=True,  peso=True,  balanza="nivel", sinfbcf=False)),
    ("7 log+peso+balanza",          dict(log=True,  peso=True,  balanza="norm",  sinfbcf=False)),
    ("8 log+peso+sin FBCF",         dict(log=True,  peso=True,  balanza="nivel", sinfbcf=True)),
    ("9 log+peso+balanza+sin FBCF", dict(log=True,  peso=True,  balanza="norm",  sinfbcf=True)),
]

DESCRIPCION = {
    "2 log":
        "Logaritmos en productividad y remuneracion media, log1p en impo/VBP y expo/VBP. "
        "FBCF por ocupado no admite logaritmo (tiene ceros y negativos) y queda en nivel.",
    "3 peso empleo":
        "Ponderacion por empleo: momentos ponderados en la estandarizacion y masas en el "
        "criterio de Ward, coste (Wa*Wb/(Wa+Wb))*||ca-cb||^2.",
    "4 balanza ajustada":
        "Balanza comercial normalizada (X-M)/(X+M) en vez de X-M en niveles. Es la unica "
        "de las siete que no era un ratio.",
    "5 sin FBCF":
        "Se saca x_capital_trabajo (FBCF por ocupado), que la COU asigna por producto y no "
        "mide intensidad de capital. Quedan seis variables.",
    "6 log+peso":
        "Logaritmos mas ponderacion por empleo. Balanza todavia en niveles.",
    "7 log+peso+balanza":
        "Logaritmos, ponderacion por empleo y balanza normalizada. Es la especificacion "
        "preferida del capitulo.",
    "8 log+peso+sin FBCF":
        "Logaritmos y ponderacion por empleo, sin FBCF por ocupado. Balanza en niveles.",
    "9 log+peso+balanza+sin FBCF":
        "Las cuatro correcciones juntas: logaritmos, ponderacion por empleo, balanza "
        "normalizada y sin FBCF por ocupado. Seis variables.",
}


def variables_de(sinfbcf: bool, comercio: bool = False) -> list[str]:
    """Las variables de una especificacion.

    `comercio=True` usa la reparametrizacion de `uam.VARS_COMERCIO`: seis
    columnas en vez de siete, con apertura y balanza normalizada en lugar de las
    TRES de comercio de la UAM, de las cuales una era funcion exacta de las otras
    dos. `comercio=False` (el default) deja las siete originales, para no alterar
    las planillas T_I22 a T_I25 ya escritas.
    """
    base = uam.VARS_COMERCIO if comercio else uam.VARS_UAM
    return [v for v in base if not (sinfbcf and v == "x_capital_trabajo")]


def matriz_z(d: pd.DataFrame, log: bool, peso: bool, balanza: str,
             sinfbcf: bool, comercio: bool = False,
             log1p_ratios: bool = True) -> tuple[np.ndarray, np.ndarray | None]:
    """Los puntajes z de una especificacion, y sus pesos (o None)."""
    x = uam.construir_variables(d, balanza=balanza)
    variables = variables_de(sinfbcf, comercio)
    if log:
        x = uam.transformar_logs(x, variables, log1p_ratios=log1p_ratios)
    m = x[variables].to_numpy(dtype=float)
    if not np.isfinite(m).all():
        malas = [v for i, v in enumerate(variables) if not np.isfinite(m[:, i]).all()]
        raise ValueError(f"variables con valores no finitos: {malas}")
    pesos = d["empleo"].to_numpy(dtype=float) if peso else None
    return uam.estandarizar(x, variables, pesos), pesos


def enlace(z: np.ndarray, pesos: np.ndarray | None) -> np.ndarray:
    from scipy.cluster.hierarchy import linkage
    return linkage(z, method="ward", metric="euclidean") if pesos is None \
        else uam.ward_ponderado(z, pesos)


def clasificar(d: pd.DataFrame, k: int, **opc) -> np.ndarray:
    from scipy.cluster.hierarchy import fcluster
    z, pesos = matriz_z(d, **opc)
    return fcluster(enlace(z, pesos), t=k, criterion="maxclust")


def renumerar(cl: np.ndarray, empleo: np.ndarray) -> np.ndarray:
    """Reetiqueta de mayor a menor empleo, para que el numero signifique lo mismo
    entre hojas."""
    orden = (pd.Series(empleo).groupby(pd.Series(cl)).sum()
             .sort_values(ascending=False).index.tolist())
    mapa = {c: i + 1 for i, c in enumerate(orden)}
    return np.array([mapa[c] for c in cl])


def bloque(d: pd.DataFrame, periodo: int, cl: np.ndarray,
           marcadas: set[str] | None = None) -> pd.DataFrame:
    """Las cinco columnas del formato UAM, ordenadas por conglomerado."""
    b = pd.DataFrame({
        "pais": "CHILE",
        "periodo": periodo,
        "cluster": cl,
        "codigo_3dig": d["codigo_3dig"].to_numpy(),
        "desc": d["desc"].to_numpy(),
    })
    b["marcada"] = b["codigo_3dig"].isin(marcadas or set())
    b["_ord"] = pd.to_numeric(b["codigo_3dig"], errors="coerce")
    return b.sort_values(["cluster", "_ord"]).drop(columns="_ord").reset_index(drop=True)


# --------------------------------------------------------------------------
# Los dos criterios para elegir k
# --------------------------------------------------------------------------

def costes_de_fusion(Z: np.ndarray) -> np.ndarray:
    """Coste de Ward de cada fusion. Tanto en scipy como en `ward_ponderado` la
    altura h cumple coste = h^2 / 2 (el incremento de la inercia intra-grupo)."""
    return Z[:, 2] ** 2 / 2.0


def salto_de_costes(Z: np.ndarray, kmax: int = 8) -> pd.Series:
    """Cuanto mas caro es bajar de k+1 a k grupos que bajar de k+2 a k+1.

    No hay codo en estas curvas —bajan suave—, asi que lo que discrimina es el
    salto entre fusiones consecutivas, no el nivel.
    """
    c = costes_de_fusion(Z)
    n = len(c) + 1
    out = {}
    for k in range(2, min(kmax, n - 1) + 1):
        c_k = c[n - 1 - k]          # la fusion que deja k grupos
        c_k1 = c[n - 2 - k]         # la que deja k+1
        out[k] = float(c_k / c_k1) if c_k1 > 0 else float("nan")
    return pd.Series(out, name="salto")


def eta2(cl: np.ndarray, y: np.ndarray, w: np.ndarray) -> float:
    """Fraccion de la varianza ponderada de `y` que explica la particion."""
    w = np.asarray(w, dtype=float)
    w = w / w.sum()
    mu = float((w * y).sum())
    total = float((w * (y - mu) ** 2).sum())
    if total <= 0:
        return float("nan")
    entre = 0.0
    for c in np.unique(cl):
        m = cl == c
        wc = w[m].sum()
        entre += wc * (float((w[m] * y[m]).sum() / wc) - mu) ** 2
    return entre / total


def curva_eta2(d: pd.DataFrame, Z: np.ndarray, kmax: int = 8) -> pd.Series:
    """eta^2 del log de la productividad, ponderado por empleo, para k = 2..kmax.

    Es la misma medida con que se piensan contrastar las particiones rivales del
    capitulo, y no premia aislar un extremo como si hace la silueta.
    """
    from scipy.cluster.hierarchy import fcluster
    p = (d["valor_agrega"] / d["empleo"].replace(0, np.nan)).to_numpy(dtype=float)
    piso = np.nanmin(p[p > 0]) / 2.0
    y = np.log(np.clip(np.nan_to_num(p, nan=piso), piso, None))
    w = d["empleo"].to_numpy(dtype=float)
    return pd.Series({k: eta2(fcluster(Z, t=k, criterion="maxclust"), y, w)
                      for k in range(2, kmax + 1)}, name="eta2")


def elegir_k(d: pd.DataFrame, kmax: int = 8, **opc) -> dict:
    """Aplica los dos criterios a ESTA especificacion y devuelve ambos.

    - `k_costes`: el k que maximiza el salto de costes de Ward.
    - `k_eta2`: el k donde mas sube el eta^2, o sea donde empieza la meseta.

    Manda `k_costes`, que es el criterio interno al algoritmo; `k_eta2` va como
    segunda opinion y su desacuerdo se reporta, no se esconde.
    """
    z, pesos = matriz_z(d, **opc)
    Z = enlace(z, pesos)
    salto = salto_de_costes(Z, kmax)
    eta = curva_eta2(d, Z, kmax)
    incremento = eta.diff().dropna()          # k=3..kmax
    k_costes = int(salto.idxmax())
    k_eta2 = int(incremento.idxmax()) if len(incremento) else k_costes
    return {"k": k_costes, "k_costes": k_costes, "k_eta2": k_eta2,
            "salto": salto, "eta2": eta, "Z": Z}


# --------------------------------------------------------------------------
# Escritura de la planilla
# --------------------------------------------------------------------------

COLORES = ["DDEBF7", "FCE4D6", "E2EFDA", "FFF2CC", "E4DFEC", "F2F2F2", "DEEAF6", "FBE5D6"]
FINO = Side(style="thin", color="BFBFBF")
GRUESO = Side(style="medium", color="404040")
ROJO = "C00000"

ENCABEZADOS = ["pais", "periodo", "cluster", "codigo_3dig", "desc"]
COL1, COL2 = 1, 7


def escribir_bloque(ws, b: pd.DataFrame, col0: int, titulo: str, fila_titulo: int,
                    con_marca: bool) -> None:
    fila_enc = fila_titulo + 1
    ws.cell(fila_titulo, col0, titulo).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=fila_titulo, start_column=col0,
                   end_row=fila_titulo, end_column=col0 + 4)
    ws.cell(fila_titulo, col0).alignment = Alignment(horizontal="center")

    for j, h in enumerate(ENCABEZADOS):
        c = ws.cell(fila_enc, col0 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)

    anterior = None
    for i, fila in enumerate(b.itertuples(index=False)):
        r = fila_enc + 1 + i
        nuevo = fila.cluster != anterior
        relleno = PatternFill("solid", fgColor=COLORES[(fila.cluster - 1) % len(COLORES)])
        for j, v in enumerate(fila[:5]):
            c = ws.cell(r, col0 + j, v)
            c.fill = relleno
            c.border = Border(left=FINO, right=FINO,
                              top=GRUESO if nuevo else FINO, bottom=FINO)
            if con_marca and fila.marcada:
                c.font = Font(color=ROJO, italic=True, bold=(j == 2))
            elif j == 2:
                c.font = Font(bold=True)
            if j in (1, 2, 3):
                c.alignment = Alignment(horizontal="center")
        anterior = fila.cluster

    ultima = fila_enc + len(b)
    for j in range(5):
        ws.cell(ultima, col0 + j).border = Border(left=FINO, right=FINO,
                                                  top=FINO, bottom=GRUESO)

    r = ultima + 2
    cab = ["conglomerado", "actividades"] + (["de ellas, marcadas"] if con_marca else [])
    for j, h in enumerate(cab):
        ws.cell(r, col0 + j, h).font = Font(bold=True, italic=True)
    conteo = b.groupby("cluster").agg(n=("cluster", "size"), marc=("marcada", "sum"))
    for i, (cl, fila) in enumerate(conteo.iterrows()):
        ws.cell(r + 1 + i, col0, int(cl)).fill = PatternFill(
            "solid", fgColor=COLORES[(int(cl) - 1) % len(COLORES)])
        ws.cell(r + 1 + i, col0).font = Font(bold=True)
        ws.cell(r + 1 + i, col0 + 1, int(fila["n"]))
        if con_marca:
            c = ws.cell(r + 1 + i, col0 + 2, int(fila["marc"]))
            if fila["marc"]:
                c.font = Font(color=ROJO, italic=True)


def escribir_tabla(ws, t: pd.DataFrame, fila: int, col0: int, titulo: str) -> int:
    """Escribe una tabla con el indice como primera columna. Devuelve la fila
    siguiente a la ultima escrita."""
    ws.cell(fila, col0, titulo).font = Font(bold=True, size=10)
    fila += 1
    for j, h in enumerate([t.index.name or ""] + list(t.columns)):
        c = ws.cell(fila, col0 + j, h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="404040")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, (idx, r) in enumerate(t.iterrows()):
        f = fila + 1 + i
        c = ws.cell(f, col0, int(idx) if isinstance(idx, (int, float)) else idx)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=COLORES[(int(idx) - 1) % len(COLORES)]
                             if isinstance(idx, (int, float)) else "FFFFFF")
        c.alignment = Alignment(horizontal="center")
        for j, v in enumerate(r):
            cc = ws.cell(f, col0 + 1 + j, None if pd.isna(v) else float(v))
            cc.number_format = "0.00"
            cc.border = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)
    return fila + len(t) + 2


def escribir_hoja(wb: Workbook, nombre: str, notas: list[tuple[str, str]],
                  bloques: list[tuple[pd.DataFrame, str]], con_marca: bool = False,
                  tablas: list[list[tuple[pd.DataFrame, str]]] | None = None,
                  ancho_tablas: int = 14) -> None:
    """`notas` son (texto, color); `bloques` son (tabla, titulo), izquierda y derecha.

    `tablas` son las medias por conglomerado de cada periodo. Van al pie de la
    hoja y no debajo de cada bloque, porque el bloque de la derecha ocupa las
    columnas 7 a 11 y una tabla de medias no cabe en las cinco de la izquierda.
    """
    ws = wb.create_sheet(nombre)
    ws.cell(1, 1, nombre).font = Font(bold=True, size=13)
    for i, (texto, color) in enumerate(notas):
        ws.cell(2 + i, 1, texto).font = Font(italic=True, size=9, color=color)
    fila_titulo = 2 + len(notas) + 1

    fin = fila_titulo
    for (b, titulo), col0 in zip(bloques, (COL1, COL2)):
        escribir_bloque(ws, b, col0, titulo, fila_titulo, con_marca)
        fin = max(fin, fila_titulo + 1 + len(b) + 3 + b["cluster"].nunique())

    if tablas:
        for grupo, col0 in zip(tablas, (COL1, COL1 + ancho_tablas)):
            f = fin + 2
            for t, titulo in grupo:
                f = escribir_tabla(ws, t, f, col0, titulo)

    for col0 in (COL1, COL2):
        for off, w in zip(range(5), (8, 8, 8, 11, 46)):
            ws.column_dimensions[get_column_letter(col0 + off)].width = w
    ws.column_dimensions[get_column_letter(COL2 - 1)].width = 3
    ws.freeze_panes = ws.cell(fila_titulo + 2, 1)


def medias_en_niveles(d: pd.DataFrame, cl: np.ndarray) -> pd.DataFrame:
    """Las siete variables por conglomerado, en unidades legibles.

    Los ratios se calculan como COCIENTE DE LAS SUMAS, no como promedio de los
    cocientes: la productividad de un grupo es su valor agregado total sobre su
    empleo total. Productividad y remuneracion media van como indice con la
    economia = 100, que es lo unico comparable entre 2003 y 2023.
    """
    g = d.assign(_c=cl).groupby("_c")
    s = g[["valor_agrega", "vbp", "remunera", "empleo", "fbcf", "expo", "impo"]].sum()
    tot = d[["valor_agrega", "empleo", "expo", "remunera"]].sum()

    t = pd.DataFrame(index=s.index)
    t["n"] = g.size()
    t["% empleo"] = 100 * s["empleo"] / tot["empleo"]
    t["% VA"] = 100 * s["valor_agrega"] / tot["valor_agrega"]
    t["% expo"] = 100 * s["expo"] / tot["expo"]
    t["productividad (=100)"] = 100 * (s["valor_agrega"] / tot["valor_agrega"]) / \
                                (s["empleo"] / tot["empleo"])
    t["remun. media (=100)"] = 100 * (s["remunera"] / tot["remunera"]) / \
                               (s["empleo"] / tot["empleo"])
    t["FBCF / ocupado (=100)"] = 100 * (s["fbcf"] / s["fbcf"].sum()) / \
                                 (s["empleo"] / tot["empleo"])
    t["impo / VBP"] = s["impo"] / s["vbp"]
    t["expo / VBP"] = s["expo"] / s["vbp"]
    t["(X+M) / VBP"] = (s["expo"] + s["impo"]) / s["vbp"]
    t["(X-M) / (X+M)"] = np.where((s["expo"] + s["impo"]) > 0,
                                  (s["expo"] - s["impo"]) / (s["expo"] + s["impo"]), 0.0)
    t["VA / VBP"] = s["valor_agrega"] / s["vbp"]
    t["masa salarial / VA"] = s["remunera"] / s["valor_agrega"]
    t.index.name = "conglomerado"
    return t


def medias_en_z(d: pd.DataFrame, cl: np.ndarray, **opc) -> pd.DataFrame:
    """Los centroides en puntajes z: el equivalente de la hoja
    `estadisticas_cluster` de los libros 3a y 3b.

    Son las mismas variables transformadas y estandarizadas con que se agrupo, y
    ponderadas por empleo si la especificacion pondera — o sea, los centroides
    que el propio algoritmo minimiza, no un promedio recalculado aparte.
    """
    z, pesos = matriz_z(d, **opc)
    variables = variables_de(opc["sinfbcf"], opc.get("comercio", False))
    w = np.ones(len(d)) if pesos is None else np.asarray(pesos, dtype=float)
    filas = {}
    for c in np.unique(cl):
        m = cl == c
        filas[int(c)] = (w[m, None] * z[m]).sum(axis=0) / w[m].sum()
    t = pd.DataFrame(filas, index=variables).T.sort_index()
    t.index.name = "conglomerado"
    return t


def diagnostico(b: pd.DataFrame, d: pd.DataFrame) -> tuple[str, str, int]:
    """Tamanos, reparto del empleo y cuantos grupos pesan menos del 1 %."""
    emp = pd.Series(d["empleo"].to_numpy(), index=d["codigo_3dig"].to_numpy())
    e = b.groupby("cluster")["codigo_3dig"].apply(lambda s: emp[s].sum())
    e = 100 * e / emp.sum()
    return (" / ".join(str(n) for n in b["cluster"].value_counts().sort_index()),
            " / ".join(f"{v:.1f}" for v in e.sort_index()),
            int((e < 1.0).sum()))

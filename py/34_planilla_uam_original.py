"""Bloque I.13 — la clasificacion ORIGINAL de la UAM, con formato y medias al pie.

Es la planilla de referencia de la serie T_I22-T_I27: no reagrupa nada. Toma la
particion tal como llego en los libros `3a` y `3b` —el universo que ellos
clasificaron (62 actividades en 2003 y 101 en 2023, con el cobre y las demas
excluidas ya fuera), su especificacion y su corte (k = 4 en 2003, k = 2 en
2023)— y la escribe con el mismo formato de colores de las otras planillas, con
las dos tablas de medias por conglomerado al pie de la hoja:

    en niveles     las siete variables en unidades legibles, mas n, participacion
                   en empleo, VA y exportaciones, y masa salarial sobre VA. Los
                   ratios son cociente de las sumas, no promedio de los cocientes.
    en puntajes z  los centroides con que trabaja el algoritmo, o sea el
                   equivalente de la hoja `estadisticas_cluster` de 3a y 3b.

La numeracion de los conglomerados es la de la UAM, sin renumerar por empleo:
esta hoja es el dato recibido y tiene que poder cruzarse contra sus libros.

Entrada:  datos/intermediate/uam_replica_p{1,2}.csv  (correr antes py/20)
Salida:   datos/output/AAAA.MM.DD T_I28 clasificacion original de la UAM.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402
import planilla_clusters as pc  # noqa: E402

OPC = dict(log=False, peso=False, balanza="nivel", sinfbcf=False)
K_UAM = {1: 4, 2: 2}
SALIDA = rutas.OUT / "2026.08.13 T_I28 clasificacion original de la UAM.xlsx"


def cargar(periodo: int) -> pd.DataFrame:
    d = pd.read_csv(rutas.exigir(
        rutas.INTER / f"uam_replica_p{periodo}.csv",
        "la replica de la UAM; correr antes 20_replica_hclust.py"))
    d["codigo_3dig"] = d["codigo_3dig"].astype(str)
    if d["cluster"].nunique() != K_UAM[periodo]:
        raise SystemExit(f"ERROR periodo {periodo}: {d['cluster'].nunique()} "
                         f"conglomerados en el archivo, se esperaban {K_UAM[periodo]}")
    return d.reset_index(drop=True)


def main() -> None:
    rutas.preparar_directorios()
    salida = SALIDA

    wb = Workbook()
    wb.remove(wb.active)

    bloques, tablas = [], []
    for periodo, anio in ((1, 2003), (2, 2023)):
        d = cargar(periodo)
        cl = d["cluster"].to_numpy()
        print(f"{anio}: {len(d)} actividades, k = {K_UAM[periodo]}, "
              f"tamanos {'/'.join(str(n) for n in pd.Series(cl).value_counts().sort_index())}")
        bloques.append((pc.bloque(d, periodo, cl),
                        f"Periodo {periodo} — {anio} ({len(d)} actividades, k = {K_UAM[periodo]})"))
        tablas.append([
            (pc.medias_en_niveles(d, cl),
             f"Periodo {periodo} — {anio}: medias por conglomerado, en niveles "
             "(ratios = cociente de las sumas; indices con la economia clasificada = 100)"),
            (pc.medias_en_z(d, cl, **OPC),
             f"Periodo {periodo} — {anio}: centroides en puntajes z, las siete variables tal "
             "como entran al algoritmo"),
        ])

    notas = [
        ("La clasificacion de la UAM tal como llego en los libros 3a y 3b: su universo, su "
         "especificacion (siete variables en niveles, sin ponderar, balanza comercial en nivel) "
         "y su corte (k = 4 en 2003, k = 2 en 2023). Aca no se reagrupa nada.", "000000"),
        ("El cobre esta fuera, junto con las demas actividades que ellos excluyeron: quedan 62 "
         "actividades en 2003 y 101 en 2023. Los porcentajes de las tablas de medias son sobre "
         "el total CLASIFICADO, no sobre la economia.", "000000"),
        ("Numeracion de conglomerados: la de la UAM, sin renumerar por empleo, para poder "
         "cruzarla contra sus libros. En las demas planillas de la serie se renumera.", "000000"),
    ]
    pc.escribir_hoja(wb, "UAM original", notas, bloques, con_marca=False, tablas=tablas)

    ws = wb.create_sheet("0 Lectura", 0)
    ws.cell(1, 1, "T_I28 — la clasificacion original de la UAM, con formato y medias al pie"
            ).font = Font(bold=True, size=13)
    for i, t in enumerate([
        "Es la hoja de referencia contra la que se leen T_I22 a T_I27: el punto de partida, sin "
        "ninguna correccion de especificacion ni de universo.",
        "Fuente: datos/intermediate/uam_replica_p1.csv y _p2.csv, que py/20 reconstruye desde el "
        "insumo y los libros de resultados de la UAM (puntajes z con error 1e-15 y particion "
        "identica sector por sector).",
        "Al pie de la hoja, dos tablas por periodo: medias en niveles (legibles) y centroides en "
        "puntajes z (los que minimiza el algoritmo).",
    ]):
        ws.cell(2 + i, 1, t).font = Font(italic=True, size=9)
    ws.column_dimensions["A"].width = 120

    wb.save(salida)
    print(f"\nescrito: {salida}")


if __name__ == "__main__":
    main()

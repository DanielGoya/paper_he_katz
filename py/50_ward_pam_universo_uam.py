"""Bloque K.7 — Ward contra PAM ponderado+ajustado, k = 5, con el universo de
sectores de CHILE restringido al que usa la UAM (mas el cobre reintegrado).

LA DIFERENCIA CON EL BLOQUE K.6 (DG, 2026-08-13). `49_pam_ponderado_logs.py`
(T_K13) corrio la COU COMPLETA en los cuatro paises, Chile incluido (73 y 111
actividades). Este script cambia UNA sola cosa: en Chile, en vez de la COU
completa, usa la lista de actividades que la UAM efectivamente clasifico —62 en
2003 y 101 en 2023— MAS el cobre reintegrado (63 y 102), que es exactamente el
universo de `T_I27` (bloque I.12). Brasil, Mexico y Argentina NO cambian: siguen
con la COU completa, porque ahi no hay lista de la UAM que restringir —su
criterio de exclusion no esta documentado (bloque K.4) y el cobre no es su
problema—.

POR QUE IMPORTA LA DIFERENCIA. Con la COU completa, el algoritmo tiene que
resolver DOS problemas de Chile a la vez: encontrar el enclave exportador Y
absorber a las 10-11 actividades que la UAM excluyo por otras razones (ver
bloque K.4 y la nota `2026-08-11-lectura-de-las-planillas-de-la-UAM`). Restringir
al universo de la UAM aisla el efecto de reintegrar SOLO el cobre, que es la
pregunta que motivo T_I27 y ahora se repite con k=5, ponderado y con logaritmos.

LAS DOS ESPECIFICACIONES (mismas que T_K13, solo cambia el universo chileno):

    Ward         sobre distancia euclidiana, las siete variables de la UAM en
                 niveles, sin ponderar, balanza comercial en niveles. Es la
                 especificacion publicada, con el corte comun k = 5.
    PAM+ajustes  PAM con distancia de Manhattan, PONDERADO por empleo, con
                 logaritmos en productividad y remuneracion media y balanza
                 normalizada (X-M)/(X+M). Es la especificacion 7 de T_I13.

DONDE SE MARCA EL COBRE. En Chile se conoce la actividad exacta —el cobre
reintegrado— y se marca esa, igual que en T_I27. En los otros tres paises no hay
un cobre conocido, asi que se sigue marcando el sustituto transparente del bloque
K.4/K.6: la actividad de mayor max |z| en las siete variables, y el mayor
exportador.

SALIDA. `T_K14`, un libro con nueve hojas: lectura, y dos por pais (Chile,
Brasil, Mexico, Argentina), llamadas «<pais> Ward» y «<pais> PAM+ajustes».

DEPENDE DE. `40_uam_metodo.py`, `41_cou_multipais.py`, `planilla_clusters.py` y
`26_pam_medoides.py`. El cargador de Chile es el de `33_planilla_pam.py`
(T_I27), traido aca para no importar un script que no expone funciones sueltas.
"""

from __future__ import annotations

import importlib.util
import sys
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402
import planilla_clusters as pc  # noqa: E402


def _cargar_modulo(archivo: str, nombre: str):
    """Los modulos de `py/` empiezan con digito y no se pueden importar por nombre."""
    spec = importlib.util.spec_from_file_location(
        nombre, Path(__file__).resolve().parent / archivo)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


pam_mod = _cargar_modulo("26_pam_medoides.py", "pam_medoides")
uam = pc.uam

HOY = date.today().strftime("%Y.%m.%d")
SALIDA = rutas.OUT / f"{HOY} T_K14 Ward contra PAM+ajustes k=5, universo UAM en Chile.xlsx"

K = 5

# Chile primero; sus dos anios salen del cargador restringido. Los otros tres, sin cambios.
PAISES = [
    ("Chile", [2003, 2023]),
    ("Brasil", [2003, 2021]),
    ("Mexico", [2008, 2018]),
    ("Argentina", [1997]),
]

SPEC_WARD = dict(log=False, peso=False, balanza="nivel", sinfbcf=False, comercio=False)
SPEC_PAM = dict(log=True, peso=True, balanza="norm", sinfbcf=False, comercio=False)

NOTAS = {
    "Ward": [
        ("Ward sobre distancia euclidiana, las siete variables de la UAM en niveles, sin "
         "ponderar y con la balanza comercial en niveles. Es la especificacion publicada; lo "
         f"unico que se le impone es el corte comun k = {K}, para que las dos hojas del pais "
         "sean comparables.", "000000"),
        ("Conglomerados numerados de mayor a menor empleo. Al pie, para cada anio, las medias "
         "por conglomerado en niveles y en puntajes z.", "000000"),
    ],
    "PAM+ajustes": [
        (f"PAM (k-medoides) con distancia de Manhattan, PONDERADO por empleo, con logaritmos en "
         f"productividad y remuneracion media y balanza normalizada (X-M)/(X+M). k = {K}. Es la "
         "especificacion 7 de T_I13: combina la estrategia de robustez del proyecto con la "
         "ortodoxa.", "000000"),
        ("Ponderar y logaritmar van juntos a proposito: logaritmar SIN ponderar empeora la "
         "invariancia a la desagregacion; con pesos la invariancia es exacta.", "000000"),
        ("PAM trabaja con MEDOIDES, no con medias: las tablas del pie van como descripcion. "
         "Conglomerados numerados de mayor a menor empleo.", "000000"),
    ],
}

NOTA_UNIVERSO = {
    "Chile": ("Universo: las actividades que la UAM CLASIFICA (62 en 2003, 101 en 2023) MAS el "
              "cobre reintegrado (63 y 102) — el mismo universo de T_I27. En rojo y cursiva, el "
              "cobre. Las demas exclusiones de la UAM se respetan: no es la COU completa.",
              pc.ROJO),
    "otros": ("Universo: la COU COMPLETA, sin excluir ninguna actividad — no cambia respecto de "
              "T_K13. En rojo y cursiva, la actividad mas extrema (mayor max |z| en las siete "
              "variables) y el mayor exportador: son el sustituto del cobre en un pais sin "
              "criterio de exclusion documentado.", pc.ROJO),
}


# ==========================================================================
# Carga
# ==========================================================================

def cargar_chile_uam() -> tuple[pd.DataFrame, dict[int, str]]:
    """Chile restringido al universo de la UAM mas el cobre — el de T_I27.

    Replica `33_planilla_pam.cargar()`: por periodo, las actividades que la UAM
    clasifica (hoja `sectores_cluster` de los libros 3a/3b) mas la UNICA excluida
    cuya descripcion contiene 'cobre'. Falla si no hay exactamente una, o si el
    tamano final no cuadra — igual que en T_I27.
    """
    ins = pd.read_excel(rutas.exigir(rutas.D_UAM_INSUMO, "el insumo de la UAM"))
    partes, cobres = [], {}
    for periodo, anio in ((1, 2003), (2, 2023)):
        d = ins[ins["periodo"] == periodo].copy()
        d["codigo_3dig"] = d["codigo_3dig"].astype(str)

        res = pd.read_excel(
            rutas.exigir(rutas.D_UAM_HCLUST[periodo], f"los resultados del periodo {periodo}"),
            sheet_name="sectores_cluster")
        dentro = {str(c) for c in res["codigo_3dig"]}

        fuera = d[~d["codigo_3dig"].isin(dentro)]
        hit = fuera["desc"].map(rutas.normalizar).str.contains("cobre")
        if hit.sum() != 1:
            raise SystemExit(
                f"ERROR periodo {periodo}: entre las {len(fuera)} excluidas por la UAM hay "
                f"{hit.sum()} con 'cobre' en la descripcion, no 1")
        cobre = str(fuera.loc[hit, "codigo_3dig"].iloc[0])
        cobres[anio] = cobre

        d = d[d["codigo_3dig"].isin(dentro | {cobre})].reset_index(drop=True)
        if len(d) != len(dentro) + 1:
            raise SystemExit(f"ERROR periodo {periodo}: quedaron {len(d)} actividades, "
                             f"esperaba {len(dentro) + 1}")

        partes.append(pd.DataFrame({
            "pais": "Chile", "anio": anio, "periodo": periodo,
            "codigo": d["codigo_3dig"], "desc": d["desc"],
            "vbp": d["vbp"], "valor_agrega": d["valor_agrega"],
            "remunera": d["remunera"], "empleo": d["empleo"],
            "fbcf": d["fbcf"], "expo": d["expo"], "impo": d["impo"],
        }))
    return pd.concat(partes, ignore_index=True), cobres


def cargar() -> tuple[pd.DataFrame, dict[int, str]]:
    chile, cobres = cargar_chile_uam()
    resto = pd.read_csv(rutas.exigir(rutas.INTER / "cou_multipais.csv", "la extraccion multipais"))
    # Una actividad sin remuneraciones declaradas se conserva con cero, que es lo
    # que dice la fuente; en NaN caeria despues en la matriz de distancias.
    resto["remunera"] = resto["remunera"].fillna(0.0)
    resto["codigo"] = resto["codigo"].astype(str)
    resto = resto[resto.pais != "Chile"]  # Chile sale del cargador restringido, no de aca.
    return pd.concat([chile, resto], ignore_index=True), cobres


def extremas(d: pd.DataFrame) -> np.ndarray:
    """max |z| de cada actividad en las siete variables, especificacion Ward.
    Sustituto del cobre para paises sin lista de la UAM que restringir."""
    r = uam.construir_variables(d, balanza="nivel")
    return np.abs(uam.estandarizar(r, uam.VARS_UAM)).max(axis=1)


# ==========================================================================
# Particion
# ==========================================================================

def clasificar(d: pd.DataFrame, metodo: str, k: int) -> np.ndarray:
    """Devuelve la particion, renumerada de mayor a menor empleo."""
    if metodo == "Ward":
        cl = pc.clasificar(d, k, **SPEC_WARD)
    elif metodo == "PAM+ajustes":
        z, pesos = pc.matriz_z(d, **SPEC_PAM)
        D = pam_mod.matriz_distancias(z, "manhattan")
        _, et = pam_mod.pam(D, k, pesos)
        cl = et + 1
    else:
        raise ValueError(metodo)
    return pc.renumerar(cl, d["empleo"].to_numpy())


def bloque(d: pd.DataFrame, pais: str, anio: int, cl: np.ndarray,
           marcadas: set[str]) -> pd.DataFrame:
    """Las cinco columnas del formato UAM, ordenadas por conglomerado."""
    b = pd.DataFrame({
        "pais": pais.upper(),
        "periodo": anio,
        "cluster": cl,
        "codigo_3dig": d["codigo"].to_numpy(),
        "desc": d["desc"].to_numpy(),
    })
    b["marcada"] = b["codigo_3dig"].isin(marcadas)
    b["_ord"] = pd.to_numeric(b["codigo_3dig"], errors="coerce")
    return b.sort_values(["cluster", "_ord"]).drop(columns="_ord").reset_index(drop=True)


def main() -> None:
    rutas.preparar_directorios()
    pam_mod.verificar()

    datos, cobres = cargar()
    print(f"Chile: {(datos.pais == 'Chile').sum()} filas "
          f"({(datos[(datos.pais == 'Chile') & (datos.anio == 2003)]).shape[0]} en 2003, "
          f"{(datos[(datos.pais == 'Chile') & (datos.anio == 2023)]).shape[0]} en 2023) "
          "— clasificadas por la UAM + cobre reintegrado\n")

    wb = Workbook()
    wb.remove(wb.active)
    resumen = []

    for pais, anios in PAISES:
        for metodo in ("Ward", "PAM+ajustes"):
            bloques, tablas = [], []
            spec = SPEC_WARD if metodo == "Ward" else SPEC_PAM

            for anio in anios:
                d = datos[(datos.pais == pais) & (datos.anio == anio)].reset_index(drop=True)
                if d.empty:
                    print(f"  AVISO: no hay datos de {pais} {anio}")
                    continue

                if pais == "Chile":
                    marcadas = {cobres[anio]}
                else:
                    ext = extremas(d)
                    i_ext, i_exp = int(np.argmax(ext)), int(np.argmax(d["expo"].to_numpy()))
                    marcadas = {d.loc[i_ext, "codigo"], d.loc[i_exp, "codigo"]}

                cl = clasificar(d, metodo, K)
                titulo = f"{pais} {anio} ({len(d)} actividades, k = {K})"
                bloques.append((bloque(d, pais, anio, cl, marcadas), titulo))

                t_niv = pc.medias_en_niveles(d, cl)
                tablas.append([
                    (t_niv, f"{anio}: medias por conglomerado, en niveles (ratios = cociente "
                            "de las sumas; indices con la economia = 100)"),
                    (pc.medias_en_z(d, cl, **spec),
                     f"{anio}: medias en puntajes z, con las variables tal como entran a ESTA "
                     "especificacion"),
                ])

                # El conglomerado que concentra mas exportaciones (no siempre es
                # el enclave chico: al ponderar, el peso puede repartirlas).
                g = int(t_niv["% expo"].idxmax())
                resumen.append({
                    "pais": pais, "anio": anio, "metodo": metodo,
                    "n actividades": len(d),
                    "tamanos": " / ".join(str(x) for x in t_niv["n"].astype(int)),
                    "% empleo": " / ".join(f"{x:.1f}" for x in t_niv["% empleo"]),
                    "productividad": " / ".join(f"{x:.0f}" for x in t_niv["productividad (=100)"]),
                    "empleo del mayor (%)": float(t_niv["% empleo"].max()),
                    "grupos <1% empleo": int((t_niv["% empleo"] < 1.0).sum()),
                    "grupo con mas expo: actividades": int(t_niv.loc[g, "n"]),
                    "grupo con mas expo: % empleo": float(t_niv.loc[g, "% empleo"]),
                    "grupo con mas expo: % expo": float(t_niv.loc[g, "% expo"]),
                    "grupo con mas expo: masa salarial/VA (%)":
                        100 * float(t_niv.loc[g, "masa salarial / VA"]),
                    "eta2 log productividad": pam_mod.eta2_productividad(
                        d, cl, d["empleo"].to_numpy(dtype=float)),
                })

            notas = NOTAS[metodo] + [NOTA_UNIVERSO["Chile" if pais == "Chile" else "otros"]]
            pc.escribir_hoja(wb, f"{pais} {metodo}", notas, bloques,
                             con_marca=True, tablas=tablas)
            print(f"  {pais:<10} {metodo:<12} escrita ({len(bloques)} periodo/s)")

    # ---- ARI entre los dos metodos, pais por pais
    T = pd.DataFrame(resumen)
    ari = []
    for pais, anios in PAISES:
        for anio in anios:
            d = datos[(datos.pais == pais) & (datos.anio == anio)].reset_index(drop=True)
            if d.empty:
                continue
            ari.append({
                "pais": pais, "anio": anio,
                "ARI Ward vs PAM+ajustes": uam.rand_ajustado(
                    clasificar(d, "Ward", K), clasificar(d, "PAM+ajustes", K)),
            })
    T = T.merge(pd.DataFrame(ari), on=["pais", "anio"], how="left")

    ws = wb.create_sheet("0 Lectura", 0)
    lineas = [
        (f"T_K14 — Ward contra PAM+ajustes, k = {K}, universo de la UAM en Chile", True, 13),
        ("Es T_K13 con UN cambio: en Chile el universo deja de ser la COU completa y pasa a ser "
         "el que la UAM efectivamente clasifico MAS el cobre reintegrado (63 actividades en "
         "2003, 102 en 2023) — el mismo de T_I27. Brasil, Mexico y Argentina NO cambian: siguen "
         "con la COU completa, porque ahi no hay lista de la UAM que restringir.", False, 9),
        ("Dos hojas por pais, llamadas «<pais> Ward» y «<pais> PAM+ajustes», con los anios lado "
         "a lado y las medias por conglomerado al pie de cada uno. Ward es Ward euclidiano con "
         "las siete variables en niveles, sin ponderar, balanza en niveles. PAM+ajustes es PAM "
         "Manhattan ponderado por empleo, con logaritmos y balanza normalizada.", False, 9),
        (f"El corte es el mismo en las dos hojas de cada pais, k = {K}: la comparacion es de "
         "METODO. No se recorre el numero de grupos.", False, 9),
        ("En Chile se marca el COBRE (la actividad reintegrada). En los otros tres, que no "
         "tienen un cobre conocido, se marca el sustituto transparente del bloque K.4/K.6: la "
         "actividad mas extrema (max |z|) y el mayor exportador.", False, 9),
        (f"Generada el {HOY} por py/50_ward_pam_universo_uam.py.", False, 9),
        ("", False, 9),
        ("Resumen (una fila por pais, anio y metodo):", True, 10),
    ]
    for i, (texto, negrita, tam) in enumerate(lineas):
        ws.cell(1 + i, 1, texto).font = Font(bold=negrita, italic=not negrita, size=tam)
    f0 = len(lineas) + 2
    for j, h in enumerate(T.columns):
        ws.cell(f0, 1 + j, h).font = Font(bold=True, size=9)
    for i, fila in enumerate(T.itertuples(index=False)):
        for j, v in enumerate(fila):
            c = ws.cell(f0 + 1 + i, 1 + j,
                        None if (isinstance(v, float) and np.isnan(v)) else v)
            if isinstance(v, float):
                c.number_format = "0.000"
    ws.column_dimensions["A"].width = 120

    wb.save(SALIDA)
    print(f"\nescrito: {SALIDA}\n")
    print(T.round(3).to_string(index=False))


if __name__ == "__main__":
    main()

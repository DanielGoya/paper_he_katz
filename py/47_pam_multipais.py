"""Bloque K.5 — PAM / k-medoides con k = 4 en Argentina, Brasil y Mexico.

LA PREGUNTA (DG, 2026-08-13). El bloque I.12 (`33_planilla_pam.py`, T_I27) mostro
que en Chile PAM resuelve el problema del cobre SIN tocar ninguna variable: con
las siete de la UAM intactas y k = 3 o 4, el cobre deja de quedar solo y entra a
un enclave de 15 actividades con 5,6 % del empleo y 82,2 % de las exportaciones,
que es el mismo objeto que encuentran la especificacion preferida del capitulo y
la reparametrizacion del comercio. Falta la misma corrida afuera: es el punto 1
de «Lo que NO se probo fuera de Chile» de MEMORY.md, y es la corrida mas barata.

QUE SE HACE. Exactamente lo de T_I27, con k = 4 (lo pedido) y sin recorrer otros
cortes:

    A  sin ajustes          las SIETE variables de la UAM intactas: en niveles,
                            sin ponderar, balanza comercial en niveles. Lo unico
                            que cambia respecto del ejercicio publicado es el
                            algoritmo: PAM Manhattan en vez de Ward euclidiano.
    B  comercio ajustado    las SEIS de la reparametrizacion del bloque I.11:
                            apertura (X+M)/VBP y balanza (X-M)/(X+M) en lugar de
                            las tres columnas de comercio de la UAM, de las
                            cuales una era funcion exacta de las otras dos.

Sin logaritmos y sin ponderacion en las dos, a proposito: la pregunta es que hace
PAM POR SI SOLO. Agregar las tres mejoras la contestaria mezclada, y ademas
`42_cluster_multipais.py` ya corre el paquete completo con Ward.

EL UNIVERSO. La COU COMPLETA de cada pais, sin excluir nada. Es la unica opcion
honesta afuera: la UAM excluyo actividades de Chile y no documento el criterio
(bloque K.4), asi que no hay subconjunto «suyo» que respetar en Brasil, Mexico ni
Argentina. Chile va como referencia, tambien con la COU completa —73 y 111
actividades, del insumo de la UAM, que es la unica fuente chilena que trae la COU
entera CON empleo—, de modo que su fila es comparable con las demas. Ojo: ese
universo no es identico al de T_I27, que corre sobre las clasificadas mas el
cobre (63 y 102); las cifras chilenas de aca se leen contra T_I23, no contra
T_I27.

QUE SE MIRA. Lo mismo que en Chile, con el analogo del cobre construido de forma
transparente, porque cada pais tiene el suyo:

  - la actividad EXTREMA: la de mayor max |z| en las siete variables, que es el
    sustituto del criterio de exclusion ya usado en el bloque K.4;
  - el mayor EXPORTADOR;
  - el ENCLAVE: el conglomerado que concentra mas exportaciones.

De cada uno se reporta si queda aislado y con que masa salarial, que es la cifra
con la que el bloque K comparo los cuatro paises (20,1 · 22,8 · 21,1 % del VA).

SALIDA. `T_K11`: catorce hojas de clasificacion (siete pais-periodos x dos
especificaciones) en el formato de los libros 3a y 3b, con las medias por
conglomerado al pie, mas cuatro hojas de diagnostico.

DEPENDE DE. `40_uam_metodo.py`, `41_cou_multipais.py` (escribe
`intermediate/cou_multipais.csv`), `planilla_clusters.py` y `26_pam_medoides.py`,
de donde sale la implementacion de PAM ya verificada.
"""

from __future__ import annotations

import importlib.util
import sys
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402
import planilla_clusters as pc  # noqa: E402


def _cargar_modulo(archivo: str, nombre: str):
    """Los modulos de `py/` empiezan con digito y no se pueden importar por nombre."""
    spec = importlib.util.spec_from_file_location(
        nombre, Path(__file__).resolve().parent / archivo)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


pam_mod = _cargar_modulo("26_pam_medoides.py", "pam_medoides")
uam = pc.uam

HOY = date.today().strftime("%Y.%m.%d")
SALIDA = rutas.OUT / f"{HOY} T_K11 PAM con k=4 en los cuatro paises.xlsx"

K = 4

# Chile primero, como referencia: es el pais donde el resultado ya esta medido.
ORDEN = [("Chile", 2003), ("Chile", 2023),
         ("Brasil", 2003), ("Brasil", 2021),
         ("Mexico", 2008), ("Mexico", 2018),
         ("Argentina", 1997)]

ESPECS = [
    ("A sin ajustes",
     dict(log=False, peso=False, balanza="nivel", sinfbcf=False, comercio=False)),
    ("B comercio ajustado",
     dict(log=False, peso=False, balanza="norm", sinfbcf=False, comercio=True)),
]

QUE_HACE = {
    "A sin ajustes":
        "Las SIETE variables de la UAM intactas: en niveles, sin ponderar, balanza comercial "
        "en niveles. Lo unico que cambia respecto del ejercicio publicado es el algoritmo: "
        "PAM con distancia de Manhattan en lugar de Ward euclidiano.",
    "B comercio ajustado":
        "Las SEIS variables de la reparametrizacion: apertura (X+M)/VBP y balanza (X-M)/(X+M) "
        "en lugar de las tres columnas de comercio de la UAM, de las cuales una era funcion "
        "exacta de las otras dos. Sigue sin logaritmos y sin ponderar.",
}


# ==========================================================================
# Carga: la COU completa de cada pais
# ==========================================================================

def cargar_chile() -> pd.DataFrame:
    """Chile con las 73 y 111 actividades, del insumo de la UAM.

    `cou_multipais.csv` trae a Chile ya recortado al subconjunto que la UAM
    clasifica (62 y 101), asi que para tener la COU completa —y con ella el
    cobre— hay que ir al insumo, que es la unica fuente chilena con todas las
    actividades Y empleo: la COU del Banco Central no publica ocupados.
    """
    ins = pd.read_excel(rutas.exigir(rutas.D_UAM_INSUMO, "el insumo de la UAM"))
    return pd.DataFrame({
        "pais": "Chile", "anio": ins["anio"], "periodo": ins["periodo"],
        "codigo": ins["codigo_3dig"].astype(str), "desc": ins["desc"],
        "vbp": ins["vbp"], "valor_agrega": ins["valor_agrega"],
        "remunera": ins["remunera"], "empleo": ins["empleo"],
        "fbcf": ins["fbcf"], "expo": ins["expo"], "impo": ins["impo"],
    })


def cargar() -> pd.DataFrame:
    d = pd.read_csv(rutas.exigir(rutas.INTER / "cou_multipais.csv", "la extraccion multipais"))
    # Una actividad sin remuneraciones declaradas se conserva con cero, que es lo
    # que dice la fuente; dejarla en NaN caeria despues en la matriz de distancias.
    d["remunera"] = d["remunera"].fillna(0.0)
    d["codigo"] = d["codigo"].astype(str)
    return pd.concat([cargar_chile(), d[d.pais != "Chile"]], ignore_index=True)


def extremas(d: pd.DataFrame) -> np.ndarray:
    """max |z| de cada actividad en las siete variables. Es el sustituto del
    criterio de exclusion del bloque K.4, usado aca solo para NOMBRAR al analogo
    del cobre de cada pais, no para sacar nada."""
    r = uam.construir_variables(d, balanza="nivel")
    return np.abs(uam.estandarizar(r, uam.VARS_UAM)).max(axis=1)


# ==========================================================================
# Particion
# ==========================================================================

def particionar(d: pd.DataFrame, k: int, metrica: str,
                **opc) -> tuple[np.ndarray, list[int], np.ndarray]:
    """PAM sobre las variables de una especificacion. Devuelve etiquetas
    renumeradas por empleo, los medoides (en el indice ORIGINAL) y la matriz de
    distancias."""
    z, pesos = pc.matriz_z(d, **opc)
    D = pam_mod.matriz_distancias(z, metrica)
    med, et = pam_mod.pam(D, k, pesos)

    # `renumerar` reordena las etiquetas por empleo; hay que arrastrar los
    # medoides al numero nuevo o dejan de corresponder con su conglomerado.
    nuevo = pc.renumerar(et + 1, d["empleo"].to_numpy())
    mapa = {int(et[m] + 1): int(nuevo[m]) for m in med}
    medoides = [med[i] for i in np.argsort([mapa[int(et[m] + 1)] for m in med])]
    return nuevo, medoides, D


def ward(d: pd.DataFrame, k: int, **opc) -> np.ndarray:
    """La misma especificacion con Ward, para comparar."""
    return pc.renumerar(pc.clasificar(d, k, **opc), d["empleo"].to_numpy())


def bloque(d: pd.DataFrame, pais: str, anio: int, cl: np.ndarray,
           marcadas: set[str]) -> pd.DataFrame:
    """Las cinco columnas del formato UAM, ordenadas por conglomerado."""
    b = pd.DataFrame({
        "pais": pais.upper(),
        "periodo": anio,
        "cluster": cl,
        "codigo_3dig": d["codigo"].to_numpy(),
        "desc": d["desc"].to_numpy(),
    })
    b["marcada"] = b["codigo_3dig"].isin(marcadas)
    b["_ord"] = pd.to_numeric(b["codigo_3dig"], errors="coerce")
    return b.sort_values(["cluster", "_ord"]).drop(columns="_ord").reset_index(drop=True)


def ficha(d: pd.DataFrame, cl: np.ndarray, grupo: int) -> dict:
    """La lectura economica de un conglomerado, con la masa salarial que es la
    cifra con que el bloque K compara los cuatro paises."""
    sub = d[cl == grupo]
    tot_x = d["expo"].sum()
    return {
        "actividades": int(len(sub)),
        "solo?": "SI — sigue aislado" if len(sub) == 1 else "no",
        "% del empleo": 100 * sub["empleo"].sum() / d["empleo"].sum(),
        "% de las expo": 100 * sub["expo"].sum() / tot_x if tot_x else np.nan,
        "productividad (=100)":
            100 * (sub["valor_agrega"].sum() / sub["empleo"].sum())
            / (d["valor_agrega"].sum() / d["empleo"].sum()),
        "masa salarial / VA (%)":
            100 * sub["remunera"].sum() / sub["valor_agrega"].sum(),
    }


# ==========================================================================
# Escritura
# ==========================================================================

def hoja_tabla(wb: Workbook, nombre: str, titulo: str, notas: list[str],
               t: pd.DataFrame, anchos: tuple) -> None:
    ws = wb.create_sheet(nombre)
    ws.cell(1, 1, titulo).font = Font(bold=True, size=13)
    for i, n in enumerate(notas):
        ws.cell(2 + i, 1, n).font = Font(italic=True, size=9)
    f0 = 2 + len(notas) + 1
    for j, h in enumerate(t.columns):
        c = ws.cell(f0, 1 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
    for i, fila in enumerate(t.itertuples(index=False)):
        for j, v in enumerate(fila):
            c = ws.cell(f0 + 1 + i, 1 + j,
                        None if (isinstance(v, float) and np.isnan(v)) else v)
            if isinstance(v, float):
                c.number_format = "0.000"
    for j, w in enumerate(anchos):
        ws.column_dimensions[chr(ord("A") + j)].width = w
    ws.freeze_panes = ws.cell(f0 + 1, 1)


def main() -> None:
    rutas.preparar_directorios()
    pam_mod.verificar()

    datos = cargar()
    wb = Workbook()
    wb.remove(wb.active)
    criterios, extremos, enclaves, comparacion, medoides_filas = [], [], [], [], []

    for nombre, opc in ESPECS:
        for pais, anio in ORDEN:
            d = datos[(datos.pais == pais) & (datos.anio == anio)].reset_index(drop=True)
            if d.empty:
                print(f"  AVISO: no hay datos de {pais} {anio}")
                continue
            w = d["empleo"].to_numpy(dtype=float)

            # --- los dos casos que hacen las veces del cobre en cada pais
            ext = extremas(d)
            i_ext = int(np.argmax(ext))
            i_exp = int(np.argmax(d["expo"].to_numpy()))
            marcadas = {d.loc[i_ext, "codigo"], d.loc[i_exp, "codigo"]}

            cl, med, D = particionar(d, K, "manhattan", **opc)

            b = bloque(d, pais, anio, cl, marcadas)
            t_niv = pc.medias_en_niveles(d, cl)
            titulo = f"{pais} {anio} ({len(d)} actividades, k = {K})"
            tablas = [[
                (t_niv, f"{titulo}: medias por conglomerado, en niveles "
                        "(ratios = cociente de las sumas; indices con la economia = 100)"),
                (pc.medias_en_z(d, cl, **opc),
                 f"{titulo}: medias en puntajes z. PAM trabaja con MEDOIDES, no con medias: "
                 "la media va como descripcion, y el medoide de cada grupo esta en la hoja "
                 "«04 Medoides»."),
            ]]
            notas = [
                (QUE_HACE[nombre], "000000"),
                (f"PAM (k-medoides) con distancia de Manhattan, k = {K}. Sin logaritmos y sin "
                 "ponderar: la pregunta es que hace el algoritmo por si solo. Conglomerados "
                 "numerados de mayor a menor empleo.", "000000"),
                ("Universo: la COU COMPLETA, sin excluir ninguna actividad. En rojo y cursiva, "
                 "la actividad mas extrema (mayor max |z| en las siete variables) y el mayor "
                 "exportador: son el analogo del cobre chileno.", pc.ROJO),
            ]
            pc.escribir_hoja(wb, f"{pais} {anio} {nombre[0]}", notas,
                             [(b, titulo)], con_marca=True, tablas=tablas)

            # ---- criterios de calidad
            criterios.append({
                "especificacion": nombre, "pais": pais, "anio": anio,
                "n actividades": len(d),
                "tamanos": " / ".join(str(x) for x in t_niv["n"].astype(int)),
                "% empleo": " / ".join(f"{x:.1f}" for x in t_niv["% empleo"]),
                "empleo del mayor (%)": float(t_niv["% empleo"].max()),
                "grupos <1% empleo": int((t_niv["% empleo"] < 1.0).sum()),
                "silueta ponderada": pam_mod.silueta(D, cl, w),
                "silueta sin ponderar": pam_mod.silueta(D, cl, None),
                "eta2 log productividad": pam_mod.eta2_productividad(d, cl, w),
            })

            # ---- donde queda cada analogo del cobre
            for etiqueta, i in (("la mas extrema (max |z|)", i_ext),
                                ("el mayor exportador", i_exp)):
                extremos.append({
                    "especificacion": nombre, "pais": pais, "anio": anio,
                    "caso": etiqueta,
                    "actividad": str(d.loc[i, "desc"])[:44],
                    "max |z|": float(ext[i]),
                    "% de las expo del pais": 100 * d.loc[i, "expo"] / d["expo"].sum(),
                    "conglomerado": int(cl[i]),
                    **ficha(d, cl, int(cl[i])),
                    "acompanantes de mas empleo": ", ".join(
                        d[(cl == cl[i]) & (d["codigo"] != d.loc[i, "codigo"])]
                        .sort_values("empleo", ascending=False)["desc"]
                        .astype(str).str[:30].head(4)),
                })

            # ---- el enclave: el conglomerado con mas exportaciones
            g_enc = int(t_niv["% expo"].idxmax())
            sub = d[cl == g_enc]
            enclaves.append({
                "especificacion": nombre, "pais": pais, "anio": anio,
                "conglomerado": g_enc, **ficha(d, cl, g_enc),
                "trae la mas extrema?": "si" if cl[i_ext] == g_enc else "no",
                "trae al mayor exportador?": "si" if cl[i_exp] == g_enc else "no",
                "actividades de mas empleo": ", ".join(
                    sub.sort_values("empleo", ascending=False)["desc"]
                    .astype(str).str[:30].head(5)),
            })

            # ---- medoides
            for i, m in enumerate(med, start=1):
                medoides_filas.append({
                    "especificacion": nombre, "pais": pais, "anio": anio,
                    "conglomerado": i,
                    "medoide (actividad representativa)": str(d.loc[m, "desc"]),
                    "codigo": d.loc[m, "codigo"],
                    "empleo del medoide": float(d.loc[m, "empleo"]),
                    "actividades en el grupo": int((cl == i).sum()),
                })

            # ---- PAM contra Ward, y Manhattan contra euclidiana
            cl_euclid, _, _ = particionar(d, K, "euclid", **opc)
            cl_ward = ward(d, K, **opc)
            comparacion.append({
                "especificacion": nombre, "pais": pais, "anio": anio,
                "ARI PAM Manhattan vs Ward": uam.rand_ajustado(cl, cl_ward),
                "ARI PAM Manhattan vs PAM euclid.": uam.rand_ajustado(cl, cl_euclid),
                "ARI PAM euclid. vs Ward": uam.rand_ajustado(cl_euclid, cl_ward),
                "tamanos con Ward": " / ".join(
                    str(x) for x in pd.Series(cl_ward).value_counts().sort_index()),
            })

            print(f"  {nombre:<22} {pais:<10} {anio}  escrita")

    # ---------------------------------------------------------------- hojas
    hoja_tabla(
        wb, "01 Criterios",
        f"Como reparte PAM con k = {K}, pais por pais",
        ["«empleo del mayor» por encima de 90 % significa que la particion no informa; es lo "
         "que le pasa a la especificacion de la UAM con Ward afuera de Chile (bloque K, "
         "T_K3). «grupos <1 %» cuenta los conglomerados gastados en actividades marginales.",
         "La silueta de Kaufman y Rousseeuw es la companera natural de PAM, pero PREMIA "
         "AISLAR UN EXTREMO: leerla junto a las dos columnas anteriores, no sola.",
         "eta2: fraccion de la varianza del log de la productividad, entre actividades y "
         "ponderada por empleo, que queda ENTRE conglomerados. Es la medida sustantiva."],
        pd.DataFrame(criterios), (22, 12, 8, 12, 22, 28, 16, 14, 16, 16, 18))

    hoja_tabla(
        wb, "02 El analogo del cobre",
        "Donde queda la actividad extrema, y donde el mayor exportador",
        ["Cada pais tiene su version del problema del cobre y no se puede nombrar a mano: se "
         "usan dos definiciones transparentes. La EXTREMA es la de mayor max |z| en las siete "
         "variables (el sustituto del criterio de exclusion del bloque K.4); el MAYOR "
         "EXPORTADOR es el que concentra mas exportaciones. En Chile las dos suelen ser el "
         "cobre; afuera no tienen por que coincidir.",
         "«solo? = SI» significa que el algoritmo gasto un conglomerado en una sola actividad, "
         "que es el modo de falla que llevo a la UAM a excluir sectores."],
        pd.DataFrame(extremos), (22, 12, 8, 24, 46, 10, 18, 14, 12, 18, 14, 14, 16, 18, 56))

    hoja_tabla(
        wb, "03 El enclave exportador",
        "El conglomerado que concentra las exportaciones, pais por pais",
        ["Es el objeto que sobrevive a todas las especificaciones en Chile: pocas actividades, "
         "poco empleo, la mayor parte de las exportaciones y una masa salarial sobre VA de "
         "alrededor del 20 %, contra 45-55 % en el resto de la economia.",
         "El bloque K encontro esa misma cifra en los tres paises con la especificacion "
         "mejorada de Ward (20,1 · 22,8 · 21,1 %). Esta hoja dice si PAM la encuentra tambien "
         "SIN ninguna de las tres mejoras."],
        pd.DataFrame(enclaves), (22, 12, 8, 14, 12, 18, 14, 14, 18, 18, 18, 22, 56))

    hoja_tabla(
        wb, "04 PAM contra Ward",
        "Cuanto cambia la particion al cambiar de algoritmo y de metrica",
        ["ARI ajustado sobre las MISMAS variables y el MISMO k: lo unico que cambia es el "
         "metodo. 1,000 seria la misma particion.",
         "Las hojas de clasificacion usan PAM Manhattan. La euclidiana se corre igual, para "
         "separar el efecto del algoritmo (aglomerar contra particionar) del de la metrica "
         "(elevar al cuadrado o no)."],
        pd.DataFrame(comparacion), (22, 12, 8, 26, 30, 26, 24))

    hoja_tabla(
        wb, "05 Medoides",
        "La actividad representativa de cada conglomerado",
        ["Es lo que PAM tiene y Ward no: el centro de cada grupo es una actividad REAL, no un "
         "punto promedio. Sirve para nombrar los estratos en el texto del capitulo.",
         "El medoide minimiza la suma de distancias al resto de su grupo: es el analogo "
         "multivariado de la mediana."],
        pd.DataFrame(medoides_filas), (22, 12, 8, 14, 56, 12, 18, 20))

    # --- Lectura
    ws = wb.create_sheet("0 Lectura", 0)
    for i, (texto, negrita, tam) in enumerate([
        (f"T_K11 — PAM / k-medoides con k = {K} en Argentina, Brasil, Chile y Mexico", True, 13),
        ("Es T_I27 llevada afuera. En Chile, PAM resuelve el problema del cobre sin tocar "
         "ninguna variable: con las siete de la UAM intactas el cobre deja de quedar solo y "
         "entra a un enclave exportador. Esta planilla corre esa misma prueba en los otros "
         "tres paises, que es el punto 1 de «lo que NO se probo fuera de Chile».", False, 9),
        ("Dos especificaciones, las dos SIN logaritmos y SIN ponderar: A, las siete variables "
         "de la UAM intactas (lo unico que cambia es el algoritmo); B, con el comercio "
         "reparametrizado (apertura y balanza normalizada en lugar de tres columnas, una de "
         "las cuales era funcion exacta de las otras dos).", False, 9),
        ("Universo: la COU COMPLETA de cada pais, sin excluir nada. Afuera de Chile no hay "
         "subconjunto de la UAM que respetar, porque su criterio de exclusion no esta "
         "documentado (bloque K.4). Chile va con sus 73 y 111 actividades, del insumo de la "
         "UAM: sus cifras se leen contra T_I23, no contra T_I27, que corre sobre 63 y 102.", False, 9),
        ("Empleo: Brasil, ocupaciones de la TRU del IBGE; Mexico, puestos de trabajo de la MIP "
         "del INEGI; Argentina, MIP del INDEC de 1997, el unico ano con remuneraciones y "
         "empleo. Los anos no son elegibles: son los ultimos con COU desagregada y empleo.", False, 9),
        ("Diagnostico: 01 criterios · 02 el analogo del cobre · 03 el enclave exportador · "
         "04 PAM contra Ward · 05 medoides.", False, 9),
        (f"Generada el {HOY} por py/47_pam_multipais.py.", False, 9),
    ]):
        ws.cell(1 + i, 1, texto).font = Font(bold=negrita, italic=not negrita, size=tam)
    ws.column_dimensions["A"].width = 200

    wb.save(SALIDA)
    print(f"\nescrito: {SALIDA}\n")

    print("== Criterios ==")
    print(pd.DataFrame(criterios)[
        ["especificacion", "pais", "anio", "tamanos", "% empleo",
         "empleo del mayor (%)", "grupos <1% empleo", "eta2 log productividad"]
    ].round(3).to_string(index=False))
    print("\n== El analogo del cobre ==")
    print(pd.DataFrame(extremos)[
        ["especificacion", "pais", "anio", "caso", "actividad", "actividades", "solo?",
         "% del empleo", "% de las expo", "masa salarial / VA (%)"]].round(1).to_string(index=False))
    print("\n== El enclave exportador ==")
    print(pd.DataFrame(enclaves)[
        ["especificacion", "pais", "anio", "actividades", "% del empleo", "% de las expo",
         "productividad (=100)", "masa salarial / VA (%)"]].round(1).to_string(index=False))
    print("\n== PAM contra Ward ==")
    print(pd.DataFrame(comparacion).round(3).to_string(index=False))


if __name__ == "__main__":
    main()

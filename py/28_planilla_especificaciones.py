"""Bloque I.7 — la clasificacion de la UAM y ocho especificaciones alternativas,
en una sola planilla con el formato de los libros 3a y 3b.

Una hoja por especificacion. Cada hoja lleva los dos periodos LADO A LADO, con
las cinco columnas de identificacion de `sectores_cluster` de la UAM
(pais, periodo, cluster, codigo_3dig, desc) y nada mas. Los conglomerados van
marcados por color de fila y separados por un borde grueso.

La hoja `1 UAM (base)` reproduce la clasificacion recibida: k = 4 en 2003 y
k = 2 en 2023, tal como vienen los libros 3a y 3b. Las ocho alternativas corren
todas con **k = 4** en los dos periodos, que es el corte que justifican el salto
de costes de Ward (bloque I.5) y el eta cuadrado (bloque I.6).

Las cuatro piezas que se combinan:

    log       logaritmos en productividad y remuneracion media (log1p en los
              ratios con ceros); la balanza pasa a (X-M)/(X+M)
    peso      ponderacion por empleo, en los momentos y en el criterio de Ward
    balanza   balanza normalizada (X-M)/(X+M) en vez de X-M en niveles
    sinfbcf   se saca x_capital_trabajo (FBCF por ocupado)

Entrada:  datos/intermediate/uam_replica_p{1,2}.csv  (los produce py/20)
Salida:   datos/output/2026.08.13 T_I22 clasificacion en nueve especificaciones.xlsx
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402

_spec = importlib.util.spec_from_file_location(
    "uam_metodo", Path(__file__).resolve().parent / "40_uam_metodo.py"
)
uam = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(uam)

K = 4
SALIDA = rutas.OUT / "2026.08.13 T_I22 clasificacion en nueve especificaciones.xlsx"

# nombre de hoja -> (logs, ponderar por empleo, balanza, sin FBCF)
ESPECIFICACIONES = [
    ("2 log",                       dict(log=True,  peso=False, balanza="nivel", sinfbcf=False)),
    ("3 peso empleo",               dict(log=False, peso=True,  balanza="nivel", sinfbcf=False)),
    ("4 balanza ajustada",          dict(log=False, peso=False, balanza="norm",  sinfbcf=False)),
    ("5 sin FBCF",                  dict(log=False, peso=False, balanza="nivel", sinfbcf=True)),
    ("6 log+peso",                  dict(log=True,  peso=True,  balanza="nivel", sinfbcf=False)),
    ("7 log+peso+balanza",          dict(log=True,  peso=True,  balanza="norm",  sinfbcf=False)),
    ("8 log+peso+sin FBCF",         dict(log=True,  peso=True,  balanza="nivel", sinfbcf=True)),
    ("9 log+peso+balanza+sin FBCF", dict(log=True,  peso=True,  balanza="norm",  sinfbcf=True)),
]

DESCRIPCION = {
    "1 UAM (base)":
        "Clasificacion recibida de la UAM (libros 3a y 3b). Siete variables en niveles, "
        "sin ponderar, balanza comercial en niveles. k = 4 en 2003 y k = 2 en 2023.",
    "2 log":
        "Logaritmos en productividad y remuneracion media, log1p en impo/VBP y expo/VBP. "
        "FBCF por ocupado no admite logaritmo (tiene ceros y negativos) y queda en nivel.",
    "3 peso empleo":
        "Ponderacion por empleo: momentos ponderados en la estandarizacion y masas en el "
        "criterio de Ward, coste (Wa*Wb/(Wa+Wb))*||ca-cb||^2.",
    "4 balanza ajustada":
        "Balanza comercial normalizada (X-M)/(X+M) en vez de X-M en niveles. Es la unica "
        "de las siete que no era un ratio.",
    "5 sin FBCF":
        "Se saca x_capital_trabajo (FBCF por ocupado), que la COU asigna por producto y no "
        "mide intensidad de capital. Quedan seis variables.",
    "6 log+peso":
        "Logaritmos mas ponderacion por empleo. Balanza todavia en niveles.",
    "7 log+peso+balanza":
        "Logaritmos, ponderacion por empleo y balanza normalizada. Es la especificacion "
        "preferida del capitulo: invariante a la desagregacion (ARI = 1,000 en la prueba de clones).",
    "8 log+peso+sin FBCF":
        "Logaritmos y ponderacion por empleo, sin FBCF por ocupado. Balanza en niveles.",
    "9 log+peso+balanza+sin FBCF":
        "Las cuatro correcciones juntas: logaritmos, ponderacion por empleo, balanza "
        "normalizada y sin FBCF por ocupado. Seis variables.",
}

# Colores de relleno por conglomerado. Suaves, para que el texto se lea.
COLORES = ["DDEBF7", "FCE4D6", "E2EFDA", "FFF2CC", "E4DFEC", "F2F2F2", "DEEAF6", "FBE5D6"]

FINO = Side(style="thin", color="BFBFBF")
GRUESO = Side(style="medium", color="404040")


def cargar(periodo: int) -> pd.DataFrame:
    d = pd.read_csv(rutas.exigir(rutas.INTER / f"uam_replica_p{periodo}.csv",
                                 f"la replica del periodo {periodo}"))
    d["codigo_3dig"] = d["codigo_3dig"].astype(str)
    return d


def clasificar(d: pd.DataFrame, log: bool, peso: bool, balanza: str,
               sinfbcf: bool) -> np.ndarray:
    """Devuelve la particion en K grupos bajo una especificacion."""
    # Las cuatro piezas son independientes: logaritmar NO arrastra la balanza
    # normalizada. Es lo que permite ver, comparando las hojas 6 y 7, cuanto
    # aporta cada correccion por separado. La balanza en niveles no admite
    # logaritmo (es negativa en buena parte de las actividades) y por eso queda
    # cruda al lado de las variables logaritmadas cuando `balanza='nivel'`.
    x = uam.construir_variables(d, balanza=balanza)

    variables = [v for v in uam.VARS_UAM if not (sinfbcf and v == "x_capital_trabajo")]
    if log:
        x = uam.transformar_logs(x, variables)

    pesos = d["empleo"].to_numpy(dtype=float) if peso else None
    return uam.agrupar(x, variables, K, pesos)


def renumerar(cl: np.ndarray, empleo: np.ndarray) -> np.ndarray:
    """Reetiqueta los conglomerados de mayor a menor empleo, para que el numero
    signifique lo mismo entre especificaciones."""
    orden = (pd.Series(empleo).groupby(pd.Series(cl)).sum()
             .sort_values(ascending=False).index.tolist())
    mapa = {c: i + 1 for i, c in enumerate(orden)}
    return np.array([mapa[c] for c in cl])


def bloque(d: pd.DataFrame, periodo: int, cl: np.ndarray) -> pd.DataFrame:
    """Las cinco columnas del formato UAM, ordenadas por conglomerado."""
    b = pd.DataFrame({
        "pais": "CHILE",
        "periodo": periodo,
        "cluster": cl,
        "codigo_3dig": d["codigo_3dig"].to_numpy(),
        "desc": d["desc"].to_numpy(),
    })
    b["_ord"] = pd.to_numeric(b["codigo_3dig"], errors="coerce")
    return b.sort_values(["cluster", "_ord"]).drop(columns="_ord").reset_index(drop=True)


# --------------------------------------------------------------------------
# Escritura
# --------------------------------------------------------------------------

ENCABEZADOS = ["pais", "periodo", "cluster", "codigo_3dig", "desc"]
COL1, COL2 = 1, 7          # columnas de arranque de cada bloque
FILA_TITULO, FILA_ENC = 3, 4


def escribir_bloque(ws, b: pd.DataFrame, col0: int, titulo: str) -> None:
    ws.cell(FILA_TITULO, col0, titulo).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=FILA_TITULO, start_column=col0,
                   end_row=FILA_TITULO, end_column=col0 + 4)
    ws.cell(FILA_TITULO, col0).alignment = Alignment(horizontal="center")

    for j, h in enumerate(ENCABEZADOS):
        c = ws.cell(FILA_ENC, col0 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)

    anterior = None
    for i, fila in enumerate(b.itertuples(index=False)):
        r = FILA_ENC + 1 + i
        nuevo = fila.cluster != anterior
        relleno = PatternFill("solid", fgColor=COLORES[(fila.cluster - 1) % len(COLORES)])
        for j, v in enumerate(fila):
            c = ws.cell(r, col0 + j, v)
            c.fill = relleno
            c.border = Border(left=FINO, right=FINO,
                              top=GRUESO if nuevo else FINO, bottom=FINO)
            if j == 2:                      # la columna cluster, en negrita
                c.font = Font(bold=True)
            if j in (1, 2, 3):
                c.alignment = Alignment(horizontal="center")
        anterior = fila.cluster

    ultima = FILA_ENC + len(b)
    for j in range(5):
        ws.cell(ultima, col0 + j).border = Border(left=FINO, right=FINO,
                                                  top=FINO, bottom=GRUESO)

    # Resumen de tamanos debajo del bloque.
    r = ultima + 2
    ws.cell(r, col0, "conglomerado").font = Font(bold=True, italic=True)
    ws.cell(r, col0 + 1, "actividades").font = Font(bold=True, italic=True)
    for i, (cl, n) in enumerate(b["cluster"].value_counts().sort_index().items()):
        ws.cell(r + 1 + i, col0, int(cl)).fill = PatternFill(
            "solid", fgColor=COLORES[(int(cl) - 1) % len(COLORES)])
        ws.cell(r + 1 + i, col0).font = Font(bold=True)
        ws.cell(r + 1 + i, col0 + 1, int(n))


def escribir_hoja(wb: Workbook, nombre: str, b1: pd.DataFrame, b2: pd.DataFrame,
                  nota_k: str) -> None:
    ws = wb.create_sheet(nombre)
    ws.cell(1, 1, nombre).font = Font(bold=True, size=13)
    ws.cell(2, 1, DESCRIPCION[nombre] + " " + nota_k).font = Font(italic=True, size=9)

    escribir_bloque(ws, b1, COL1, "Periodo 1 — 2003")
    escribir_bloque(ws, b2, COL2, "Periodo 2 — 2023")

    for col0 in (COL1, COL2):
        ws.column_dimensions[get_column_letter(col0)].width = 8
        ws.column_dimensions[get_column_letter(col0 + 1)].width = 8
        ws.column_dimensions[get_column_letter(col0 + 2)].width = 8
        ws.column_dimensions[get_column_letter(col0 + 3)].width = 11
        ws.column_dimensions[get_column_letter(col0 + 4)].width = 46
    ws.column_dimensions[get_column_letter(COL2 - 1)].width = 3
    ws.freeze_panes = ws.cell(FILA_ENC + 1, 1)


def main() -> None:
    rutas.preparar_directorios()
    d1, d2 = cargar(1), cargar(2)

    wb = Workbook()
    wb.remove(wb.active)

    # --- Hoja 1: la clasificacion recibida, sin recalcular nada -------------
    b1 = bloque(d1, 1, d1["cluster"].to_numpy())
    b2 = bloque(d2, 2, d2["cluster"].to_numpy())
    escribir_hoja(wb, "1 UAM (base)", b1, b2,
                  "Los numeros de conglomerado son los del archivo original.")

    resumen = [("1 UAM (base)", len(set(b1['cluster'])), len(set(b2['cluster'])),
                "clasificacion recibida")]

    # --- Hojas 2 a 9: las especificaciones alternativas ---------------------
    for nombre, opc in ESPECIFICACIONES:
        cl1 = renumerar(clasificar(d1, **opc), d1["empleo"].to_numpy())
        cl2 = renumerar(clasificar(d2, **opc), d2["empleo"].to_numpy())
        escribir_hoja(wb, nombre, bloque(d1, 1, cl1), bloque(d2, 2, cl2),
                      f"k = {K} en los dos periodos; conglomerados numerados de mayor "
                      "a menor empleo.")
        ari1 = uam.rand_ajustado(d1["cluster"], cl1)
        ari2 = uam.rand_ajustado(d2["cluster"], cl2)
        resumen.append((nombre,
                        " / ".join(str(n) for n in pd.Series(cl1).value_counts().sort_index()),
                        " / ".join(str(n) for n in pd.Series(cl2).value_counts().sort_index()),
                        f"ARI contra la UAM: {ari1:.3f} (2003) · {ari2:.3f} (2023)"))

    # --- Hoja de lectura, al principio --------------------------------------
    ws = wb.create_sheet("0 Lectura", 0)
    ws.cell(1, 1, "T_I22 — la clasificacion en nueve especificaciones").font = Font(bold=True, size=13)
    ws.cell(2, 1, "62 actividades clasificadas en 2003 y 101 en 2023 (las que la UAM no excluyo). "
                  "Fuente: datos/intermediate/uam_replica_p{1,2}.csv, py/20.").font = Font(italic=True, size=9)
    for j, h in enumerate(["hoja", "tamanos 2003", "tamanos 2023", "nota"]):
        c = ws.cell(4, 1 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
    for i, fila in enumerate(resumen):
        for j, v in enumerate(fila):
            ws.cell(5 + i, 1 + j, v)
        ws.cell(5 + i, 1).font = Font(bold=True)
    for i, (nombre, _) in enumerate([("1 UAM (base)", None)] + ESPECIFICACIONES):
        ws.cell(5 + i, 6, DESCRIPCION[nombre])
    for col, w in zip("ABCDEF", (30, 18, 18, 42, 3, 110)):
        ws.column_dimensions[col].width = w

    wb.save(SALIDA)
    print(f"escrito: {SALIDA}")
    for f in resumen:
        print(f"  {f[0]:<30} 2003 {f[1]:<18} 2023 {f[2]:<18} {f[3]}")


if __name__ == "__main__":
    main()

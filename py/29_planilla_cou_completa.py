"""Bloque I.8 — la misma planilla de nueve especificaciones, pero SIN EXCLUIR
ninguna actividad: las 73 de 2003 y las 111 de 2023 de la COU completa.

Gemela de `py/28_planilla_especificaciones.py`, con una sola diferencia de
fondo: el universo. `py/28` clasifica las 62 y 101 actividades que la UAM dejo
adentro, heredando un criterio de exclusion que **no conocemos** —el cobre, los
combustibles, los servicios de vivienda y otras—. Aca se clasifica todo, incluida
la hoja base: es el ejercicio de la UAM tal cual, corrido sobre la matriz entera.

Las actividades que la UAM excluyo van marcadas en rojo y cursiva, para poder
ver de un vistazo en que conglomerado cae cada una.

Entrada:  datos/raw_local/UAM_Dutrenit/Chile_datos_Aug08_26_v1.xlsx  (el insumo)
          datos/intermediate/uam_replica_p{1,2}.csv  (solo para saber cuales excluyo)
Salida:   datos/output/2026.08.13 T_I23 clasificacion sin excluir sectores.xlsx
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402

_spec = importlib.util.spec_from_file_location(
    "uam_metodo", Path(__file__).resolve().parent / "40_uam_metodo.py"
)
uam = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(uam)

K = 4
# El corte con que la UAM reporto cada periodo. La hoja base lo respeta; las
# ocho alternativas van todas con k = 4.
K_UAM = {1: 4, 2: 2}
SALIDA = rutas.OUT / "2026.08.13 T_I23 clasificacion sin excluir sectores.xlsx"

ESPECIFICACIONES = [
    ("2 log",                       dict(log=True,  peso=False, balanza="nivel", sinfbcf=False)),
    ("3 peso empleo",               dict(log=False, peso=True,  balanza="nivel", sinfbcf=False)),
    ("4 balanza ajustada",          dict(log=False, peso=False, balanza="norm",  sinfbcf=False)),
    ("5 sin FBCF",                  dict(log=False, peso=False, balanza="nivel", sinfbcf=True)),
    ("6 log+peso",                  dict(log=True,  peso=True,  balanza="nivel", sinfbcf=False)),
    ("7 log+peso+balanza",          dict(log=True,  peso=True,  balanza="norm",  sinfbcf=False)),
    ("8 log+peso+sin FBCF",         dict(log=True,  peso=True,  balanza="nivel", sinfbcf=True)),
    ("9 log+peso+balanza+sin FBCF", dict(log=True,  peso=True,  balanza="norm",  sinfbcf=True)),
]

BASE = "1 UAM sin excluir"

DESCRIPCION = {
    BASE:
        "La especificacion de la UAM intacta —siete variables en niveles, sin ponderar, balanza "
        "comercial en niveles— corrida sobre la COU COMPLETA, con las actividades que ellos "
        "excluyeron adentro. k = 4 en 2003 y k = 2 en 2023, los cortes que ellos reportan.",
    "2 log":
        "Logaritmos en productividad y remuneracion media, log1p en impo/VBP y expo/VBP. "
        "FBCF por ocupado no admite logaritmo (tiene ceros y negativos) y queda en nivel.",
    "3 peso empleo":
        "Ponderacion por empleo: momentos ponderados en la estandarizacion y masas en el "
        "criterio de Ward, coste (Wa*Wb/(Wa+Wb))*||ca-cb||^2.",
    "4 balanza ajustada":
        "Balanza comercial normalizada (X-M)/(X+M) en vez de X-M en niveles. Es la unica "
        "de las siete que no era un ratio.",
    "5 sin FBCF":
        "Se saca x_capital_trabajo (FBCF por ocupado), que la COU asigna por producto y no "
        "mide intensidad de capital. Quedan seis variables.",
    "6 log+peso":
        "Logaritmos mas ponderacion por empleo. Balanza todavia en niveles.",
    "7 log+peso+balanza":
        "Logaritmos, ponderacion por empleo y balanza normalizada. Es la especificacion "
        "preferida del capitulo, y la unica donde el cobre no necesita exclusion.",
    "8 log+peso+sin FBCF":
        "Logaritmos y ponderacion por empleo, sin FBCF por ocupado. Balanza en niveles.",
    "9 log+peso+balanza+sin FBCF":
        "Las cuatro correcciones juntas: logaritmos, ponderacion por empleo, balanza "
        "normalizada y sin FBCF por ocupado. Seis variables.",
}

COLORES = ["DDEBF7", "FCE4D6", "E2EFDA", "FFF2CC", "E4DFEC", "F2F2F2", "DEEAF6", "FBE5D6"]
FINO = Side(style="thin", color="BFBFBF")
GRUESO = Side(style="medium", color="404040")
ROJO = "C00000"


def cargar(periodo: int) -> tuple[pd.DataFrame, set[str]]:
    """La COU completa del periodo, y el conjunto de codigos que la UAM excluyo."""
    d = pd.read_excel(rutas.exigir(rutas.D_UAM_INSUMO, "el insumo de la UAM"))
    d = d[d["periodo"] == periodo].copy()
    d["codigo_3dig"] = d["codigo_3dig"].astype(str)
    d = d.reset_index(drop=True)

    r = pd.read_csv(rutas.exigir(rutas.INTER / f"uam_replica_p{periodo}.csv",
                                 "la replica; correr antes 20_replica_hclust.py"))
    dentro = set(r["codigo_3dig"].astype(str))
    return d, set(d["codigo_3dig"]) - dentro


def clasificar(d: pd.DataFrame, k: int, log: bool, peso: bool, balanza: str,
               sinfbcf: bool) -> np.ndarray:
    """Las cuatro piezas se cruzan de forma independiente: logaritmar NO arrastra
    la balanza normalizada. Es la misma convencion de py/28."""
    x = uam.construir_variables(d, balanza=balanza)
    variables = [v for v in uam.VARS_UAM if not (sinfbcf and v == "x_capital_trabajo")]
    if log:
        x = uam.transformar_logs(x, variables)
    pesos = d["empleo"].to_numpy(dtype=float) if peso else None
    return uam.agrupar(x, variables, k, pesos)


def renumerar(cl: np.ndarray, empleo: np.ndarray) -> np.ndarray:
    orden = (pd.Series(empleo).groupby(pd.Series(cl)).sum()
             .sort_values(ascending=False).index.tolist())
    mapa = {c: i + 1 for i, c in enumerate(orden)}
    return np.array([mapa[c] for c in cl])


def bloque(d: pd.DataFrame, periodo: int, cl: np.ndarray, excluidas: set[str]) -> pd.DataFrame:
    b = pd.DataFrame({
        "pais": "CHILE",
        "periodo": periodo,
        "cluster": cl,
        "codigo_3dig": d["codigo_3dig"].to_numpy(),
        "desc": d["desc"].to_numpy(),
    })
    b["excluida"] = b["codigo_3dig"].isin(excluidas)
    b["_ord"] = pd.to_numeric(b["codigo_3dig"], errors="coerce")
    return b.sort_values(["cluster", "_ord"]).drop(columns="_ord").reset_index(drop=True)


# --------------------------------------------------------------------------
# Escritura
# --------------------------------------------------------------------------

ENCABEZADOS = ["pais", "periodo", "cluster", "codigo_3dig", "desc"]
COL1, COL2 = 1, 7
FILA_TITULO, FILA_ENC = 4, 5


def escribir_bloque(ws, b: pd.DataFrame, col0: int, titulo: str) -> None:
    ws.cell(FILA_TITULO, col0, titulo).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=FILA_TITULO, start_column=col0,
                   end_row=FILA_TITULO, end_column=col0 + 4)
    ws.cell(FILA_TITULO, col0).alignment = Alignment(horizontal="center")

    for j, h in enumerate(ENCABEZADOS):
        c = ws.cell(FILA_ENC, col0 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)

    anterior = None
    for i, fila in enumerate(b.itertuples(index=False)):
        r = FILA_ENC + 1 + i
        nuevo = fila.cluster != anterior
        relleno = PatternFill("solid", fgColor=COLORES[(fila.cluster - 1) % len(COLORES)])
        for j, v in enumerate(fila[:5]):
            c = ws.cell(r, col0 + j, v)
            c.fill = relleno
            c.border = Border(left=FINO, right=FINO,
                              top=GRUESO if nuevo else FINO, bottom=FINO)
            if fila.excluida:
                c.font = Font(color=ROJO, italic=True, bold=(j == 2))
            elif j == 2:
                c.font = Font(bold=True)
            if j in (1, 2, 3):
                c.alignment = Alignment(horizontal="center")
        anterior = fila.cluster

    ultima = FILA_ENC + len(b)
    for j in range(5):
        ws.cell(ultima, col0 + j).border = Border(left=FINO, right=FINO,
                                                  top=FINO, bottom=GRUESO)

    r = ultima + 2
    for j, h in enumerate(("conglomerado", "actividades", "de ellas, excluidas por la UAM")):
        ws.cell(r, col0 + j, h).font = Font(bold=True, italic=True)
    conteo = b.groupby("cluster").agg(n=("cluster", "size"), excl=("excluida", "sum"))
    for i, (cl, fila) in enumerate(conteo.iterrows()):
        ws.cell(r + 1 + i, col0, int(cl)).fill = PatternFill(
            "solid", fgColor=COLORES[(int(cl) - 1) % len(COLORES)])
        ws.cell(r + 1 + i, col0).font = Font(bold=True)
        ws.cell(r + 1 + i, col0 + 1, int(fila["n"]))
        c = ws.cell(r + 1 + i, col0 + 2, int(fila["excl"]))
        if fila["excl"]:
            c.font = Font(color=ROJO, italic=True)


def escribir_hoja(wb: Workbook, nombre: str, b1: pd.DataFrame, b2: pd.DataFrame,
                  nota_k: str) -> None:
    ws = wb.create_sheet(nombre)
    ws.cell(1, 1, nombre).font = Font(bold=True, size=13)
    ws.cell(2, 1, DESCRIPCION[nombre] + " " + nota_k).font = Font(italic=True, size=9)
    ws.cell(3, 1, "En rojo y cursiva, las actividades que la UAM excluyo de su ejercicio "
                  "(11 en 2003, 10 en 2023).").font = Font(italic=True, size=9, color=ROJO)

    escribir_bloque(ws, b1, COL1, "Periodo 1 — 2003 (73 actividades)")
    escribir_bloque(ws, b2, COL2, "Periodo 2 — 2023 (111 actividades)")

    for col0 in (COL1, COL2):
        for off, w in zip(range(5), (8, 8, 8, 11, 46)):
            ws.column_dimensions[get_column_letter(col0 + off)].width = w
    ws.column_dimensions[get_column_letter(COL2 - 1)].width = 3
    ws.freeze_panes = ws.cell(FILA_ENC + 1, 1)


def main() -> None:
    rutas.preparar_directorios()
    d1, ex1 = cargar(1)
    d2, ex2 = cargar(2)
    print(f"2003: {len(d1)} actividades, {len(ex1)} excluidas por la UAM")
    print(f"2023: {len(d2)} actividades, {len(ex2)} excluidas por la UAM")

    wb = Workbook()
    wb.remove(wb.active)
    resumen = []

    def agregar(nombre, cl1, cl2, nota_k):
        b1 = bloque(d1, 1, cl1, ex1)
        b2 = bloque(d2, 2, cl2, ex2)
        escribir_hoja(wb, nombre, b1, b2, nota_k)
        # ¿Cuantos conglomerados gasta la especificacion en grupos minusculos?
        def diagnostico(b, d):
            emp = pd.Series(d["empleo"].to_numpy(), index=d["codigo_3dig"])
            e = b.groupby("cluster")["codigo_3dig"].apply(lambda s: emp[s].sum())
            e = 100 * e / emp.sum()
            return (" / ".join(str(n) for n in b["cluster"].value_counts().sort_index()),
                    " / ".join(f"{v:.1f}" for v in e.sort_index()),
                    int((e < 1.0).sum()))
        t1, p1, m1 = diagnostico(b1, d1)
        t2, p2, m2 = diagnostico(b2, d2)
        resumen.append((nombre, t1, p1, m1, t2, p2, m2))

    # --- Hoja 1: la especificacion de la UAM sobre la COU completa ----------
    agregar(BASE,
            renumerar(clasificar(d1, K_UAM[1], log=False, peso=False,
                                 balanza="nivel", sinfbcf=False), d1["empleo"].to_numpy()),
            renumerar(clasificar(d2, K_UAM[2], log=False, peso=False,
                                 balanza="nivel", sinfbcf=False), d2["empleo"].to_numpy()),
            f"k = {K_UAM[1]} en 2003 y k = {K_UAM[2]} en 2023; conglomerados numerados de "
            "mayor a menor empleo.")

    # --- Hojas 2 a 9 --------------------------------------------------------
    for nombre, opc in ESPECIFICACIONES:
        agregar(nombre,
                renumerar(clasificar(d1, K, **opc), d1["empleo"].to_numpy()),
                renumerar(clasificar(d2, K, **opc), d2["empleo"].to_numpy()),
                f"k = {K} en los dos periodos; conglomerados numerados de mayor a menor empleo.")

    # --- Hoja de lectura ----------------------------------------------------
    ws = wb.create_sheet("0 Lectura", 0)
    ws.cell(1, 1, "T_I23 — la clasificacion en nueve especificaciones, SIN EXCLUIR sectores"
            ).font = Font(bold=True, size=13)
    ws.cell(2, 1, "Las 73 actividades de 2003 y las 111 de 2023 de la COU completa, incluidas el "
                  "cobre y las demas que la UAM dejo fuera con un criterio que no conocemos. "
                  "Fuente: Chile_datos_Aug08_26_v1.xlsx. Gemela de T_I22, que clasifica solo las "
                  "62 y 101 que ellos dejaron adentro.").font = Font(italic=True, size=9)
    ws.cell(3, 1, "La columna 'grupos < 1 % del empleo' cuenta los conglomerados que la "
                  "especificacion gasta en actividades marginales: es el sintoma del problema "
                  "que la exclusion pretendia resolver.").font = Font(italic=True, size=9)
    enc = ["hoja", "tamanos 2003", "% empleo 2003", "grupos <1% 2003",
           "tamanos 2023", "% empleo 2023", "grupos <1% 2023", "", "que hace"]
    for j, h in enumerate(enc):
        c = ws.cell(5, 1 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
    for i, fila in enumerate(resumen):
        for j, v in enumerate(fila):
            ws.cell(6 + i, 1 + j, v)
        ws.cell(6 + i, 1).font = Font(bold=True)
        ws.cell(6 + i, 9, DESCRIPCION[fila[0]])
    for col, w in zip("ABCDEFGHI", (30, 20, 24, 15, 20, 24, 15, 3, 110)):
        ws.column_dimensions[col].width = w

    wb.save(SALIDA)
    print(f"\nescrito: {SALIDA}\n")
    for f in resumen:
        print(f"  {f[0]:<30} 2003 {f[1]:<18} [{f[2]:<22}] <1%: {f[3]}"
              f"   2023 {f[4]:<18} [{f[5]:<22}] <1%: {f[6]}")


if __name__ == "__main__":
    main()

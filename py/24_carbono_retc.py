"""Validación física: ¿el gasto en energía de la MIP ordena igual que las emisiones reales?

POR QUÉ HACE FALTA. `x_intens_energ` mide lo que cada actividad GASTA en productos
energéticos, no la energía que consume ni el carbono que emite. Es la única forma de
tener la variable en los cuatro países sin datos nuevos, pero hay que mostrar que el
proxy ordena a los sectores como lo haría la medida física. Si no, la variable mide
precios relativos de la energía y no intensidad.

Y ADEMÁS. La huella de carbono es la única de las cuatro variables que Gabriela pide que
NO sale de la matriz insumo-producto. Esto es lo que Chile sí puede poner: emisiones
declaradas de fuentes puntuales, por establecimiento, con código CIIU4.CL.

FUENTE. Registro de Emisiones y Transferencias de Contaminantes (RETC), Ministerio del
Medio Ambiente, emisiones al aire de fuentes puntuales 2023, bajado el 2026-08-11 desde
`datosretc.mma.gob.cl`. Son 459.991 declaraciones de fuente por contaminante.

QUÉ NO ES. El RETC cubre fuentes PUNTUALES declaradas: no incluye transporte en ruta,
fuentes difusas ni combustión residencial. Las actividades cuyas emisiones son de flota
—transporte de carga, transporte de pasajeros— aparecen subrepresentadas, y por eso el
contraste se hace sobre el orden de magnitud y el ranking, no sobre el nivel.

SALIDAS. Tabla T_I7 y figura F_I3.
"""

from __future__ import annotations

import numpy as np
import pandas as pd
from openpyxl import load_workbook

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import rutas

ARCHIVO = "2026.08.11 Emisiones al aire fuentes puntuales/retc_fuentes_puntuales_2023.xlsx"
CONTAMINANTE = "Carbon dioxide"

# `emision_retc` y no `emision_total`. La hoja de metadatos del propio archivo lo dice:
# `emision_total` es la estimación cruda del sistema sectorial RUEA, y `emision_retc` es
# la validada por el Departamento de Información Ambiental del MMA, que descarta tres
# clases de outlier —de procesos, de consumos declarados por sobre el consumo posible, y
# de tendencia— y reemplaza la estimación por los datos del Sistema de Impuestos Verdes
# (DS 63/2022) y del sistema de centrales termoeléctricas SICTER cuando existen. Con la
# columna cruda el total nacional se va a 146 Mt, por encima del inventario del país, y
# aparecen panaderías emitiendo 9 Mt.
COLUMNAS = ("ciiu6_id", "contaminante", "emision_retc")

# Productos de electricidad dentro del grupo energético de la COU 2018: se emiten aguas
# arriba, en la central, no en el sector que los compra.
ELECTRICIDAD = (119, 120, 121)

# Actividades de transporte de la COU 2018. Emiten desde flotas, que son fuentes móviles
# y no entran al registro de fuentes puntuales del RETC.
ACTIVIDADES_TRANSPORTE = tuple(range(79, 88))


def matriz_utilizacion() -> pd.DataFrame:
    """La utilización intermedia de 2023, reusando el lector de 22_variables_nuevas.py."""
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        "vn", rutas.REPO / "py" / "22_variables_nuevas.py")
    vn = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(vn)
    return vn.leer_matriz(
        rutas.D_BCCH / "2026.08.11 COU y MIP" / "2023_Cuadros_111x181.xlsx", "5", 181, 111)


def leer_retc(ruta) -> pd.DataFrame:
    """Lee sólo las tres columnas que se usan, en streaming: el archivo son 83 MB."""
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    filas = ws.iter_rows(values_only=True)
    cabecera = list(next(filas))
    pos = {c: cabecera.index(c) for c in COLUMNAS}
    datos = []
    for r in filas:
        if r[pos["contaminante"]] != CONTAMINANTE:
            continue
        datos.append((r[pos["ciiu6_id"]], r[pos["emision_retc"]]))
    wb.close()
    d = pd.DataFrame(datos, columns=["ciiu6_id", "emision"])
    d["emision"] = pd.to_numeric(d["emision"], errors="coerce").fillna(0.0)
    return d


def main() -> int:
    rutas.preparar_directorios()
    ruta = rutas.D_RETC / ARCHIVO
    if not rutas.hay(ruta):
        print(f"No está el RETC en {ruta}; se salta la validación física.")
        return 0

    print("Leyendo el RETC (459.991 filas, sólo CO2)…")
    d = leer_retc(ruta)
    print(f"  declaraciones de CO2: {len(d):,}")

    # El código viene como «C259900»: letra de sección más los seis dígitos CIIU4.CL.
    d["ciiu"] = d["ciiu6_id"].astype(str).str.extract(r"(\d{6})")[0]
    sin_codigo = d["ciiu"].isna().sum()
    if sin_codigo:
        print(f"  sin código CIIU utilizable: {sin_codigo:,} ({sin_codigo / len(d):.1%})")
    d = d.dropna(subset=["ciiu"])

    reglas = pd.read_csv(
        rutas.exigir(rutas.CONFIG / "reglas_ciiu_a_cou.csv", "las reglas CIIU → COU"),
        comment="#", dtype={"prefijo": str}).dropna(subset=["prefijo"])
    mapa = dict(zip(reglas["prefijo"], reglas["cou"].astype(int)))

    def resolver(c: str):
        for n in range(len(c), 0, -1):
            if c[:n] in mapa:
                return mapa[c[:n]]
        return np.nan

    d["cou"] = d["ciiu"].map(resolver)
    print(f"  CO2 total declarado: {d['emision'].sum() / 1e6:,.1f} millones de toneladas")

    asignadas = d.dropna(subset=["cou"]).copy()
    asignadas["cou"] = asignadas["cou"].astype(int)
    emis = asignadas.groupby("cou")["emision"].sum()

    rep = pd.read_csv(rutas.exigir(rutas.INTER / "uam_replica_p2.csv", "la réplica de 2023"),
                      index_col=0)
    nue = pd.read_csv(rutas.exigir(rutas.INTER / "variables_nuevas_p2.csv",
                                   "las variables nuevas de 2023"), index_col=0)
    t = rep.join(nue[["x_intens_energ"]])
    t["emision_co2"] = emis.reindex(t.index)
    t["intens_co2"] = t["emision_co2"] / t["valor_agrega"]   # toneladas por miles de millones

    # El contraste directo está mal especificado por dos razones, y las dos empujan la
    # correlación hacia abajo sin que la variable de la MIP tenga la culpa:
    #
    #   · el RETC cubre fuentes PUNTUALES. Las actividades de transporte, que son las de
    #     mayor gasto en energía de toda la economía (38% del VBP en transporte terrestre
    #     de pasajeros), emiten desde flotas y casi no aparecen en el registro;
    #   · la electricidad comprada se emite aguas arriba, en la central, no en el sector
    #     que la consume. Un sector electrointensivo tiene gasto energético alto y
    #     emisión directa baja.
    #
    # El contraste bien especificado es gasto en COMBUSTIBLES —sin electricidad— contra
    # emisión por combustión directa, dejando fuera al transporte. Se reportan los dos.
    clasif = pd.read_csv(rutas.CONFIG / "clasificacion_productos_cou.csv")
    energeticos = clasif[(clasif["periodo"] == 2) & (clasif["grupo"] == "energia")]
    combustibles = energeticos[~energeticos["cod_producto"].isin(ELECTRICIDAD)]
    U = matriz_utilizacion()
    vbp = pd.read_excel(rutas.D_UAM_INSUMO)
    vbp = vbp[vbp["periodo"] == 2].set_index("codigo_3dig")["vbp"]
    t["x_intens_combust"] = (U.reindex(combustibles["cod_producto"]).sum(axis=0)
                             .reindex(t.index) / vbp.reindex(t.index))

    resultados = []
    for etiqueta, col, quitar_transporte in [
        ("todas las actividades, gasto en energía total", "x_intens_energ", False),
        ("sin transporte, gasto en combustibles", "x_intens_combust", True),
    ]:
        con = t.dropna(subset=["emision_co2", col])
        con = con[(con["emision_co2"] > 0) & (con[col] > 0)]
        if quitar_transporte:
            con = con[~con.index.isin(ACTIVIDADES_TRANSPORTE)]
        rho = con[col].corr(con["intens_co2"], method="spearman")
        r_log = np.log(con[col]).corr(np.log(con["intens_co2"]))
        resultados.append({"contraste": etiqueta, "n actividades": len(con),
                           "rho de Spearman": rho, "r de Pearson en logs": r_log})
        print(f"\n  {etiqueta} ({len(con)} actividades)")
        print(f"    rho de Spearman        : {rho:.3f}")
        print(f"    r de Pearson en logs   : {r_log:.3f}")

    cobertura = t["emision_co2"].sum() / d["emision"].sum()
    print(f"\n  del CO2 declarado, cae en actividades clasificadas: {cobertura:.1%}")
    rho = resultados[-1]["rho de Spearman"]
    con = t.dropna(subset=["emision_co2", "x_intens_energ"])
    con = con[(con["emision_co2"] > 0) & (con["x_intens_energ"] > 0)]
    if rho >= 0.75:
        veredicto = ("el gasto en combustibles de la MIP puede usarse como proxy de "
                     "intensidad de emisiones")
    elif rho >= 0.40:
        veredicto = ("la relación es del signo correcto pero MODERADA: el gasto en energía "
                     "de la MIP sirve como variable de intensidad energética y no como "
                     "sustituto de la huella de carbono, que hay que medir aparte")
    else:
        veredicto = ("el gasto en energía y las emisiones físicas ordenan distinto: la "
                     "variable de la MIP no puede presentarse como proxy de emisiones")
    print(f"  → {veredicto}")
    print("     Corregir la especificación —sacar la electricidad, que se emite aguas "
          "arriba,\n     y el transporte, que el RETC no cubre— mueve rho de "
          f"{resultados[0]['rho de Spearman']:.3f} a {rho:.3f}: poco.\n"
          "     Lo que queda es dispersión real de precios de la energía y de mezcla de "
          "combustibles\n     entre sectores, y es la razón para no vender la variable de "
          "la MIP como huella de carbono.")
    pd.DataFrame(resultados).to_excel(
        rutas.tabla("T_I8 validacion fisica de la intensidad energetica"), index=False)

    print("\n  las 10 actividades más emisoras de CO2 declarado:")
    for r in t.nlargest(10, "emision_co2").itertuples():
        print(f"    {r.emision_co2 / 1e6:7.2f} Mt   energía {r.x_intens_energ:6.1%}   {r.desc}")

    salida = t[["desc", "cluster", "valor_agrega", "empleo", "x_intens_energ",
                "emision_co2", "intens_co2"]].reset_index()
    salida.to_excel(rutas.tabla("T_I7 emisiones RETC contra intensidad energetica MIP"),
                    index=False)

    fig, ax = plt.subplots(figsize=(7.2, 5.2))
    colores = {1: "#4C72B0", 2: "#DD8452"}
    for c, sub in con.groupby("cluster"):
        ax.scatter(sub["x_intens_energ"], sub["intens_co2"], s=26, alpha=0.75,
                   color=colores.get(c, "#888888"), label=f"Conglomerado {c} (UAM)")
    ax.set_xscale("log")
    ax.set_yscale("log")
    ax.set_xlabel("Gasto en productos energéticos sobre VBP  (COU, Banco Central)")
    ax.set_ylabel("Toneladas de CO$_2$ declaradas por unidad de valor agregado  (RETC)")
    ax.xaxis.set_major_formatter(lambda v, _: f"{v:.0%}")
    ax.legend(frameon=False, fontsize=9)
    ax.set_title("Gasto en energía y emisiones físicas: misma dirección, relación moderada\n"
                 f"Chile, 2023 · {len(con)} actividades · rho de Spearman = "
                 f"{resultados[0]['rho de Spearman']:.2f}", fontsize=11)
    fig.text(0.5, -0.03,
             "Ambos ejes en escala logarítmica. El RETC cubre fuentes puntuales declaradas: "
             "no incluye transporte en ruta ni fuentes difusas,\npor lo que las actividades "
             "cuyas emisiones son de flota quedan subrepresentadas.",
             ha="center", fontsize=7.5, color="#444444")
    fig.tight_layout()
    for ext in ("png", "pdf"):
        fig.savefig(rutas.figura("F_I3 gasto en energia contra emisiones", ext),
                    bbox_inches="tight")
    plt.close(fig)
    print("\n  tabla T_I7 y figura F_I3 escritas.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

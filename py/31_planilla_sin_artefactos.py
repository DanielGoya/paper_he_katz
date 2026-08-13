"""Bloque I.9 — la planilla de nueve especificaciones sobre la COU sin las cuatro
actividades que son artefactos contables, y con el k ELEGIDO EN CADA HOJA.

Tercera de la serie. Se diferencia de las dos anteriores en dos cosas:

1. **El universo.** No es ni el de la UAM (62 y 101 actividades, con un criterio
   de exclusion que no conocemos) ni la COU completa: son las 69 de 2003 y las
   107 de 2023 que quedan al sacar solo cuatro —educacion privada, propiedad /
   servicios de vivienda, elaboracion de combustible y tabaco—. Las cuatro tienen
   la misma justificacion, y es contable, no estadistica: alquiler imputado,
   impuesto especifico y un empleo mal medido producen valores que no miden lo
   que dicen. El cobre y el petroleo SE QUEDAN: son enclaves reales.

2. **El numero de conglomerados.** En T_I22 y T_I23 el k = 4 estaba impuesto,
   heredado de dos criterios calculados una sola vez sobre la especificacion
   preferida de 2023. Aca los dos criterios se recalculan en cada hoja y en cada
   periodo, y el k cambia entre hojas si los datos lo piden:

       salto de costes de Ward   cuanto mas caro es bajar de k+1 a k grupos que
                                 bajar de k+2 a k+1. No hay codo en estas curvas,
                                 asi que lo que discrimina es el salto.
       eta^2                     fraccion de la varianza del log de la
                                 productividad, ponderada por empleo, que explica
                                 la particion. Donde mas sube, empieza la meseta.

   Manda el salto de costes, que es interno al algoritmo; el eta^2 va como
   segunda opinion y su desacuerdo se reporta en la hoja `00 Eleccion de k`.

Entrada:  datos/raw_local/UAM_Dutrenit/Chile_datos_Aug08_26_v1.xlsx
          datos/intermediate/uam_replica_p{1,2}.csv  (solo para marcar en rojo
                                                      las que la UAM excluia)
Salida:   datos/output/2026.08.13 T_I24 clasificacion sin artefactos contables.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402
import planilla_clusters as pc  # noqa: E402

KMAX = 8
SALIDA = rutas.OUT / "2026.08.13 T_I24 clasificacion sin artefactos contables.xlsx"

BASE = "1 UAM (misma especificacion)"

# Las cuatro que se sacan, por palabra clave sobre la descripcion normalizada.
# La nomenclatura cambia entre 2003 y 2023 ("propiedad de vivienda" pasa a
# "servicios de vivienda", "combustible" a "combustibles"), asi que se busca por
# contenido y se verifica que sean exactamente cuatro en cada periodo.
FUERA = {
    "educacion privada":    ("educacion privada",),
    "vivienda":             ("propiedad de vivienda", "servicios de vivienda"),
    "combustibles":         ("elaboracion de combustible",),
    "tabaco":               ("tabaco",),
}


def cargar(periodo: int) -> tuple[pd.DataFrame, set[str], pd.DataFrame]:
    """La COU del periodo sin las cuatro artefactuales, los codigos que la UAM
    excluia (para marcarlos) y las cuatro que se sacaron."""
    d = pd.read_excel(rutas.exigir(rutas.D_UAM_INSUMO, "el insumo de la UAM"))
    d = d[d["periodo"] == periodo].copy()
    d["codigo_3dig"] = d["codigo_3dig"].astype(str)
    d["_norm"] = d["desc"].map(rutas.normalizar)

    sacar = pd.Series(False, index=d.index)
    for _, patrones in FUERA.items():
        hit = d["_norm"].apply(lambda t: any(p in t for p in patrones))
        if hit.sum() != 1:
            raise SystemExit(
                f"ERROR periodo {periodo}: los patrones {patrones} calzan con "
                f"{hit.sum()} actividades, no con 1: {list(d.loc[hit, 'desc'])}")
        sacar |= hit

    fuera = d[sacar][["codigo_3dig", "desc"]].copy()
    d = d[~sacar].drop(columns="_norm").reset_index(drop=True)

    r = pd.read_csv(rutas.exigir(rutas.INTER / f"uam_replica_p{periodo}.csv",
                                 "la replica; correr antes 20_replica_hclust.py"))
    marcadas = set(d["codigo_3dig"]) - set(r["codigo_3dig"].astype(str))
    return d, marcadas, fuera


def main() -> None:
    rutas.preparar_directorios()
    d1, m1, f1 = cargar(1)
    d2, m2, f2 = cargar(2)
    print(f"2003: {len(d1)} actividades (73 - 4). Todavia excluidas por la UAM: {len(m1)}")
    print(f"2023: {len(d2)} actividades (111 - 4). Todavia excluidas por la UAM: {len(m2)}")
    print("  fuera 2003:", ", ".join(f1["desc"]))
    print("  fuera 2023:", ", ".join(f2["desc"]))

    wb = Workbook()
    wb.remove(wb.active)
    resumen, curvas = [], []

    hojas = [(BASE, dict(log=False, peso=False, balanza="nivel", sinfbcf=False))]
    hojas += pc.ESPECIFICACIONES

    for nombre, opc in hojas:
        elec = {p: pc.elegir_k(d, kmax=KMAX, **opc) for p, d in ((1, d1), (2, d2))}
        bs, diag = [], []
        for p, d, marcadas, anio in ((1, d1, m1, 2003), (2, d2, m2, 2023)):
            k = elec[p]["k"]
            cl = pc.renumerar(pc.clasificar(d, k, **opc), d["empleo"].to_numpy())
            b = pc.bloque(d, p, cl, marcadas)
            bs.append((b, f"Periodo {p} — {anio} ({len(d)} actividades, k = {k})"))
            diag.append(pc.diagnostico(b, d))
            for kk in range(2, KMAX + 1):
                curvas.append({"hoja": nombre, "anio": anio, "k": kk,
                               "salto_costes": elec[p]["salto"].get(kk),
                               "eta2": elec[p]["eta2"].get(kk),
                               "elegido": "<-- k" if kk == k else ""})

        desc = (pc.DESCRIPCION[nombre] if nombre in pc.DESCRIPCION else
                "La especificacion de la UAM intacta: siete variables en niveles, sin "
                "ponderar, balanza comercial en niveles.")
        notas = [
            (desc, "000000"),
            (f"k elegido por el salto de costes de Ward: {elec[1]['k']} en 2003 "
             f"(salto {elec[1]['salto'].max():.2f}) y {elec[2]['k']} en 2023 "
             f"(salto {elec[2]['salto'].max():.2f}). El eta^2 sugiere {elec[1]['k_eta2']} y "
             f"{elec[2]['k_eta2']}. Un salto cercano a 1 significa que la curva de costes es "
             f"plana y el k no esta identificado. Conglomerados numerados de mayor a menor empleo.",
             "000000"),
            ("En rojo y cursiva, las actividades que la UAM excluia y aca siguen adentro "
             "(el cobre entre ellas).", pc.ROJO),
        ]
        pc.escribir_hoja(wb, nombre, notas, bs, con_marca=True)
        resumen.append((nombre, elec[1]["k"], round(float(elec[1]["salto"].max()), 2),
                        elec[1]["k_eta2"], *diag[0],
                        elec[2]["k"], round(float(elec[2]["salto"].max()), 2),
                        elec[2]["k_eta2"], *diag[1]))

    # --- Hoja de eleccion de k ----------------------------------------------
    ws = wb.create_sheet("00 Eleccion de k")
    ws.cell(1, 1, "Los dos criterios, recalculados en cada hoja y cada periodo"
            ).font = Font(bold=True, size=13)
    ws.cell(2, 1, "salto_costes: cuanto mas caro es bajar de k+1 a k grupos que bajar de k+2 a "
                  "k+1. Manda este. | eta2: fraccion de la varianza del log de la productividad, "
                  "ponderada por empleo, que explica la particion; se elige donde mas sube."
            ).font = Font(italic=True, size=9)
    cur = pd.DataFrame(curvas)
    for j, h in enumerate(cur.columns):
        c = ws.cell(4, 1 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
    for i, fila in enumerate(cur.itertuples(index=False)):
        for j, v in enumerate(fila):
            c = ws.cell(5 + i, 1 + j, v)
            if isinstance(v, float):
                c.number_format = "0.000"
            if fila.elegido:
                c.font = Font(bold=True)
    for col, w in zip("ABCDEF", (30, 8, 6, 14, 10, 10)):
        ws.column_dimensions[col].width = w

    # --- Hoja de lectura ----------------------------------------------------
    ws = wb.create_sheet("0 Lectura", 0)
    ws.cell(1, 1, "T_I24 — sin los cuatro artefactos contables, con el k elegido en cada hoja"
            ).font = Font(bold=True, size=13)
    ws.cell(2, 1, f"69 actividades en 2003 y 107 en 2023: la COU completa menos "
                  f"{', '.join(f2['desc'])}. El cobre y el petroleo se quedan adentro."
            ).font = Font(italic=True, size=9)
    ws.cell(3, 1, "Tercera de la serie. T_I22 clasifica solo las 62 y 101 que la UAM dejo "
                  "adentro; T_I23, la COU completa. Las dos con k = 4 impuesto."
            ).font = Font(italic=True, size=9)
    enc = ["hoja", "k 2003", "salto", "(eta2)", "tamanos 2003", "% empleo 2003", "<1% 2003",
           "k 2023", "salto", "(eta2)", "tamanos 2023", "% empleo 2023", "<1% 2023",
           "", "que hace"]
    for j, h in enumerate(enc):
        c = ws.cell(5, 1 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
    for i, fila in enumerate(resumen):
        for j, v in enumerate(fila):
            ws.cell(6 + i, 1 + j, v)
        ws.cell(6 + i, 1).font = Font(bold=True)
        ws.cell(6 + i, 15, pc.DESCRIPCION.get(fila[0], "especificacion de la UAM intacta"))
        # Marcar en rojo donde los dos criterios no coinciden.
        for cc, (ka, ke) in ((2, (fila[1], fila[3])), (8, (fila[7], fila[9]))):
            if ka != ke:
                ws.cell(6 + i, cc).font = Font(bold=True, color=pc.ROJO)
                ws.cell(6 + i, cc + 2).font = Font(italic=True, color=pc.ROJO)
    for col, w in zip("ABCDEFGHIJKLMNO",
                      (30, 8, 8, 8, 20, 26, 9, 8, 8, 8, 20, 26, 9, 3, 100)):
        ws.column_dimensions[col].width = w

    wb.save(SALIDA)
    print(f"\nescrito: {SALIDA}\n")
    for f in resumen:
        print(f"  {f[0]:<30} 2003 k={f[1]} (salto {f[2]:.2f}, eta2 {f[3]}) {f[4]:<16}"
              f" [{f[5]:<22}] <1%: {f[6]}"
              f"   2023 k={f[7]} (salto {f[8]:.2f}, eta2 {f[9]}) {f[10]:<16}"
              f" [{f[11]:<22}] <1%: {f[12]}")


if __name__ == "__main__":
    main()

"""Bloque I.11 — la reparametrizacion del bloque de comercio, sobre la replica de
la UAM con el cobre adentro.

LA PREGUNTA (DG, 2026-08-13). Las siete variables de la UAM traen TRES columnas
de comercio: impo/VBP, expo/VBP y la balanza. Con la balanza corregida —la
mejora que el proyecto ya adopto— la tercera es funcion EXACTA de las otras dos:
con o = X/VBP y d = M/VBP el VBP se cancela y (X-M)/(X+M) == (o-d)/(o+d). La
variante (X-M)/VBP es peor: es exactamente o - d, o sea el bloque tiene rango 2
y no 3. Pero estandarizar iguala la varianza de cada columna, asi que el
comercio se lleva 3/7 = 42,9 % de la inercia por una direccion contada dos
veces. De ahi el hallazgo del bloque K.4: sacar un coeficiente de comercio mueve
la particion mas que sacar la productividad, que es el objeto declarado del
ejercicio.

LA SALIDA NO ES PODAR SINO REPARAMETRIZAR. (o, d) -> (apertura, balanza) es
biyectiva —o = a(1+b)/2, d = a(1-b)/2— asi que no se pierde informacion, y
separa los dos conceptos que hoy estan superpuestos:

    x_apertura       (X+M)/VBP     cuanto comercia la actividad
    x_balanza_norm   (X-M)/(X+M)   en que direccion, acotado en [-1, 1]

Seis variables en vez de siete, y el comercio pasa de 42,9 % a 2/6 = 33,3 % de
la inercia, ya sin redundancia. Ambas son ratios, asi que la invariancia exacta
a la desagregacion se conserva: se verifica con la prueba de clones.

EL UNIVERSO: LA REPLICA DE LA UAM, MENOS SU EXCLUSION DEL COBRE. Se respeta el
criterio de exclusion de la UAM —el que no conocemos y no es inferible— con una
sola excepcion, la que el capitulo discute: el cobre vuelve a entrar. Quedan 63
actividades en 2003 (62 + cobre) y 102 en 2023 (101 + cobre). Es el universo que
aisla la pregunta: todo lo demas identico a ellos, y el cobre adentro para ver
si la reparametrizacion lo absorbe sin que haya que excluirlo.

LAS SEIS ESPECIFICACIONES, acumulativas salvo la cuarta:

    1 UAM base                              las siete, sin nada
    2 comercio                              + reparametrizacion
    3 comercio + peso                       + ponderacion por empleo
    4 comercio + logs                       reparametrizacion + logs, SIN peso
    5 comercio + peso + logs                las tres juntas
    6 comercio + peso + logs + sin FBCF     + sacar la FBCF por ocupado

La 4 esta a proposito fuera de la escalera: es la que mide si sigue valiendo el
resultado propio del proyecto de que logaritmar SIN ponderar empeora la
invariancia (ARI -0,086 en 2003 k=2 con la parametrizacion vieja).

EL k. Impuesto en 4 en las hojas de clasificacion, para que las seis sean
comparables entre si —igual que T_I22 y T_I23—, y los dos criterios (salto de
costes de Ward y eta^2 del log de la productividad) recalculados y reportados
en la hoja `00 Eleccion de k`. Donde discrepan con 4, se dice.

SALIDA. `T_I26`, con seis hojas de clasificacion en el formato de los libros 3a
y 3b de la UAM —los dos periodos lado a lado, y al pie de cada una las medias
por conglomerado en niveles y los centroides en puntajes z— mas cinco hojas de
diagnostico.

DEPENDE DE. `40_uam_metodo.py` y `planilla_clusters.py`.
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402
import planilla_clusters as pc  # noqa: E402

uam = pc.uam

FECHA = "2026.08.13"
SALIDA = rutas.OUT / f"{FECHA} T_I26 reparametrizacion del comercio con el cobre adentro.xlsx"

K = 4          # impuesto en las hojas de clasificacion, por comparabilidad
KMAX = 8

# Las columnas que se dividen al partir una actividad en clones: son magnitudes,
# no ratios. `empleo` incluido, que es el peso.
SUMABLES = ["vbp", "valor_agrega", "remunera", "empleo", "fbcf", "expo", "impo"]

ESPECS = [
    ("1 UAM base",
     dict(log=False, peso=False, balanza="nivel", sinfbcf=False, comercio=False)),
    ("2 comercio",
     dict(log=False, peso=False, balanza="norm", sinfbcf=False, comercio=True)),
    ("3 comercio + peso",
     dict(log=False, peso=True, balanza="norm", sinfbcf=False, comercio=True)),
    ("4 comercio + logs",
     dict(log=True, peso=False, balanza="norm", sinfbcf=False, comercio=True)),
    ("5 comercio + peso + logs",
     dict(log=True, peso=True, balanza="norm", sinfbcf=False, comercio=True)),
    # El nombre es el de la hoja de Excel, que no admite mas de 31 caracteres.
    ("6 idem 5, sin FBCF",
     dict(log=True, peso=True, balanza="norm", sinfbcf=True, comercio=True)),
]

QUE_HACE = {
    "1 UAM base":
        "Las siete variables de la UAM, sin ponderar, en niveles, con la balanza comercial "
        "en niveles. Es su especificacion intacta; lo unico distinto es el universo (el "
        "cobre esta adentro).",
    "2 comercio":
        "Reemplaza las TRES columnas de comercio por DOS: apertura (X+M)/VBP y balanza "
        "normalizada (X-M)/(X+M). La transformacion es biyectiva, asi que no se pierde "
        "informacion; lo que se quita es que una direccion se contara dos veces.",
    "3 comercio + peso":
        "Mas ponderacion por empleo: momentos ponderados en la estandarizacion y masas en "
        "el criterio de Ward, coste (Wa*Wb/(Wa+Wb))*||ca-cb||^2.",
    "4 comercio + logs":
        "Reparametrizacion mas logaritmos, SIN ponderar. Fuera de la escalera a proposito: "
        "mide si sigue valiendo que logaritmar sin ponderar empeora la invariancia.",
    "5 comercio + peso + logs":
        "Las tres correcciones juntas: reparametrizacion del comercio, ponderacion por "
        "empleo y logaritmos en productividad y remuneracion media.",
    "6 idem 5, sin FBCF":
        "Comercio + peso + logs, mas sacar x_capital_trabajo (FBCF por ocupado), que la COU "
        "asigna por producto y que no mide intensidad de capital. Cinco variables.",
}


# ==========================================================================
# El universo: la replica de la UAM con el cobre devuelto
# ==========================================================================

def cargar(periodo: int) -> tuple[pd.DataFrame, str]:
    """Las que la UAM clasifica, mas el cobre. Devuelve tambien su codigo."""
    d = pd.read_excel(rutas.exigir(rutas.D_UAM_INSUMO, "el insumo de la UAM"))
    d = d[d["periodo"] == periodo].copy()
    d["codigo_3dig"] = d["codigo_3dig"].astype(str)

    res = pd.read_excel(
        rutas.exigir(rutas.D_UAM_HCLUST[periodo], f"los resultados del periodo {periodo}"),
        sheet_name="sectores_cluster")
    dentro = {str(c) for c in res["codigo_3dig"]}

    fuera = d[~d["codigo_3dig"].isin(dentro)]
    hit = fuera["desc"].map(rutas.normalizar).str.contains("cobre")
    if hit.sum() != 1:
        raise SystemExit(
            f"ERROR periodo {periodo}: entre las {len(fuera)} excluidas por la UAM hay "
            f"{hit.sum()} con 'cobre' en la descripcion, no 1: {list(fuera.loc[hit, 'desc'])}")
    cobre = str(fuera.loc[hit, "codigo_3dig"].iloc[0])

    d = d[d["codigo_3dig"].isin(dentro | {cobre})].reset_index(drop=True)
    if len(d) != len(dentro) + 1:
        raise SystemExit(f"ERROR periodo {periodo}: quedaron {len(d)} actividades, "
                         f"no {len(dentro) + 1}")
    return d, cobre


# ==========================================================================
# Diagnostico del bloque de comercio: cuanta informacion independiente tiene
# ==========================================================================

def diagnostico_comercio(d: pd.DataFrame, anio: int) -> list[dict]:
    """Rango efectivo del bloque de comercio, antes y despues de reparametrizar."""
    filas = []
    casos = [
        ("UAM, balanza en niveles", "nivel",
         ["x_demanda_externa", "x_oferta_externa", "x_balanza_com"]),
        ("UAM, balanza normalizada", "norm",
         ["x_demanda_externa", "x_oferta_externa", "x_balanza_com"]),
        ("UAM, balanza sobre VBP", "vbp",
         ["x_demanda_externa", "x_oferta_externa", "x_balanza_com"]),
        ("reparametrizado (apertura, balanza)", "norm",
         ["x_apertura", "x_balanza_norm"]),
    ]
    for etiq, bal, cols in casos:
        x = uam.construir_variables(d, balanza=bal)
        variables = uam.VARS_COMERCIO if len(cols) == 2 else uam.VARS_UAM
        z = uam.estandarizar(x, variables)
        zc = z[:, [variables.index(v) for v in cols]]
        sv = np.linalg.svd(zc, compute_uv=False)
        p = sv ** 2 / (sv ** 2).sum()
        # R2 de reconstruir la ultima columna del bloque con las anteriores
        A = np.column_stack([np.ones(len(d))] + [zc[:, i] for i in range(zc.shape[1] - 1)])
        y = zc[:, -1]
        beta, *_ = np.linalg.lstsq(A, y, rcond=None)
        r2 = 1 - ((y - A @ beta) ** 2).sum() / ((y - y.mean()) ** 2).sum()
        filas.append({
            "anio": anio, "bloque de comercio": etiq,
            "columnas": len(cols),
            "valores singulares": " / ".join(f"{v:.2f}" for v in sv),
            "menor / mayor": sv[-1] / sv[0],
            "dimension efectiva": float(np.exp(-(p * np.log(p)).sum())),
            "R2 de la ultima sobre las otras": r2,
            "% de la inercia total": 100 * (zc ** 2).sum() / (z ** 2).sum(),
        })
    return filas


def prueba_de_clones(d: pd.DataFrame, anio: int, cobre: str) -> list[dict]:
    """Partir en 3 las 5 actividades de mas empleo no debe cambiar la particion.

    Es la prueba que decide si una especificacion puede comparar nomenclaturas
    de distinta finura (73 contra 111 actividades). Un ratio la pasa; una
    magnitud en niveles no.
    """
    grandes = d["empleo"].nlargest(5).index
    clones = []
    for i, r in d.iterrows():
        veces = 3 if i in grandes else 1
        for j in range(veces):
            s = r.copy()
            for col in SUMABLES:
                s[col] = r[col] / veces
            s["_id"] = f"{i}_{j}"
            clones.append(s)
    dc = pd.DataFrame(clones).reset_index(drop=True)
    primeros = dc.index[dc["_id"].isin(f"{i}_0" for i in d.index)]

    filas = []
    for nombre, opc in ESPECS:
        fila = {"anio": anio, "especificacion": nombre,
                "partidas": ", ".join(d.loc[grandes, "desc"].str[:28])}
        for k in (2, 4):
            a = pc.clasificar(d, k, **opc)
            b = pc.clasificar(dc, k, **opc)
            fila[f"ARI k={k}"] = uam.rand_ajustado(a, b[primeros])
        filas.append(fila)
    return filas


def sensibilidad_log1p(d: pd.DataFrame, anio: int) -> list[dict]:
    """Cuanto depende la particion de la constante arbitraria del log1p.

    Chen y Roth (2024, QJE): log(1+y) no es invariante a las unidades. En un
    ratio el problema es mas debil, pero la constante sigue siendo arbitraria.
    Se compara la especificacion preferida con la apertura en log1p contra la
    misma con la apertura en nivel.
    """
    filas = []
    for nombre, opc in ESPECS:
        if not opc["log"]:
            continue
        for k in (3, 4):
            a = pc.clasificar(d, k, **opc)
            b = pc.clasificar(d, k, **{**opc, "log1p_ratios": False})
            filas.append({"anio": anio, "especificacion": nombre, "k": k,
                          "ARI log1p vs apertura en nivel": uam.rand_ajustado(a, b)})
    return filas


# ==========================================================================
# Escritura
# ==========================================================================

def hoja_tabla(wb: Workbook, nombre: str, titulo: str, notas: list[str],
               t: pd.DataFrame, anchos: tuple) -> None:
    ws = wb.create_sheet(nombre)
    ws.cell(1, 1, titulo).font = Font(bold=True, size=13)
    for i, n in enumerate(notas):
        ws.cell(2 + i, 1, n).font = Font(italic=True, size=9)
    f0 = 2 + len(notas) + 1
    for j, h in enumerate(t.columns):
        c = ws.cell(f0, 1 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
    for i, fila in enumerate(t.itertuples(index=False)):
        for j, v in enumerate(fila):
            c = ws.cell(f0 + 1 + i, 1 + j, None if (isinstance(v, float) and np.isnan(v)) else v)
            if isinstance(v, float):
                c.number_format = "0.000"
    for j, w in enumerate(anchos):
        ws.column_dimensions[chr(ord("A") + j)].width = w
    ws.freeze_panes = ws.cell(f0 + 1, 1)


def main() -> None:
    rutas.preparar_directorios()
    d1, cobre1 = cargar(1)
    d2, cobre2 = cargar(2)
    print(f"2003: {len(d1)} actividades (62 de la UAM + cobre, codigo {cobre1})")
    print(f"2023: {len(d2)} actividades (101 de la UAM + cobre, codigo {cobre2})")

    periodos = [(1, d1, cobre1, 2003), (2, d2, cobre2, 2023)]

    wb = Workbook()
    wb.remove(wb.active)
    resumen, curvas, cobre_filas, comparacion = [], [], [], []
    particiones: dict[tuple[str, int], np.ndarray] = {}

    for nombre, opc in ESPECS:
        bs, tablas, diag, elec = [], [], [], {}
        for p, d, cobre, anio in periodos:
            elec[p] = pc.elegir_k(d, kmax=KMAX, **opc)
            cl = pc.renumerar(pc.clasificar(d, K, **opc), d["empleo"].to_numpy())
            particiones[(nombre, p)] = cl

            b = pc.bloque(d, p, cl, marcadas={cobre})
            bs.append((b, f"Periodo {p} — {anio} ({len(d)} actividades, k = {K})"))
            diag.append(pc.diagnostico(b, d))
            tablas.append([
                (pc.medias_en_niveles(d, cl),
                 f"Periodo {p} — {anio}: medias por conglomerado, en niveles "
                 "(ratios = cociente de las sumas; indices con la economia = 100)"),
                (pc.medias_en_z(d, cl, **opc),
                 f"Periodo {p} — {anio}: centroides en puntajes z, las variables tal como "
                 "entran al algoritmo" + (", ponderados por empleo" if opc["peso"] else "")),
            ])

            # --- donde queda el cobre, y con quien
            ic = int(np.where(d["codigo_3dig"].to_numpy() == cobre)[0][0])
            grupo = cl == cl[ic]
            sub = d[grupo]
            cobre_filas.append({
                "anio": anio, "especificacion": nombre,
                "conglomerado del cobre": int(cl[ic]),
                "actividades en el": int(grupo.sum()),
                "solo?": "SI — sigue aislado" if grupo.sum() == 1 else "no",
                "% del empleo del grupo": 100 * sub["empleo"].sum() / d["empleo"].sum(),
                "% de las expo del grupo": 100 * sub["expo"].sum() / d["expo"].sum(),
                "productividad del grupo (=100)":
                    100 * (sub["valor_agrega"].sum() / sub["empleo"].sum())
                    / (d["valor_agrega"].sum() / d["empleo"].sum()),
                "masa salarial / VA del grupo (%)":
                    100 * sub["remunera"].sum() / sub["valor_agrega"].sum(),
                "acompanantes de mas empleo": ", ".join(
                    sub[sub["codigo_3dig"] != cobre]
                    .sort_values("empleo", ascending=False)["desc"].str[:34].head(4)),
            })

            for kk in range(2, KMAX + 1):
                curvas.append({"especificacion": nombre, "anio": anio, "k": kk,
                               "salto de costes": elec[p]["salto"].get(kk),
                               "eta2": elec[p]["eta2"].get(kk),
                               "elegido": "<--" if kk == elec[p]["k"] else ""})

        notas = [
            (QUE_HACE[nombre], "000000"),
            (f"k = {K} impuesto, para que las seis especificaciones sean comparables. Los dos "
             f"criterios pediran {elec[1]['k']} en 2003 (salto {elec[1]['salto'].max():.2f}; "
             f"eta2 {elec[1]['k_eta2']}) y {elec[2]['k']} en 2023 "
             f"(salto {elec[2]['salto'].max():.2f}; eta2 {elec[2]['k_eta2']}). "
             f"Conglomerados numerados de mayor a menor empleo.", "000000"),
            ("Universo: las que la UAM clasifica MAS el cobre, que va en rojo y cursiva. "
             "Las demas exclusiones de la UAM se respetan.", pc.ROJO),
            ("Al pie: medias por conglomerado en niveles (legibles) y centroides en puntajes z "
             "(los que minimiza el algoritmo).", "000000"),
        ]
        pc.escribir_hoja(wb, nombre, notas, bs, con_marca=True, tablas=tablas)

        resumen.append({
            "especificacion": nombre,
            "variables": len(pc.variables_de(opc["sinfbcf"], opc["comercio"])),
            "tamanos 2003": diag[0][0], "% empleo 2003": diag[0][1], "<1% 2003": diag[0][2],
            "k que pediria 2003": elec[1]["k"], "(eta2) 2003": elec[1]["k_eta2"],
            "tamanos 2023": diag[1][0], "% empleo 2023": diag[1][1], "<1% 2023": diag[1][2],
            "k que pediria 2023": elec[2]["k"], "(eta2) 2023": elec[2]["k_eta2"],
            "que hace": QUE_HACE[nombre],
        })

    # --- comparacion entre especificaciones: contra la base y contra la anterior
    for p, d, cobre, anio in periodos:
        for i, (nombre, _) in enumerate(ESPECS):
            comparacion.append({
                "anio": anio, "especificacion": nombre,
                "ARI contra 1 UAM base": uam.rand_ajustado(
                    particiones[(ESPECS[0][0], p)], particiones[(nombre, p)]),
                "ARI contra la anterior": np.nan if i == 0 else uam.rand_ajustado(
                    particiones[(ESPECS[i - 1][0], p)], particiones[(nombre, p)]),
                "ARI contra 5 (las tres juntas)": uam.rand_ajustado(
                    particiones[(ESPECS[4][0], p)], particiones[(nombre, p)]),
            })

    diag_com = [f for p, d, c, anio in periodos for f in diagnostico_comercio(d, anio)]
    clones = [f for p, d, c, anio in periodos for f in prueba_de_clones(d, anio, c)]
    log1p = [f for p, d, c, anio in periodos for f in sensibilidad_log1p(d, anio)]

    # ---------------------------------------------------------------- hojas
    hoja_tabla(
        wb, "01 Comparacion",
        "Cuanto mueve la particion cada correccion (k = 4)",
        ["ARI ajustado. 1,000 = la misma particion; 0 = la coincidencia de dos particiones al azar.",
         "«Contra la anterior» mide el aporte incremental de cada paso de la escalera. La 4 esta "
         "fuera de la escalera (logs sin peso), asi que su columna incremental se lee contra la 3."],
        pd.DataFrame(comparacion), (8, 36, 22, 22, 28))

    hoja_tabla(
        wb, "02 El cobre",
        "Donde queda el cobre en cada especificacion",
        ["La pregunta del universo: con la especificacion de la UAM el cobre queda solo en su "
         "propio conglomerado y por eso ellos lo excluyen.",
         "«solo? = SI» significa que el problema sigue; un grupo con varias actividades y una "
         "masa salarial baja es el enclave exportador que el capitulo describe."],
        pd.DataFrame(cobre_filas), (8, 36, 12, 12, 18, 14, 14, 16, 16, 60))

    hoja_tabla(
        wb, "03 Bloque de comercio",
        "Cuanta informacion independiente tienen las columnas de comercio",
        ["Un valor singular en 0 y un R2 de 1,000 significan que la ultima columna es "
         "combinacion exacta de las otras: el bloque tiene rango 2 aunque ocupe 3 columnas.",
         "«% de la inercia total» es mecanico: al estandarizar, cada columna aporta lo mismo, "
         "asi que 3 de 7 columnas se llevan 42,9 % pase lo que pase. Ese es el problema.",
         "La ultima fila de cada anio es la reparametrizacion: 2 columnas, 33,3 % de la inercia "
         "y sin redundancia."],
        pd.DataFrame(diag_com), (8, 38, 11, 24, 14, 16, 22, 18))

    hoja_tabla(
        wb, "04 Invariancia a clonar",
        "Prueba de clones: partir en 3 las 5 actividades de mas empleo",
        ["Es lo que decide si una especificacion puede comparar 2003 (73 actividades) con 2023 "
         "(111): mas desagregacion no debe cambiar la particion.",
         "ARI = 1,000 es invariancia exacta. Las dos variables nuevas son ratios, asi que la "
         "reparametrizacion deberia conservarla cuando se pondera."],
        pd.DataFrame(clones), (8, 36, 60, 12, 12))

    hoja_tabla(
        wb, "05 Sensibilidad al log1p",
        "Cuanto depende la particion de la constante arbitraria del log1p",
        ["Chen y Roth (2024, QJE): log(1+y) no es invariante a las unidades. La apertura es un "
         "ratio, asi que el problema es mas debil, pero la constante sigue siendo arbitraria.",
         "Se compara cada especificacion con logs contra la misma dejando la apertura en nivel. "
         "Un ARI alto significa que la eleccion no importa y se puede reportar en una linea."],
        pd.DataFrame(log1p), (8, 36, 8, 30))

    ws = wb.create_sheet("00 Eleccion de k", 0)
    cur = pd.DataFrame(curvas)
    ws.cell(1, 1, "Los dos criterios, recalculados en cada especificacion y periodo"
            ).font = Font(bold=True, size=13)
    ws.cell(2, 1, "En las hojas de clasificacion el k esta IMPUESTO en 4, por comparabilidad. "
                  "Esta hoja dice que habria pedido cada especificacion por su cuenta."
            ).font = Font(italic=True, size=9)
    ws.cell(3, 1, "salto de costes: cuanto mas caro es bajar de k+1 a k grupos que bajar de k+2 "
                  "a k+1; manda este. | eta2: fraccion de la varianza del log de la "
                  "productividad, ponderada por empleo, que explica la particion."
            ).font = Font(italic=True, size=9)
    for j, h in enumerate(cur.columns):
        c = ws.cell(5, 1 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
    for i, fila in enumerate(cur.itertuples(index=False)):
        for j, v in enumerate(fila):
            c = ws.cell(6 + i, 1 + j, v)
            if isinstance(v, float):
                c.number_format = "0.000"
            if fila.elegido:
                c.font = Font(bold=True)
    for col, w in zip("ABCDEF", (36, 8, 6, 16, 10, 9)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("0 Lectura", 0)
    for i, (texto, negrita, tam) in enumerate([
        ("T_I26 — reparametrizar el bloque de comercio, con el cobre adentro", True, 13),
        ("Universo: las 62 y 101 actividades que la UAM clasifica, MAS el cobre. Su criterio de "
         "exclusion se respeta en todo lo demas, aunque no lo conozcamos.", False, 9),
        ("La correccion: las TRES columnas de comercio de la UAM (impo/VBP, expo/VBP, balanza) "
         "pasan a DOS —apertura (X+M)/VBP y balanza (X-M)/(X+M)—. Con la balanza corregida la "
         "tercera era funcion exacta de las otras dos, asi que el comercio se llevaba 42,9 % de "
         "la inercia contando una direccion dos veces. La transformacion es biyectiva: no se "
         "pierde informacion, se quita la duplicacion.", False, 9),
        ("Hojas: 00 eleccion de k · las seis especificaciones · 01 comparacion · 02 el cobre · "
         "03 bloque de comercio · 04 invariancia a clonar · 05 sensibilidad al log1p.", False, 9),
        ("En cada hoja de clasificacion, al pie, las medias por conglomerado en niveles y los "
         "centroides en puntajes z, por periodo.", False, 9),
    ]):
        ws.cell(1 + i, 1, texto).font = Font(bold=negrita, italic=not negrita, size=tam)

    t = pd.DataFrame(resumen)
    for j, h in enumerate(t.columns):
        c = ws.cell(7, 1 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
    for i, fila in enumerate(t.itertuples(index=False)):
        for j, v in enumerate(fila):
            ws.cell(8 + i, 1 + j, v)
        ws.cell(8 + i, 1).font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJKLM",
                      (36, 10, 18, 26, 9, 12, 11, 18, 26, 9, 12, 11, 100)):
        ws.column_dimensions[col].width = w

    wb.save(SALIDA)
    print(f"\nescrito: {SALIDA}\n")

    print("== ARI contra la especificacion de la UAM, k = 4 ==")
    print(pd.DataFrame(comparacion).round(3).to_string(index=False))
    print("\n== Donde queda el cobre ==")
    print(pd.DataFrame(cobre_filas)[
        ["anio", "especificacion", "actividades en el", "solo?", "% del empleo del grupo",
         "% de las expo del grupo", "masa salarial / VA del grupo (%)"]
    ].round(2).to_string(index=False))
    print("\n== Bloque de comercio ==")
    print(pd.DataFrame(diag_com)[
        ["anio", "bloque de comercio", "columnas", "menor / mayor", "dimension efectiva",
         "R2 de la ultima sobre las otras", "% de la inercia total"]
    ].round(3).to_string(index=False))
    print("\n== Invariancia a clonar ==")
    print(pd.DataFrame(clones)[
        ["anio", "especificacion", "ARI k=2", "ARI k=4"]].round(3).to_string(index=False))
    print("\n== Sensibilidad al log1p ==")
    print(pd.DataFrame(log1p).round(3).to_string(index=False))


if __name__ == "__main__":
    main()

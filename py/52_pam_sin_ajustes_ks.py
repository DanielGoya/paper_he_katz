"""Bloque K.9 — PAM Manhattan SIN NINGUN AJUSTE, k = 3, 4 y 5, en los cuatro
paises, con Ward como contraste.

POR QUE EXISTE (DG, 2026-08-13). La corrida «PAM sin ajustes» —las siete
variables de la UAM intactas, en niveles, sin ponderar y con la balanza comercial
en niveles, cambiando UNICAMENTE el algoritmo— ya estaba hecha para los cuatro
paises, pero **solo con k = 4** (`T_K11`, especificacion A; `T_K12` es la misma
en formato plano). Faltaban k = 3 y k = 5.

Y falta ademas una alineacion: en `T_K11` Chile corria con la COU COMPLETA (73 y
111 actividades), mientras que las corridas posteriores —`T_K14` y `T_K15`— usan
el universo que la UAM efectivamente clasifica MAS el cobre reintegrado (63 y
102). Aca se usa el segundo, para que las hojas chilenas sean comparables con las
ultimas y no con las primeras.

QUE SE CORRE. Dos metodos, tres cortes, cuatro paises:

    Ward   sobre distancia euclidiana, las siete variables en niveles, sin
           ponderar, balanza en niveles. Es la especificacion publicada. Va como
           contraste en cada corte, igual que en T_K14 y T_K15.
    PAM    con distancia de Manhattan, sobre EXACTAMENTE las mismas siete
           variables, tambien en niveles y sin ponderar. Lo unico que cambia
           entre las dos hojas de un mismo corte es el algoritmo: aglomerar
           contra particionar.

Es la comparacion mas limpia disponible —una sola pieza distinta— y por eso vale
la pena tenerla en los tres cortes: si el enclave exportador aparece con el
algoritmo solo, sin tocar ninguna variable, deja de depender de las correcciones
de especificacion del proyecto.

EL UNIVERSO, igual que en T_K14 y T_K15: Chile, las actividades que la UAM
clasifica MAS el cobre reintegrado (63 en 2003, 102 en 2023); Brasil, Mexico y
Argentina, la COU completa, porque ahi no hay lista de la UAM que restringir —su
criterio de exclusion no esta documentado (bloque K.4)—.

SALIDA. `T_K16`, un libro con veinticinco hojas: lectura, y seis por pais
(«<pais> Ward k=3» … «<pais> PAM k=5»), con los dos anios lado a lado y las
medias por conglomerado, en niveles y en puntajes z, al pie de cada anio.

DEPENDE DE. `50_ward_pam_universo_uam.py`, de donde salen los cargadores del
universo y el armado de bloques; `planilla_clusters.py` y `26_pam_medoides.py`.
"""

from __future__ import annotations

import importlib.util
import sys
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402
import planilla_clusters as pc  # noqa: E402


def _cargar_modulo(archivo: str, nombre: str):
    """Los modulos de `py/` empiezan con digito y no se pueden importar por nombre."""
    spec = importlib.util.spec_from_file_location(
        nombre, Path(__file__).resolve().parent / archivo)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


pam_mod = _cargar_modulo("26_pam_medoides.py", "pam_medoides")
# De aca salen `cargar()` (universo UAM+cobre para Chile, COU completa para el
# resto), `bloque()`, `extremas()` y `NOTA_UNIVERSO`: identicos, no se duplican.
k14 = _cargar_modulo("50_ward_pam_universo_uam.py", "ward_pam_universo_uam")
uam = pc.uam

HOY = date.today().strftime("%Y.%m.%d")
SALIDA = rutas.OUT / f"{HOY} T_K16 PAM sin ajustes con k=3,4,5.xlsx"

KS = (3, 4, 5)
PAISES = k14.PAISES

# Una sola especificacion de variables para los dos metodos: las siete de la UAM,
# en niveles, sin ponderar, balanza en niveles. Lo unico que cambia es el algoritmo.
SPEC = dict(log=False, peso=False, balanza="nivel", sinfbcf=False, comercio=False)

NOTAS = {
    "Ward": [
        ("Ward sobre distancia euclidiana, las siete variables de la UAM en niveles, sin "
         "ponderar y con la balanza comercial en niveles: la especificacion publicada, SIN "
         "ningun ajuste. Va como contraste de la hoja PAM del mismo corte.", "000000"),
    ],
    "PAM": [
        ("PAM (k-medoides) con distancia de Manhattan sobre EXACTAMENTE las mismas siete "
         "variables: en niveles, sin ponderar, balanza comercial en niveles. SIN ningun ajuste "
         "de especificacion. Lo unico que cambia respecto de la hoja Ward del mismo corte es el "
         "ALGORITMO: aglomerar contra particionar.", "000000"),
        ("PAM trabaja con MEDOIDES —el centro de cada grupo es una actividad real, no una "
         "media— y su fase SWAP puede reasignar una actividad mal colocada, cosa que Ward, "
         "aglomerativo, nunca revisa.", "000000"),
    ],
}


def clasificar(d: pd.DataFrame, metodo: str, k: int) -> np.ndarray:
    """Devuelve la particion, renumerada de mayor a menor empleo."""
    if metodo == "Ward":
        cl = pc.clasificar(d, k, **SPEC)
    elif metodo == "PAM":
        z, pesos = pc.matriz_z(d, **SPEC)          # pesos es None: SPEC no pondera
        D = pam_mod.matriz_distancias(z, "manhattan")
        _, et = pam_mod.pam(D, k, pesos)
        cl = et + 1
    else:
        raise ValueError(metodo)
    return pc.renumerar(cl, d["empleo"].to_numpy())


def main() -> None:
    rutas.preparar_directorios()
    pam_mod.verificar()

    datos, cobres = k14.cargar()
    print(f"Chile: {(datos[(datos.pais == 'Chile') & (datos.anio == 2003)]).shape[0]} actividades "
          f"en 2003 y {(datos[(datos.pais == 'Chile') & (datos.anio == 2023)]).shape[0]} en 2023 "
          "— clasificadas por la UAM + cobre reintegrado\n")

    wb = Workbook()
    wb.remove(wb.active)
    resumen = []

    for pais, anios in PAISES:
        for k in KS:
            for metodo in ("Ward", "PAM"):
                bloques, tablas = [], []

                for anio in anios:
                    d = datos[(datos.pais == pais) & (datos.anio == anio)].reset_index(drop=True)
                    if d.empty:
                        print(f"  AVISO: no hay datos de {pais} {anio}")
                        continue

                    if pais == "Chile":
                        marcadas = {cobres[anio]}
                    else:
                        ext = k14.extremas(d)
                        i_ext = int(np.argmax(ext))
                        i_exp = int(np.argmax(d["expo"].to_numpy()))
                        marcadas = {d.loc[i_ext, "codigo"], d.loc[i_exp, "codigo"]}

                    cl = clasificar(d, metodo, k)
                    titulo = f"{pais} {anio} ({len(d)} actividades, k = {k})"
                    bloques.append((k14.bloque(d, pais, anio, cl, marcadas), titulo))

                    t_niv = pc.medias_en_niveles(d, cl)
                    tablas.append([
                        (t_niv, f"{anio}: medias por conglomerado, en niveles (ratios = "
                                "cociente de las sumas; indices con la economia = 100)"),
                        (pc.medias_en_z(d, cl, **SPEC),
                         f"{anio}: medias en puntajes z de las siete variables, tal como entran "
                         "al agrupamiento"),
                    ])

                    g = int(t_niv["% expo"].idxmax())
                    fila = {
                        "pais": pais, "anio": anio, "metodo": metodo, "k": k,
                        "n actividades": len(d),
                        "tamanos": " / ".join(str(x) for x in t_niv["n"].astype(int)),
                        "% empleo": " / ".join(f"{x:.1f}" for x in t_niv["% empleo"]),
                        "productividad": " / ".join(
                            f"{x:.0f}" for x in t_niv["productividad (=100)"]),
                        "empleo del mayor (%)": float(t_niv["% empleo"].max()),
                        "grupos <1% empleo": int((t_niv["% empleo"] < 1.0).sum()),
                        "grupos de 1 actividad": int((t_niv["n"] == 1).sum()),
                        "grupo con mas expo: actividades": int(t_niv.loc[g, "n"]),
                        "grupo con mas expo: % empleo": float(t_niv.loc[g, "% empleo"]),
                        "grupo con mas expo: % expo": float(t_niv.loc[g, "% expo"]),
                        "grupo con mas expo: masa salarial/VA (%)":
                            100 * float(t_niv.loc[g, "masa salarial / VA"]),
                        "eta2 log productividad": pam_mod.eta2_productividad(
                            d, cl, d["empleo"].to_numpy(dtype=float)),
                    }
                    if pais == "Chile":
                        ic = int(np.where(d["codigo"].to_numpy() == cobres[anio])[0][0])
                        sub = d[cl == cl[ic]]
                        fila["cobre: actividades en su grupo"] = int(len(sub))
                        fila["cobre: % expo de su grupo"] = (
                            100 * sub["expo"].sum() / d["expo"].sum())
                        fila["cobre: masa salarial/VA de su grupo (%)"] = (
                            100 * sub["remunera"].sum() / sub["valor_agrega"].sum())
                    resumen.append(fila)

                notas = NOTAS[metodo] + [
                    (f"Corte k = {k}. Conglomerados numerados de mayor a menor empleo. Al pie, "
                     "para cada anio, las medias por conglomerado en niveles y en puntajes z.",
                     "000000"),
                    k14.NOTA_UNIVERSO["Chile" if pais == "Chile" else "otros"],
                ]
                pc.escribir_hoja(wb, f"{pais} {metodo} k={k}", notas, bloques,
                                 con_marca=True, tablas=tablas)
            print(f"  {pais:<10} k={k}  Ward y PAM escritas")

    T = pd.DataFrame(resumen)
    ari = []
    for pais, anios in PAISES:
        for anio in anios:
            d = datos[(datos.pais == pais) & (datos.anio == anio)].reset_index(drop=True)
            if d.empty:
                continue
            for k in KS:
                ari.append({
                    "pais": pais, "anio": anio, "k": k,
                    "ARI Ward vs PAM": uam.rand_ajustado(
                        clasificar(d, "Ward", k), clasificar(d, "PAM", k)),
                })
    T = T.merge(pd.DataFrame(ari), on=["pais", "anio", "k"], how="left")

    ws = wb.create_sheet("0 Lectura", 0)
    lineas = [
        ("T_K16 — PAM Manhattan SIN ningun ajuste, k = 3, 4 y 5, con Ward como contraste",
         True, 13),
        ("Las dos hojas de cada corte usan EXACTAMENTE las mismas siete variables de la UAM, en "
         "niveles, sin ponderar y con la balanza comercial en niveles. Lo unico que cambia es "
         "el ALGORITMO: Ward aglomerativo euclidiano contra PAM particional con distancia de "
         "Manhattan. Es la comparacion mas limpia disponible, con una sola pieza distinta.",
         False, 9),
        ("Completa una corrida que existia solo con k = 4 (T_K11, especificacion A; T_K12 es la "
         "misma en formato plano). Aca van k = 3, 4 y 5.", False, 9),
        ("Ademas alinea el universo chileno: T_K11 corria Chile con la COU completa (73 y 111 "
         "actividades), mientras que este libro usa el de T_K14 y T_K15 —las que la UAM "
         "clasifica MAS el cobre reintegrado, 63 y 102—, para que las hojas chilenas sean "
         "comparables con las ultimas corridas y no con las primeras.", False, 9),
        ("Universo: Chile, las que la UAM clasifica mas el cobre. Brasil, Mexico y Argentina, "
         "la COU completa, porque ahi no hay lista de la UAM que restringir. En rojo y cursiva, "
         "el cobre en Chile; en los otros tres, la actividad mas extrema (max |z|) y el mayor "
         "exportador, que son su sustituto.", False, 9),
        (f"Generada el {HOY} por py/52_pam_sin_ajustes_ks.py.", False, 9),
        ("", False, 9),
        ("Resumen (una fila por pais, anio, metodo y corte):", True, 10),
    ]
    for i, (texto, negrita, tam) in enumerate(lineas):
        ws.cell(1 + i, 1, texto).font = Font(bold=negrita, italic=not negrita, size=tam)
    f0 = len(lineas) + 2
    for j, h in enumerate(T.columns):
        ws.cell(f0, 1 + j, h).font = Font(bold=True, size=9)
    for i, fila in enumerate(T.itertuples(index=False)):
        for j, v in enumerate(fila):
            c = ws.cell(f0 + 1 + i, 1 + j,
                        None if (isinstance(v, float) and np.isnan(v)) else v)
            if isinstance(v, float):
                c.number_format = "0.000"
    ws.column_dimensions["A"].width = 120

    wb.save(SALIDA)
    print(f"\nescrito: {SALIDA}\n")
    cols = ["pais", "anio", "metodo", "k", "tamanos", "empleo del mayor (%)",
            "grupos <1% empleo", "grupo con mas expo: actividades",
            "grupo con mas expo: % empleo", "grupo con mas expo: % expo",
            "grupo con mas expo: masa salarial/VA (%)", "ARI Ward vs PAM"]
    print(T[cols].round(2).to_string(index=False))
    if "cobre: actividades en su grupo" in T.columns:
        print("\n== Donde queda el cobre (Chile) ==")
        print(T[T.pais == "Chile"][
            ["anio", "metodo", "k", "cobre: actividades en su grupo",
             "cobre: % expo de su grupo", "cobre: masa salarial/VA de su grupo (%)"]
        ].round(1).to_string(index=False))


if __name__ == "__main__":
    main()

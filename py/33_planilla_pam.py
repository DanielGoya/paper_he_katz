"""Bloque I.12 — PAM / k-medoides como alternativa completa a Ward, en planilla.

LA PREGUNTA (DG, 2026-08-13). En 2003 la especificacion 5 de T_I26 —comercio
reparametrizado + peso + logs, con Ward— deja al cobre dentro de un conglomerado
de 48 actividades con el 44,8 % del empleo, que no es un estrato sino «el resto
de la economia». El mecanismo es el criterio de Ward ponderado: el coste de
fusionar es (Wa*Wb/(Wa+Wb))*d^2, y con el cobre pesando 0,29 % del empleo ese
factor queda dominado por la masa chica casi sin importar la distancia. El grupo
del cobre se vuelve el destino mas barato y absorbe bloques enteros por
aritmetica, no por afinidad.

PAM no tiene ese problema porque no fusiona: reparte. No es una variante de Ward
sino otra familia —particional en vez de aglomerativa—, con tres diferencias que
importan aca:

    el centro de un grupo es una OBSERVACION real (el medoide, analogo
      multivariado de la mediana), no una media que un extremo arrastra;
    la distancia de Manhattan no eleva al cuadrado, asi que una variable
      extrema no domina el resto;
    la fase SWAP REASIGNA: puede sacar de un grupo a una actividad mal
      colocada, cosa que Ward, aglomerativo, nunca revisa.

El precio: PAM no produce jerarquia —no hay dendrograma ni particiones anidadas
al bajar k— y el k hay que fijarlo de antemano. De ahi que esta planilla recorra
varios k en vez de uno.

LAS DOS ESPECIFICACIONES, que son las que pidio Daniel:

    A  sin arreglar nada mas   las SIETE variables de la UAM intactas: en
                               niveles, sin ponderar, balanza comercial en
                               niveles. Lo unico que cambia respecto del
                               ejercicio publicado es el algoritmo.
    B  comercio ajustado       las SEIS de la reparametrizacion del bloque I.11:
                               apertura (X+M)/VBP y balanza (X-M)/(X+M) en lugar
                               de las tres columnas de comercio, de las cuales
                               una era funcion exacta de las otras dos.

Sin logaritmos y sin ponderacion en las dos, a proposito: la pregunta es que
hace PAM POR SI SOLO, y agregar correcciones la contestaria mezclada.

EL UNIVERSO. El mismo de T_I26, para poder comparar hoja contra hoja: las
actividades que la UAM clasifica MAS el cobre (63 en 2003, 102 en 2023). Su
criterio de exclusion se respeta en todo lo demas.

LA METRICA. Manhattan en las hojas de clasificacion, que es la que el bloque I.6
dejo como eleccion del proyecto. La euclidiana se corre igual y va en la hoja de
comparacion, para que el efecto de la metrica quede separado del algoritmo.

SALIDA. `T_I27`: diez hojas de clasificacion (dos especificaciones x k = 2..6) en
el formato de los libros 3a y 3b, con las medias por conglomerado al pie, mas
cinco hojas de diagnostico.

DEPENDE DE. `40_uam_metodo.py`, `planilla_clusters.py` y `26_pam_medoides.py`,
de donde sale la implementacion de PAM ya verificada (cuatro controles, entre
ellos el optimo global por enumeracion exhaustiva).
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
import rutas  # noqa: E402
import planilla_clusters as pc  # noqa: E402


def _cargar_modulo(archivo: str, nombre: str):
    """Los modulos de `py/` empiezan con digito y no se pueden importar por nombre."""
    spec = importlib.util.spec_from_file_location(
        nombre, Path(__file__).resolve().parent / archivo)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


pam_mod = _cargar_modulo("26_pam_medoides.py", "pam_medoides")
uam = pc.uam

FECHA = "2026.08.13"
SALIDA = rutas.OUT / f"{FECHA} T_I27 PAM con distintos k y comercio ajustado.xlsx"

KS = (2, 3, 4, 5, 6)
KMAX = 8

ESPECS = [
    ("A sin ajustes",
     dict(log=False, peso=False, balanza="nivel", sinfbcf=False, comercio=False)),
    ("B comercio ajustado",
     dict(log=False, peso=False, balanza="norm", sinfbcf=False, comercio=True)),
]

QUE_HACE = {
    "A sin ajustes":
        "Las SIETE variables de la UAM intactas: en niveles, sin ponderar, balanza comercial "
        "en niveles. Lo unico que cambia respecto del ejercicio publicado es el algoritmo: "
        "PAM con distancia de Manhattan en lugar de Ward euclidiano.",
    "B comercio ajustado":
        "Las SEIS variables de la reparametrizacion: apertura (X+M)/VBP y balanza (X-M)/(X+M) "
        "en lugar de las tres columnas de comercio de la UAM, de las cuales una era funcion "
        "exacta de las otras dos. Sigue sin logaritmos y sin ponderar.",
}


# ==========================================================================
# Universo: el mismo de T_I26
# ==========================================================================

def cargar(periodo: int) -> tuple[pd.DataFrame, str]:
    """Las que la UAM clasifica, mas el cobre. Devuelve tambien su codigo."""
    d = pd.read_excel(rutas.exigir(rutas.D_UAM_INSUMO, "el insumo de la UAM"))
    d = d[d["periodo"] == periodo].copy()
    d["codigo_3dig"] = d["codigo_3dig"].astype(str)

    res = pd.read_excel(
        rutas.exigir(rutas.D_UAM_HCLUST[periodo], f"los resultados del periodo {periodo}"),
        sheet_name="sectores_cluster")
    dentro = {str(c) for c in res["codigo_3dig"]}

    fuera = d[~d["codigo_3dig"].isin(dentro)]
    hit = fuera["desc"].map(rutas.normalizar).str.contains("cobre")
    if hit.sum() != 1:
        raise SystemExit(
            f"ERROR periodo {periodo}: entre las {len(fuera)} excluidas por la UAM hay "
            f"{hit.sum()} con 'cobre' en la descripcion, no 1")
    cobre = str(fuera.loc[hit, "codigo_3dig"].iloc[0])

    d = d[d["codigo_3dig"].isin(dentro | {cobre})].reset_index(drop=True)
    if len(d) != len(dentro) + 1:
        raise SystemExit(f"ERROR periodo {periodo}: quedaron {len(d)} actividades")
    return d, cobre


# ==========================================================================
# Particion
# ==========================================================================

def particionar(d: pd.DataFrame, k: int, metrica: str,
                **opc) -> tuple[np.ndarray, list[int], np.ndarray]:
    """PAM sobre las variables de una especificacion. Devuelve etiquetas
    renumeradas por empleo, los medoides (en el indice ORIGINAL) y la matriz de
    distancias."""
    z, pesos = pc.matriz_z(d, **opc)
    D = pam_mod.matriz_distancias(z, metrica)
    med, et = pam_mod.pam(D, k, pesos)

    # `renumerar` reordena las etiquetas por empleo; hay que arrastrar los
    # medoides al numero nuevo o dejan de corresponder con su conglomerado.
    nuevo = pc.renumerar(et + 1, d["empleo"].to_numpy())
    mapa = {int(et[m] + 1): int(nuevo[m]) for m in med}
    medoides = [med[i] for i in np.argsort([mapa[int(et[m] + 1)] for m in med])]
    return nuevo, medoides, D


def ward(d: pd.DataFrame, k: int, **opc) -> np.ndarray:
    """La misma especificacion con Ward, para comparar."""
    return pc.renumerar(pc.clasificar(d, k, **opc), d["empleo"].to_numpy())


# ==========================================================================
# Escritura
# ==========================================================================

def hoja_tabla(wb: Workbook, nombre: str, titulo: str, notas: list[str],
               t: pd.DataFrame, anchos: tuple) -> None:
    ws = wb.create_sheet(nombre)
    ws.cell(1, 1, titulo).font = Font(bold=True, size=13)
    for i, n in enumerate(notas):
        ws.cell(2 + i, 1, n).font = Font(italic=True, size=9)
    f0 = 2 + len(notas) + 1
    for j, h in enumerate(t.columns):
        c = ws.cell(f0, 1 + j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
    for i, fila in enumerate(t.itertuples(index=False)):
        for j, v in enumerate(fila):
            c = ws.cell(f0 + 1 + i, 1 + j,
                        None if (isinstance(v, float) and np.isnan(v)) else v)
            if isinstance(v, float):
                c.number_format = "0.000"
    for j, w in enumerate(anchos):
        ws.column_dimensions[chr(ord("A") + j)].width = w
    ws.freeze_panes = ws.cell(f0 + 1, 1)


def main() -> None:
    rutas.preparar_directorios()
    pam_mod.verificar()

    d1, cobre1 = cargar(1)
    d2, cobre2 = cargar(2)
    print(f"\n2003: {len(d1)} actividades (62 de la UAM + cobre)")
    print(f"2023: {len(d2)} actividades (101 de la UAM + cobre)\n")
    periodos = [(1, d1, cobre1, 2003), (2, d2, cobre2, 2023)]

    wb = Workbook()
    wb.remove(wb.active)
    criterios, cobre_filas, comparacion, medoides_filas = [], [], [], []

    for nombre, opc in ESPECS:
        for k in KS:
            bs, tablas = [], []
            for p, d, cobre, anio in periodos:
                cl, med, D = particionar(d, k, "manhattan", **opc)
                w = d["empleo"].to_numpy(dtype=float)

                b = pc.bloque(d, p, cl, marcadas={cobre})
                bs.append((b, f"Periodo {p} — {anio} ({len(d)} actividades, k = {k})"))
                t_niv = pc.medias_en_niveles(d, cl)
                tablas.append([
                    (t_niv,
                     f"Periodo {p} — {anio}: medias por conglomerado, en niveles "
                     "(ratios = cociente de las sumas; indices con la economia = 100)"),
                    (pc.medias_en_z(d, cl, **opc),
                     f"Periodo {p} — {anio}: medias en puntajes z. PAM trabaja con MEDOIDES, "
                     "no con medias: la media va como descripcion, y el medoide de cada grupo "
                     "esta en la hoja «04 Medoides»."),
                ])

                # --- criterios de calidad, por k
                criterios.append({
                    "especificacion": nombre, "anio": anio, "k": k,
                    "tamanos": " / ".join(str(x) for x in t_niv["n"].astype(int)),
                    "% empleo": " / ".join(f"{x:.1f}" for x in t_niv["% empleo"]),
                    "empleo del mayor (%)": float(t_niv["% empleo"].max()),
                    "grupos <1% empleo": int((t_niv["% empleo"] < 1.0).sum()),
                    "silueta ponderada": pam_mod.silueta(D, cl, w),
                    "silueta sin ponderar": pam_mod.silueta(D, cl, None),
                    "eta2 log productividad": pam_mod.eta2_productividad(d, cl, w),
                })

                # --- donde queda el cobre
                ic = int(np.where(d["codigo_3dig"].to_numpy() == cobre)[0][0])
                grupo = cl == cl[ic]
                sub = d[grupo]
                cobre_filas.append({
                    "especificacion": nombre, "anio": anio, "k": k,
                    "conglomerado del cobre": int(cl[ic]),
                    "actividades en el": int(grupo.sum()),
                    "solo?": "SI — sigue aislado" if grupo.sum() == 1 else "no",
                    "% del empleo": 100 * sub["empleo"].sum() / d["empleo"].sum(),
                    "% de las expo": 100 * sub["expo"].sum() / d["expo"].sum(),
                    "productividad (=100)":
                        100 * (sub["valor_agrega"].sum() / sub["empleo"].sum())
                        / (d["valor_agrega"].sum() / d["empleo"].sum()),
                    "masa salarial / VA (%)":
                        100 * sub["remunera"].sum() / sub["valor_agrega"].sum(),
                    "acompanantes de mas empleo": ", ".join(
                        sub[sub["codigo_3dig"] != cobre]
                        .sort_values("empleo", ascending=False)["desc"].str[:32].head(4)),
                })

                # --- medoides: la actividad representativa de cada grupo
                for i, m in enumerate(med, start=1):
                    medoides_filas.append({
                        "especificacion": nombre, "anio": anio, "k": k,
                        "conglomerado": i,
                        "medoide (actividad representativa)": d.loc[m, "desc"],
                        "codigo": d.loc[m, "codigo_3dig"],
                        "empleo del medoide": float(d.loc[m, "empleo"]),
                        "actividades en el grupo": int((cl == i).sum()),
                    })

                # --- PAM contra Ward, y Manhattan contra euclidiana
                cl_euclid, _, _ = particionar(d, k, "euclid", **opc)
                comparacion.append({
                    "especificacion": nombre, "anio": anio, "k": k,
                    "ARI PAM Manhattan vs Ward": uam.rand_ajustado(cl, ward(d, k, **opc)),
                    "ARI PAM Manhattan vs PAM euclid.": uam.rand_ajustado(cl, cl_euclid),
                    "ARI PAM euclid. vs Ward": uam.rand_ajustado(cl_euclid, ward(d, k, **opc)),
                })

            notas = [
                (QUE_HACE[nombre], "000000"),
                (f"PAM (k-medoides) con distancia de Manhattan, k = {k}. Sin logaritmos y sin "
                 "ponderar: la pregunta es que hace el algoritmo por si solo. Conglomerados "
                 "numerados de mayor a menor empleo.", "000000"),
                ("Universo: las que la UAM clasifica MAS el cobre, que va en rojo y cursiva. "
                 "Las demas exclusiones de la UAM se respetan.", pc.ROJO),
            ]
            pc.escribir_hoja(wb, f"{nombre} k={k}", notas, bs, con_marca=True, tablas=tablas)
            print(f"  {nombre:<22} k={k}  escrita")

    # ---------------------------------------------------------------- hojas
    hoja_tabla(
        wb, "01 Criterios por k",
        "Que k elegir: silueta, degeneracion y eta cuadrado",
        ["La silueta de Kaufman y Rousseeuw es la companera natural de PAM, pero PREMIA AISLAR "
         "UN EXTREMO: sube justo donde un conglomerado se lleva casi todo el empleo. Leerla "
         "junto a las dos columnas siguientes, no sola.",
         "«empleo del mayor» por encima de 90 % significa que la particion no informa. "
         "«grupos <1 %» cuenta los conglomerados gastados en actividades marginales.",
         "eta2: fraccion de la varianza del log de la productividad, entre actividades y "
         "ponderada por empleo, que queda ENTRE conglomerados. Es la medida sustantiva."],
        pd.DataFrame(criterios), (22, 8, 6, 26, 30, 16, 14, 14, 16, 16))

    hoja_tabla(
        wb, "02 El cobre",
        "Donde queda el cobre en cada combinacion",
        ["Es la pregunta que motiva la planilla. Con Ward y la especificacion de la UAM el cobre "
         "queda solo en su propio conglomerado, y por eso ellos lo excluyen.",
         "«solo? = SI» significa que el problema sigue. Un grupo de varias actividades con "
         "exportaciones altas y masa salarial baja es el enclave exportador del capitulo."],
        pd.DataFrame(cobre_filas), (22, 8, 6, 12, 12, 18, 12, 12, 14, 16, 56))

    hoja_tabla(
        wb, "03 PAM contra Ward",
        "Cuanto cambia la particion al cambiar de algoritmo y de metrica",
        ["ARI ajustado sobre las MISMAS variables y el MISMO k: lo unico que cambia es el "
         "metodo. 1,000 seria la misma particion.",
         "Las hojas de clasificacion usan PAM Manhattan. La euclidiana se corre igual, para "
         "separar el efecto del algoritmo (aglomerar contra particionar) del de la metrica "
         "(elevar al cuadrado o no)."],
        pd.DataFrame(comparacion), (22, 8, 6, 26, 30, 26))

    hoja_tabla(
        wb, "04 Medoides",
        "La actividad representativa de cada conglomerado",
        ["Es lo que PAM tiene y Ward no: el centro de cada grupo es una actividad REAL, no un "
         "punto promedio. Sirve para nombrar los estratos en el texto del capitulo.",
         "El medoide minimiza la suma de distancias al resto de su grupo: es el analogo "
         "multivariado de la mediana."],
        pd.DataFrame(medoides_filas), (22, 8, 6, 14, 56, 10, 18, 20))

    # --- Lectura
    ws = wb.create_sheet("0 Lectura", 0)
    for i, (texto, negrita, tam) in enumerate([
        ("T_I27 — PAM / k-medoides, con distintos k y con el comercio ajustado", True, 13),
        ("PAM es una alternativa COMPLETA a Ward, no una variante: no fusiona sino que reparte, "
         "el centro de cada grupo es una actividad real (el medoide) y no una media, la "
         "distancia de Manhattan no eleva al cuadrado, y la fase SWAP puede reasignar una "
         "actividad mal colocada. Lo que se pierde: no hay jerarquia ni dendrograma, y el k hay "
         "que fijarlo de antemano — por eso esta planilla recorre k = 2 a 6.", False, 9),
        ("Motivo: en 2003 la especificacion 5 de T_I26 (Ward + peso + logs) deja al cobre en un "
         "conglomerado de 48 actividades con el 44,8 % del empleo. El coste de Ward ponderado es "
         "(Wa*Wb/(Wa+Wb))*d^2, y con el cobre pesando 0,29 % del empleo ese factor lo domina la "
         "masa chica casi sin importar la distancia: el grupo del cobre se vuelve el destino mas "
         "barato y absorbe bloques enteros por aritmetica, no por afinidad.", False, 9),
        ("Universo: las 62 y 101 actividades que la UAM clasifica MAS el cobre (63 y 102), el "
         "mismo de T_I26, para poder comparar hoja contra hoja.", False, 9),
        ("Diez hojas de clasificacion: A (las siete variables de la UAM intactas) y B (comercio "
         "reparametrizado), por k = 2 a 6. Sin logaritmos y sin ponderar en las dos: la pregunta "
         "es que hace PAM por si solo.", False, 9),
        ("Diagnostico: 01 criterios por k · 02 el cobre · 03 PAM contra Ward · 04 medoides.",
         False, 9),
    ]):
        ws.cell(1 + i, 1, texto).font = Font(bold=negrita, italic=not negrita, size=tam)
    ws.column_dimensions["A"].width = 200

    wb.save(SALIDA)
    print(f"\nescrito: {SALIDA}\n")

    print("== Donde queda el cobre ==")
    print(pd.DataFrame(cobre_filas)[
        ["especificacion", "anio", "k", "actividades en el", "solo?", "% del empleo",
         "% de las expo", "masa salarial / VA (%)"]].round(2).to_string(index=False))
    print("\n== Criterios por k ==")
    print(pd.DataFrame(criterios)[
        ["especificacion", "anio", "k", "tamanos", "empleo del mayor (%)",
         "grupos <1% empleo", "silueta ponderada", "eta2 log productividad"]
    ].round(3).to_string(index=False))
    print("\n== PAM contra Ward ==")
    print(pd.DataFrame(comparacion).round(3).to_string(index=False))


if __name__ == "__main__":
    main()
